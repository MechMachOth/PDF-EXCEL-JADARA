# L'application est un outil puissant conçu pour faciliter la gestion des relevés bancaires de la banque "Umnia Bank".
# Elle offre aux utilisateurs deux options principales dans son menu.

# Option "Charger relevé bancaire PDF":
# Cette option permet à l'utilisateur de sélectionner un relevé bancaire au format PDF émanant de la banque "Umnia Bank".
# Une fois le fichier PDF chargé, l'application demande à l'utilisateur de spécifier l'année associée au relevé, car celle-ci n'est pas directement mentionnée dans le document.
# En utilisant l'année fournie par l'utilisateur, l'application extrait les données suivantes du relevé :
# Date (au format DD/MM/AA)
# Mois (en toutes lettres)
# Type de paiement
# Nom du donneur d'ordre (si disponible)
# Montant (en chiffres)
# Montant (en lettres)
# Numéro d'attestation unique pour chaque paiement
# De plus, l'application offre des fonctionnalités supplémentaires telles que la possibilité de supprimer, modifier et ajouter des paiements.
# Elle permet également d'éditer une seule attestation ou plusieurs attestations en fonction des paiements sélectionnés avant de cliquer sur le bouton "Attestation(s)".
# Pour faciliter la gestion des données, l'application permet également de télécharger un fichier Excel au format CSV en cliquant sur le bouton "Télécharger CSV",
# créant ainsi une copie des données chargées.

# Option "Créer attestation":
# Cette option permet à l'utilisateur de créer manuellement une attestation en fournissant les informations suivantes :
# Date
# Mois
# Type de paiement
# Montant (en chiffres)
# Chiffre final du numéro d'attestation unique
# En utilisant ces informations, l'application génère une attestation personnalisée.

# En résumé, cette application offre une solution complète pour importer des relevés bancaires au format PDF, extraire automatiquement les données pertinentes,
# effectuer des opérations de gestion sur ces données et générer des attestations personnalisées.
# Elle facilite ainsi la gestion financière en automatisant les tâches fastidieuses et en fournissant des fonctionnalités pratiques pour la manipulation des données bancaires.*

import sys
import os
import re
from fpdf import FPDF
from tkinter import messagebox
import csv
import xlsxwriter
import tabula
from numpy import NaN
from PIL import ImageTk, Image
from tkPDFViewer import tkPDFViewer as pdf
from tkinter import simpledialog, filedialog, ttk
from tkinter import *
import openpyxl
import xlrd
# ==========States=======
mdp = [('')]
année = 0
mois = ''
logged = 0
traité = 0
ope = 0
bordercolor = 3
bgcolor = 0
v2 = ""
v4 = ""
v6 = ""
RBUmnia = ''
RBAwb = ''
RBCmi = ''
selectionsRB = []
safichargi = 0
rje3 = 0
deb = 0
showpdf = 0
rech = 0
Dattta = []
ka = 0
reloulo = 0
kn9lb3la = ''
bach = 'att'
mselec = 0
mytag = ''
laDate = ''
changerowcolo = 0
télo = 0
Organ = 5 
NumAtt= 0 
tableAWB=[]
tableUmnia=[]
tableCMI=[]
bases = [("MILLIARD ", 1e9), ("MILLION ", 1e6), ("MILLE ", 1e3), ("CENT ", 1e2), ("QUATRE VINGT ", 80),
         ("SOIXANTE ", 60), ("CINQUANTE ", 50), ("QUARANTE ", 40), ("TRENTE ", 30), ("VINGT ", 20), ("DIX ", 10)]
units = ["ZERO ", "UN ", "DEUX ", "TROIS ", "QUATRE ", "CINQ ", "SIX ", "SEPT ",
         "HUIT ", "NEUF ", None, "ONZE ", "DOUZE ", "TREIZE ", "QUATORZE ", "QUINZE ", "SEIZE "]
# =======window========================
window = Tk()
icon = PhotoImage(file='logo-light.png')
window.tk.call('wm', 'iconphoto', window._w, icon)
style = ttk.Style(window)
style.theme_use("clam")

# ============Login==========

window.resizable(width=1, height=1)

checkVar1 = BooleanVar(value=False)
checkVar2 = BooleanVar(value=False)
checkVar3 = BooleanVar(value=False)
checkVar4 = BooleanVar(value=False)
checkVar5 = BooleanVar(value=False)
checkVar6 = BooleanVar(value=False)
checkVar7 = BooleanVar(value=False)
checkVar8 = BooleanVar(value=True)
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x_cordinate = int((screen_width/2) - 160)
y_cordinate = int((screen_height/2) - 140)
window.geometry("{}x{}+{}+{}".format(350, 200, x_cordinate, y_cordinate))
window.resizable(width=0, height=0)
window.resizable(width=0, height=0)
window.title("  Login")
window.config(background='black')
page_frame = Frame(window)
mdp_lb = Label(page_frame, text='Mot de passe',
               font=('Bold', 15), fg='white', bg='black')
mdp_lb.pack(pady=10)
mdp_entry = Entry(page_frame, font=('Bold', 15), bd=0, show='*')
mdp_entry.pack(pady=10)
mdp_entry.focus()


def show_passeword():
    if mdp_entry.cget('show') == '*':
        mdp_entry.config(show='')
        show_hide.config(bg='green')
    else:
        mdp_entry.config(show='*')
        show_hide.config(bg='red')


show_hide = Button(mdp_entry, bg='red', command=show_passeword)
show_hide.place(width=20, height=20, x=240, y=3)
login_btn = Button(page_frame, text='Connexion', font=(
    'Bold', 15), bd=0, bg='#158aff', fg='black', command=lambda: chek_mdp())
login_btn.pack()
error = Label(page_frame, text="", bg='black')
error.pack(pady=5)
page_frame.pack(pady=20)
page_frame.pack_propagate(False)
page_frame.configure(width=350, height=500, bg='black')
window.bind("<Return>", lambda e: chek_mdp())
window.bind("<Control-a>", lambda e: mdp_entry.select_range(0, 'end'))
mdp_entry.bind("<Escape>", lambda e: window.destroy())


def chek_mdp():
    global logged
    if mdp_entry.get() in mdp:
        if logged != 1:
            logged = 1
            error.config(text="Connexion...", bg='green')
            page_frame.after(500, lambda: page_frame.forget())
            window.after(500, lambda: acceuil())
        return
    else:
        error.config(text="try again", bg='red')
    return

# ============Acceuil==========


def acceuil():
    window.resizable(width=1, height=1)
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_cordinate = int((screen_width/2) - 160)
    y_cordinate = int((screen_height/2) - 140)
    window.geometry("{}x{}+{}+{}".format(350, 200, x_cordinate, y_cordinate))
    window.resizable(width=0, height=0)
    global rje3
    rje3 = 0
    global RBUmnia
    global RBAwb
    global RBCmi
    global selectionsRB
    global Dattta
    global télo
    global safichargi
    safichargi = 0
    télo = 0
    selectionsRB = []
    Dattta = []
    RBUmnia = ''
    page_frame.destroy()

    def secondaire():
        zeroun = '01'
        topsec = Toplevel()
        topsec.title("Attestation de don")
        topsec.geometry("500x400")
        topsec.resizable(width=0, height=0)
        icon = PhotoImage(file='logo-light.png')
        window.tk.call('wm', 'iconphoto', topsec._w, icon)
        l80 = Label(topsec, text="Date", width=20,
                    font=('Times', 11, 'bold'))
        e80 = Entry(
            topsec,  width=25)
        l80.place(x=50, y=70)
        e80.place(x=200, y=70)
        l89 = Label(topsec, text="Mois", width=20,
                    font=('Times', 11, 'bold'))
        e89 = Entry(
            topsec,  width=25)
        l89.place(x=50, y=110)
        e89.place(x=200, y=110)
        l81 = Label(topsec, text="Type", width=20,
                    font=('Times', 11, 'bold'))

        options = [
            "VIREMENT PERMANENT",
            "VIREMENT",
            "chèque".upper(),
            "CARTE BANCAIRE",
            "espèces".upper(),
        ]

        # datatype of menu text
        clicked = StringVar()

        # initial menu text
        clicked.set("VIREMENT")

        # Create Dropdown menu
        drop = OptionMenu(topsec, clicked, *options)
        drop.place(x=200, y=145)
        l81.place(x=50, y=150)
        l82 = Label(topsec, text="Nom donneur d'ordre",
                    width=20, font=('Times', 11, 'bold'))
        e82 = Entry(
            topsec, width=25)
        l82.place(x=50, y=190)
        e82.place(x=200, y=190)
        l83 = Label(topsec, text="Montant",
                    width=20, font=('Times', 11, 'bold'))
        e83 = Entry(
            topsec, width=25)
        l83.place(x=50, y=230)
        e83.place(x=200, y=230)
        l85 = Label(topsec, text="Montant en lettre",
                    width=20, font=('Times', 11, 'bold'))
        e85 = Entry(
            topsec,  width=25)
        l85.place(x=50, y=270)
        e85.place(x=200, y=270)

        l86 = Label(topsec, text="N° Attestation",
                    width=20, font=('Times', 11, 'bold'))
        l87 = Label(topsec, text="D-AA-MM-",
                    width=20, font=('Times', 11, 'bold'))
        e86 = Entry(
            topsec, width=10)
        l86.place(x=50, y=310)
        l87.place(x=170, y=315)
        e86.place(x=300, y=310)
        e80.focus()
        e89.configure(state="disabled")
        e85.configure(state="disabled")

        def idcode(e):
            global Dattta
            nonlocal options
            if e == "C":
                # Supprimez les options existantes du menu déroulant
                drop['menu'].delete(0, 'end')
                # Recréez le menu déroulant avec les options d'origine
                for option in options:
                    drop['menu'].add_command(label=option, command=lambda value=option: clicked.set(value))
                # initial menu text
                clicked.set("VIREMENT")

                C.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                P.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                S.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                D.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                e82.configure(state="normal")
            if e == "S":
                # Supprimez les options existantes du menu déroulant
                drop['menu'].delete(0, 'end')
                # Recréez le menu déroulant avec les options d'origine
                for option in options:
                    drop['menu'].add_command(label=option, command=lambda value=option: clicked.set(value))
                # initial menu text
                clicked.set("VIREMENT")
                S.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                P.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                C.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                D.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                e82.configure(state="normal")
            if e == "D":
                # Supprimez les options existantes du menu déroulant
                drop['menu'].delete(0, 'end')
                # Recréez le menu déroulant avec les options d'origine
                for option in options:
                    drop['menu'].add_command(label=option, command=lambda value=option: clicked.set(value))
                # initial menu text
                clicked.set("VIREMENT")
                D.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                C.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                S.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                P.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                e82.configure(state="normal")
            if e == "P":
                clicked.set("")
                drop['menu'].delete(0, 'end')
                P.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                C.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                S.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                D.configure(
                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                e82.configure(state="disabled")
            if D.cget('bg') == 'blue' and C.cget('bg') == 'blue' and P.cget('bg') == 'blue' and S.cget('bg') == 'blue':
                l87.configure(
                    text='D-'+e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5]+'-')
                topsec.title("Attestation de Don")
                if e80.get() == '':
                    l87.configure(text='D-AA-MM-')
            if D.cget('bg') == 'green':
                l87.configure(
                    text='D-'+e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5]+'-')
                topsec.title("Attestation de Don")
                if e80.get() == '':
                    l87.configure(text='D-AA-MM-')
            if C.cget('bg') == 'green':
                l87.configure(
                    text='C-'+e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5]+'-')
                topsec.title("Attestation de Cotisation")
                if e80.get() == '':
                    l87.configure(text='C-AA-MM-')
            if P.cget('bg') == 'green':
                l87.configure(
                    text='P-'+e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5]+'-')
                topsec.title("Attestation de Don en nature")
                if e80.get() == '':
                    l87.configure(text='P-AA-MM-')
            if S.cget('bg') == 'green':
                l87.configure(
                    text='S-'+e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5]+'-')
                topsec.title("Attestation de Sponsoring")
                if e80.get() == '':
                    l87.configure(text='S-AA-MM-')

        D = Button(
            topsec, text='D', command=lambda: idcode("D"))
        D.configure(
            font=('Times', 11, 'bold'), bg='blue', fg='white')
        D.place(x=200, y=20)
        S = Button(
            topsec, text='S', command=lambda: idcode("S"))
        S.configure(
            font=('Times', 11, 'bold'), bg='blue', fg='white')
        S.place(x=230, y=20)
        P = Button(
            topsec, text='P', command=lambda: idcode("P"))
        P.configure(
            font=('Times', 11, 'bold'), bg='blue', fg='white')
        P.place(x=260, y=20)
        C = Button(
            topsec, text='C', command=lambda: idcode("C"))
        C.configure(
            font=('Times', 11, 'bold'), bg='blue', fg='white')
        C.place(x=290, y=20)

        e80.bind("<KeyRelease>", idcode)

        def hantaTchouf():
            e80.delete(0, END)
            e82.delete(0, END)
            e83.delete(0, END)
            e85.delete(0, END)
            e86.delete(0, END)
            topsec.destroy()

        def attsec():
            attoto = 0

            Fait_le = e80.get().upper()
            heywdi = e86.get().upper()
            if heywdi == '1' or heywdi == '':
                heywdi = '01'
            if heywdi == '2':
                heywdi = '02'
            if heywdi == '3':
                heywdi = '03'
            if heywdi == '4':
                heywdi = '04'
            if heywdi == '5':
                heywdi = '05'
            if heywdi == '6':
                heywdi = '06'
            if heywdi == '7':
                heywdi = '07'
            if heywdi == '8':
                heywdi = '08'
            if heywdi == '9':
                heywdi = '09'
            
            D_22_09_10 = l87.cget("text")+heywdi
            Montant = e83.get().upper()
            Montant_en_lettre = e85.get().upper()
            Nom = e82.get().upper()
            Type = clicked.get().upper()
            mois = ''
            kok = 0

            # Créer le nom du dossier basé sur le nom du fichier Excel
            nom_dossier = os.path.splitext(e80.get()[6: len(e80.get())]+'-'+e80.get()[3:5])[0]

            # Créer le dossier s'il n'existe pas déjà
            if not os.path.exists(nom_dossier):
                os.makedirs(nom_dossier)
            if os.path.exists(nom_dossier+"/"+str(D_22_09_10)+".pdf"):
                attoto = 1

            def chiftolett(value, skip=-1):
                if int(value) < len(units) and units[int(value)]:
                    return [] if int(value) <= skip else [units[int(value)]]
                for name, v in bases:
                    if int(value) >= v:
                        return chiftolett(int(int(value)/v), 1 if v <= 1000 else -1) + [name] + chiftolett(int(int(value) % v), 0)

            if Montant == '' or Fait_le == '':
                messagebox.showinfo(
                    title='Erreur !!', message="la date et le montant sont obligatoire.", parent=topsec)
                ope = 0
                return

            if Montant.count(',') == 1:
                if len(Montant)-Montant.index(',') != 3:
                    messagebox.showinfo(
                        title='Erreur !!', message="Montant invalide, merci de saisir deux chiffres après la virgule.", parent=topsec)
                    return
                else:
                    for i in Montant:
                        if i == ',' or i == '.':
                            pass
                        elif not i.isnumeric():
                            messagebox.showinfo(
                                title='Erreur !!', message="Montant invalide", parent=topsec)
                            return
                    po = Montant[len(Montant)-2] + \
                        Montant[len(Montant)-1]
                    if po == '00':
                        Montant_en_lettre = "".join(chiftolett(
                            int(int(Montant.replace(',', '').replace('.', ''))/100)))
                    else:
                        Montant_en_lettre = "".join(chiftolett(
                            int(int(Montant.replace(',', '').replace('.', ''))/100)))
                        Montant_en_lettre += 'VIRGULE '
                        if po == "01" or po == "02" or po == "03" or po == "04" or po == "05" or po == "06" or po == "07" or po == "08" or po == "09":
                            Montant_en_lettre += "ZERO "
                        Montant_en_lettre += "".join(chiftolett(int(po)))
                x = ''
                Montant = Montant.replace('.', '')
                if len(Montant) > 4:
                    bima = Montant[len(Montant)-3: len(Montant)]
                    Montant = Montant[0:len(Montant)-3]
                    for i in range(len(Montant)-1, -1, -1):
                        if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21:
                            x = x+'.'
                            x = x+Montant[i]
                        else:
                            x = x+Montant[i]
                        kok = kok+1
                if kok != 0:
                    Montant = ''
                    for i in range(len(x)-1, -1, -1):
                        Montant = Montant+x[i]
                    Montant = Montant+bima
            else:
                x = ''
                Montant = Montant.replace('.', '')
                if len(Montant) > 4:
                    for i in range(len(Montant)-1, -1, -1):
                        if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21:
                            x = x+'.'
                            x = x+Montant[i]
                        else:
                            x = x+Montant[i]
                        kok = kok+1
                if kok != 0:
                    Montant = ''
                    for i in range(len(x)-1, -1, -1):
                        Montant = Montant+x[i]
                for i in Montant:
                    if i == '.':
                        pass
                    elif not i.isnumeric():
                        messagebox.showinfo(
                            title='Erreur !!', message="Montant invalide", parent=topsec)
                        return
                Montant_en_lettre = "".join(chiftolett(
                    int(int(Montant.replace(',', '').replace('.', '')))))
                Montant = "".join((Montant, ',00'))

            if Fait_le[0:2].isnumeric() and Fait_le[3:5].isnumeric() and Fait_le[6:8].isnumeric() and len(Fait_le) == 8:
                if int(Fait_le[3:5]) > 12 or int(Fait_le[0:2]) > 31:
                    messagebox.showinfo(
                        title='Erreur !!', message="* Le nombre de jour ne doit pas être superieur à 31. \n\n* Le nombre de mois ne doit pas être superieur à 12.", parent=topsec)
                    return
            else:
                messagebox.showinfo(
                    title='Erreur !!', message="La date doit être sous la forme suivante :\n\n            'JJxMMxAA'", parent=topsec)
                return
            hoho = list(Fait_le)
            hoho[2] = '/'
            hoho[5] = '/'
            Fait_le = ''.join(hoho)
            if Fait_le[3:5] == '01':
                mois = 'Janvier'
            if Fait_le[3:5] == '02':
                mois = 'Février'
            if Fait_le[3:5] == '03':
                mois = 'Mars'
            if Fait_le[3:5] == '04':
                mois = 'Avril'
            if Fait_le[3:5] == '05':
                mois = 'Mai'
            if Fait_le[3:5] == '06':
                mois = 'Juin'
            if Fait_le[3:5] == '07':
                mois = 'Juillet'
            if Fait_le[3:5] == '08':
                mois = 'Août'
            if Fait_le[3:5] == '09':
                mois = 'Septembre'
            if Fait_le[3:5] == '10':
                mois = 'Octobre'
            if Fait_le[3:5] == '11':
                mois = 'Novembre'
            if Fait_le[3:5] == '12':
                mois = 'Décembre'
            pdf = FPDF(orientation='P', format='A4')
            pdf.add_page()

            if D_22_09_10[0] == "D":
                hantaTchouf()
                pdf.set_xy(90, 65)
                pdf.set_font("times", size=21, style='BU')
                pdf.cell(txt='Attestation de Don',
                            w=30, align='C')
                pdf.image('logo-att.png', 80, 10, w=50, h=40)
                pdf.set_xy(91.5, 75)
                pdf.cell(txt=D_22_09_10, w=28, align='C')
                text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                text2 = Montant + \
                    " dirhams ("+Montant_en_lettre+" dirhams)"
                if Type.lower() == "espèce" or Type.lower() == "espece":
                    text3 = " en "
                else:
                    text3 = " par "
                text4 = Type+" "
                text5 = "de "
                text6 = Nom+"."
                text7 = "La contribution de "
                text8 = Nom+" "
                text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                text10 = Nom+" "
                text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                text14 = "Fait à Casablanca, le "
                text15 = Fait_le[0:2]+" "+mois + \
                    " "+"20"+Fait_le[6:len(Fait_le)]

                text16 = "Bouchra OUTAGHANI"

                pdf.set_auto_page_break("ON", margin=0.0)
                pdf.set_font("times", size=12)
                pdf.set_xy(20, 105)
                pdf.multi_cell(w=170, h=5, txt=text1+"**"+text2+"**"+text3+"**"+text4+"**"+text5+"**"+text6+"**"+"\n\n"+text7+"**"+text8+"**"+text9+"\n\n"+"**"+text10+"**"+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                align='L')

                pdf.set_font("times", size=11)
                pdf.set_xy(100, 200)
                pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                "\n\n"+"**"+text16+"**", markdown=True, align='R')
                pdf.set_xy(100, 215)
                pdf.multi_cell(
                    w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                pdf.set_font("times", size=9)
                pdf.set_xy(100, 220)
                pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                markdown=True, align='R')
                pdf.set_xy(100, 225)
                pdf.multi_cell(
                    w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                pdf.set_xy(100, 230)
                pdf.multi_cell(
                    w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                pdf.set_fill_color(193, 153, 9)
                pdf.set_xy(8, 275)
                pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                pdf.set_text_color(45, 82, 158)
                pdf.set_font("times", size=14, style="B")
                pdf.set_xy(8, 280)
                pdf.multi_cell(
                    w=0, h=5, txt="JADARA Foundation")

                pdf.set_text_color(193, 153, 9)
                pdf.set_font("times", size=7.5, style="")
                pdf.set_xy(8, 285)
                pdf.multi_cell(
                    w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                pdf.set_text_color(0, 0, 0)
                pdf.set_font("times", size=7.5, style="")
                pdf.set_xy(8, 289)
                pdf.multi_cell(
                    w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                pdf.set_font("times", size=8, style="")
                pdf.set_xy(107, 279)
                pdf.multi_cell(
                    w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                pdf.set_font("times", size=8, style="")
                pdf.set_xy(158, 283)
                pdf.multi_cell(
                    w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                pdf.set_xy(152, 275)
                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                pdf.set_xy(102, 275)
                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                nom_fichier_pdf = os.path.join(
                    nom_dossier, str(D_22_09_10) + ".pdf")

                pdf.output(nom_fichier_pdf)
            if D_22_09_10[0] == "S":
                hantaTchouf()
                def swbatts() : 
                    pdf.set_xy(90, 65)
                    pdf.set_font("times", size=21, style='BU')
                    pdf.cell(txt='Attestation de Don',
                            w=30, align='C')
                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                    pdf.set_xy(91.5, 75)
                    pdf.cell(txt=D_22_09_10, w=28, align='C')
                    nameevent = e808.get()
                    text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                    text2 = "**"+Montant + \
                        " dirhams ("+Montant_en_lettre+" dirhams)"+"**"
                    if Type.lower() == "espèce" or Type.lower() == "espece":
                        text3 = " en "
                    else:
                        text3 = " par "
                    text4 = "**"+Type+" "+"**"
                    text5 = "de"
                    text6 = "**"+" "+Nom+"**"+"."
                    text7 = "La contribution de "
                    text8 = "**"+Nom+" "+"**"
                    text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                    textX = "Cette contribution participera au financement de l'évènement :"
                    textY = "**"+" "+nameevent+"**"
                    text10 = "**"+" "+Nom+" "+"**"
                    text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                    text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                    text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                    text14 = "Fait à Casablanca, le "
                    text15 = Fait_le[0:2]+" "+mois + \
                        " "+"20"+Fait_le[6:len(Fait_le)]

                    text16 = "Bouchra OUTAGHANI"

                    pdf.set_auto_page_break("ON", margin=0.0)
                    pdf.set_font("times", size=12)
                    pdf.set_xy(20, 105)
                    pdf.multi_cell(w=170, h=5, txt=text1+text2+text3+text4+text5+text6+"\n\n"+text7+text8+text9+"\n\n"+textX+textY+"\n\n"+text10+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                align='L')

                    pdf.set_font("times", size=11)
                    pdf.set_xy(100, 200)
                    pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                "\n\n"+"**"+text16+"**", markdown=True, align='R')
                    pdf.set_xy(100, 215)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                    pdf.set_font("times", size=9)
                    pdf.set_xy(100, 220)
                    pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                markdown=True, align='R')
                    pdf.set_xy(100, 225)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                    pdf.set_xy(100, 230)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                    pdf.set_fill_color(193, 153, 9)
                    pdf.set_xy(8, 275)
                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                    pdf.set_text_color(45, 82, 158)
                    pdf.set_font("times", size=14, style="B")
                    pdf.set_xy(8, 280)
                    pdf.multi_cell(
                        w=0, h=5, txt="JADARA Foundation")

                    pdf.set_text_color(193, 153, 9)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 285)
                    pdf.multi_cell(
                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 289)
                    pdf.multi_cell(
                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(107, 279)
                    pdf.multi_cell(
                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(158, 283)
                    pdf.multi_cell(
                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                    pdf.set_xy(152, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                    pdf.set_xy(102, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                    e808.delete(0, END)
                    topatts.destroy()

                    nom_fichier_pdf = os.path.join(
                        nom_dossier, str(D_22_09_10) + ".pdf")

                    pdf.output(nom_fichier_pdf)
                topatts = Toplevel()
                topatts.title(D_22_09_10)
                topatts.geometry("500x200")
                topatts.resizable(width=0, height=0)
                icon = PhotoImage(file='logo-light.png')
                window.tk.call('wm', 'iconphoto', topatts._w, icon)
                l800 = Label(topatts, text=Type+" ("+Montant+" DH) de "+Nom,
                            font=('Times', 11, 'bold'))
                l800.place(x=50, y=25)
                l808 = Label(topatts, text="Cette contribution participera au financement de l'évènement :",
                            font=('Times', 11, 'bold'))
                e808 = Entry(
                    topatts,  width=25)
                l808.place(x=50, y=72)
                e808.place(x=150, y=95)
                def hantaTchoufatts():
                    nonlocal attoto
                    attoto = 2
                    e808.delete(0, END)
                    topatts.destroy()
                submitbuttonatts = Button(
                    topatts, text='Enregistrer', command=lambda: swbatts())
                submitbuttonatts.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                submitbuttonatts.place(x=250, y=150)
                cancelbuttonatts = Button(
                    topatts, text='Annulé', command=lambda: hantaTchoufatts())
                cancelbuttonatts.configure(
                    font=('Times', 11, 'bold'), bg='red', fg='white')
                cancelbuttonatts.place(x=150, y=150)
                topatts.protocol(
                    "WM_DELETE_WINDOW", hantaTchoufatts)
                topatts.bind(
                    "<Return>", lambda e: swbatts())
                topatts.bind(
                    "<Escape>", lambda e: hantaTchoufatts())
                topatts.wait_window()
            if D_22_09_10[0] == "P":
                hantaTchouf()
                def swbattp():
                    pdf.set_xy(90, 65)
                    pdf.set_font("times", size=21, style='BU')
                    pdf.cell(txt='Attestation de Don en nature',
                            w=30, align='C')
                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                    pdf.set_xy(91.5, 75)
                    pdf.cell(txt=D_22_09_10+"-"+e802.get()+"DN", w=30, align='C')
                    text10 = "Je soussignée, Mme Bochra CHABBOUBA ELIDRISSI, Responsable Administrative et Financière de JADARA FOUBDATION, atteste par la présente que l'association a bénéficié au titre de l'année scolaire"
                    text11 = "**"+" "+e801.get()+"**"
                    text12 = " d'un don en nature de la part de :"
                    text13 = "**"+" "+e808.get()+"**"
                    text14 = "Ce don est sous forme d'une place pédagogique gracieusement offerte au profit du boursier inscrit régulièrement au titre de l'année universitaire"
                    text15 = "**"+"Nom : "+e803.get()+"**"
                    text16 = "**"+"CIN : "+e804.get()+"**"
                    text17 = "**"+"Etudiant en : "+e805.get()+"**"
                    text18 = "Ce don est valorisé dans les livres comptables de notre association au titre de l'exercice"+"**"+" "+e802.get()+"**"
                    text19 ="**"+ Montant + \
                        " dirhams ("+Montant_en_lettre+" dirhams)."+"**"
                    text199="Cette contribution participe au financement de la mission de JADARA FOUNDATION dont l'objet est de financer des bourses d'études supérieures pour les bacheliers méritants issus de milieux défavorisés."
                    text20 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                    text21 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                    text22 = "Fait à Casablanca, le "
                    text23 = Fait_le[0:2]+" "+mois + \
                        " "+"20"+Fait_le[6:len(Fait_le)]


                    pdf.set_auto_page_break("ON", margin=0.0)
                    pdf.set_font("times", size=12)
                    pdf.set_xy(20, 100)
                    pdf.multi_cell(w=170, h=5, txt=text10+text11+text12+"\n\n"+"                                               "+text13+"\n\n"+text14+text11+" :"+"\n\n"+"       "+text15+"\n\n"+"       "+text16+"\n\n"+"       "+text17+"\n\n"+text18+" à hauteur de "+text19+"\n\n"+text199+"\n\n"+text20+"\n\n"+text21, markdown=True,
                                align='L')

                    pdf.set_font("times", size=11)
                    pdf.set_xy(100, 240)
                    pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                    pdf.set_xy(100, 250)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                    pdf.set_xy(100, 255)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                    pdf.set_fill_color(193, 153, 9)
                    pdf.set_xy(8, 275)
                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                    pdf.set_text_color(45, 82, 158)
                    pdf.set_font("times", size=14, style="B")
                    pdf.set_xy(8, 280)
                    pdf.multi_cell(
                        w=0, h=5, txt="JADARA Foundation")

                    pdf.set_text_color(193, 153, 9)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 285)
                    pdf.multi_cell(
                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 289)
                    pdf.multi_cell(
                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(107, 279)
                    pdf.multi_cell(
                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(158, 283)
                    pdf.multi_cell(
                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                    pdf.set_xy(152, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                    pdf.set_xy(102, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                    e808.delete(0, END)
                    e801.delete(0, END)
                    e802.delete(0, END)
                    e803.delete(0, END)
                    e804.delete(0, END)
                    e805.delete(0, END)
                    topattp.destroy()

                    nom_fichier_pdf = os.path.join(
                        nom_dossier, str(D_22_09_10) + ".pdf")
                    pdf.output(nom_fichier_pdf)
                topattp = Toplevel()
                topattp.title(D_22_09_10)
                topattp.geometry("500x360")
                topattp.resizable(width=0, height=0)
                icon = PhotoImage(file='logo-light.png')
                window.tk.call('wm', 'iconphoto', topattp._w, icon)
                l800 = Label(topattp, text=Type+" ("+Montant+" DH) de "+Nom,
                            font=('Times', 11, 'bold'))
                l800.place(x=50, y=25)
                l808 = Label(topattp, text="De la part de : ",
                            font=('Times', 11, 'bold'))
                e808 = Entry(
                    topattp,  width=25)
                l808.place(x=50, y=72)
                e808.place(x=200, y=70)
                l802 = Label(topattp, text="Exercice : ",
                            font=('Times', 11, 'bold'))
                e802 = Entry(
                    topattp,  width=25)
                l802.place(x=50, y=112)
                e802.place(x=200, y=110)
                l801 = Label(topattp, text="Année scolaire : ",
                            font=('Times', 11, 'bold'))
                e801 = Entry(
                    topattp,  width=25)
                l801.place(x=50, y=152)
                e801.place(x=200, y=150)
                l803 = Label(topattp, text="Nom étudiant : ",
                            font=('Times', 11, 'bold'))
                e803 = Entry(
                    topattp,  width=25)
                l803.place(x=50, y=192)
                e803.place(x=200, y=190)
                l804 = Label(topattp, text="CIN étudiant : ",
                            font=('Times', 11, 'bold'))
                e804 = Entry(
                    topattp,  width=25)
                l804.place(x=50, y=232)
                e804.place(x=200, y=230)
                l805 = Label(topattp, text="Etudiant en : ",
                            font=('Times', 11, 'bold'))
                e805 = Entry(
                    topattp,  width=25)
                l805.place(x=50, y=272)
                e805.place(x=200, y=270)
                def hantaTchoufattp():
                    nonlocal attoto
                    attoto = 2
                    e808.delete(0, END)
                    e801.delete(0, END)
                    e802.delete(0, END)
                    e803.delete(0, END)
                    e804.delete(0, END)
                    e805.delete(0, END)
                    topattp.destroy()
                submitbuttonattp = Button(
                    topattp, text='Enregistrer', command=lambda: swbattp())
                submitbuttonattp.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                submitbuttonattp.place(x=300, y=320)
                cancelbuttonattp = Button(
                    topattp, text='Annulé', command=lambda: hantaTchoufattp())
                cancelbuttonattp.configure(
                    font=('Times', 11, 'bold'), bg='red', fg='white')
                cancelbuttonattp.place(x=200, y=320)
                topattp.protocol(
                    "WM_DELETE_WINDOW", hantaTchoufattp)
                topattp.bind(
                    "<Return>", lambda e: swbattp())
                topattp.bind(
                    "<Escape>", lambda e: hantaTchoufattp())
                topattp.wait_window()
            if D_22_09_10[0] == "C":
                hantaTchouf()
                def swbattc() : 
                    pdf.set_xy(90, 65)
                    pdf.set_font("times", size=21, style='BU')
                    pdf.cell(txt='Attestation de Cotisation',
                            w=30, align='C')
                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                    pdf.set_xy(91.5, 75)
                    pdf.cell(txt=D_22_09_10, w=28, align='C')
                    text1 = "Nous, JADARA FOUNDATION, attestons par la présente avoir reçu la somme de "
                    text2 = "**"+Montant + \
                        " dirhams ("+Montant_en_lettre+" dirhams) "+"**"
                    text5 = "de "
                    text6 ="**"+Nom+" "+"**"
                    text7 = labelc.cget('text')
                    text8 = "**"+" "+e908.get()+"**"+"."
                    text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                    text22 = "Fait à Casablanca, le "
                    text23 = Fait_le[0:2]+" "+mois + \
                        " "+"20"+Fait_le[6:len(Fait_le)]

                    pdf.set_auto_page_break("ON", margin=0.0)
                    pdf.set_font("times", size=13)
                    pdf.set_xy(20, 120)
                    pdf.multi_cell(w=170, h=5, txt=text1+text2+text5+text6+text7+text8+"\n\n"+text13, markdown=True,
                                align='L')
                    pdf.set_font("times", size=11)
                    pdf.set_xy(100, 200)
                    pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                    pdf.set_xy(100, 210)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                    pdf.set_xy(100, 215)
                    pdf.multi_cell(
                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                    pdf.set_fill_color(193, 153, 9)
                    pdf.set_xy(8, 275)
                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                    pdf.set_text_color(45, 82, 158)
                    pdf.set_font("times", size=14, style="B")
                    pdf.set_xy(8, 280)
                    pdf.multi_cell(
                        w=0, h=5, txt="JADARA Foundation")

                    pdf.set_text_color(193, 153, 9)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 285)
                    pdf.multi_cell(
                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font("times", size=7.5, style="")
                    pdf.set_xy(8, 289)
                    pdf.multi_cell(
                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(107, 279)
                    pdf.multi_cell(
                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                    pdf.set_font("times", size=8, style="")
                    pdf.set_xy(158, 283)
                    pdf.multi_cell(
                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                    pdf.set_xy(152, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                    pdf.set_xy(102, 275)
                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)
                    
                    e908.delete(0, END)
                    topattc.destroy()
                    nom_fichier_pdf = os.path.join(
                        nom_dossier, str(D_22_09_10) + ".pdf")

                    pdf.output(nom_fichier_pdf)

                topattc = Toplevel()
                topattc.title(D_22_09_10)
                topattc.geometry("500x220")
                topattc.resizable(width=0, height=0)
                icon = PhotoImage(file='logo-light.png')
                window.tk.call('wm', 'iconphoto', topattc._w, icon)
                l900 = Label(topattc, text=Type+" ("+Montant+" DH) de "+Nom,
                            font=('Times', 11, 'bold'))
                l900.place(x=50, y=25)
                def update_label():
                    if var.get() == 1:
                        labelc.config(text="en tant que membre de l'association au titre de l'année")
                    else:
                        labelc.config(text="en tant que membre de l'association au titre des années :")

                var = IntVar()
                var.set(1)  # Coche le premier checkbutton par défaut

                checkbuttonc1 = Checkbutton(topattc, text="Une année", variable=var, onvalue=1, offvalue=0, command=update_label)
                checkbuttonc1.place(x=90, y=60)

                checkbuttonc2 = Checkbutton(topattc, text="Plusieurs années", variable=var, onvalue=0, offvalue=1, command=update_label)
                checkbuttonc2.place(x=240, y=60)

                labelc = Label(topattc, text="en tant que membre de l'association au titre de l'année")
                labelc.place(x=50, y=90)

                e908 = Entry(topattc, width=25)
                e908.pack()
                e908.place(x=150, y=120)
                def hantaTchoufattc():
                    nonlocal attoto
                    attoto = 2
                    e908.delete(0, END)
                    topattc.destroy()
                submitbuttonattc = Button(
                    topattc, text='Enregistrer', command=lambda: swbattc())
                submitbuttonattc.configure(
                    font=('Times', 11, 'bold'), bg='green', fg='white')
                submitbuttonattc.place(x=250, y=170)
                cancelbuttonattc = Button(
                    topattc, text='Annulé', command=lambda: hantaTchoufattc())
                cancelbuttonattc.configure(
                    font=('Times', 11, 'bold'), bg='red', fg='white')
                cancelbuttonattc.place(x=150, y=170)
                topattc.protocol(
                    "WM_DELETE_WINDOW", hantaTchoufattc)
                topattc.bind(
                    "<Return>", lambda e: swbattc())
                topattc.bind(
                    "<Escape>", lambda e: hantaTchoufattc())
                topattc.wait_window()
            if attoto == 0:
                messagebox.showinfo(
                    title='', message="Fichier "+str(D_22_09_10)+".pdf créé. ", parent=window)
            elif attoto == 1:
                messagebox.showinfo(
                    title='', message="Fichier "+str(D_22_09_10)+".pdf mis à jour. ", parent=window)

        submitbutton = Button(
            topsec, text='Enregistrer', command=lambda: attsec())
        submitbutton.configure(
            font=('Times', 11, 'bold'), bg='green', fg='white')
        submitbutton.place(x=300, y=350)
        cancelbutton = Button(
            topsec, text='Annulé', command=lambda: hantaTchouf())
        cancelbutton.configure(
            font=('Times', 11, 'bold'), bg='red', fg='white')
        cancelbutton.place(x=200, y=350)
        topsec.protocol(
            "WM_DELETE_WINDOW", hantaTchouf)
        topsec.bind(
            "<Return>", lambda e: attsec())
        topsec.bind(
            "<Escape>", lambda e: hantaTchouf())
        return

    def principal():
        global rje3
        global v2
        global v1
        global RBUmnia
        global RBAwb
        global RBCmi
        global selectionsRB
        global bordercolor
        global bgcolor
        global safichargi
        rje3 = 1
        b.place_forget()
        a.place_forget()
        lo13.place_forget()
        lo14.place_forget()
        window.resizable(width=1, height=1)
        window.geometry("5000x5000")

        def reset():
            sur = messagebox.askquestion(
                'les données seront perdues !', "êtes-vous sûr de vouloir relancer l'application ?", icon='warning')
            if sur == 'yes':
                python = sys.executable
                os.execl(python, python, * sys.argv)
        frame2 = Frame(window)
        frame1 = Frame(window)
        frame4 = Frame(frame1)
        frame3 = Frame(frame1)
        frame5 = Frame(frame1, height=350, width=45, bg='black')
        frame4.pack(padx=0, pady=5)
        frame5.place(x=5, y=202)
        thblack = Button(frame4, text='',
                         command=lambda: borco())
        thblack.pack(pady=5, padx=30, side='left')
        thblackbg = Button(frame4, text='',
                           command=lambda: bgco())
        thblackbg.pack(pady=5, padx=30, side='left')

        if bordercolor == 2:
            frame1.configure(bg='black',
                             highlightbackground="yellow", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="yellow", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="yellow", highlightthickness=2)
            frame4.configure(bg='black',
                             highlightbackground="yellow", highlightthickness=2)
            thblack.configure(background='#000080')

        elif bordercolor == 1:
            frame4.configure(bg='black',
                             highlightbackground="red", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="red", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="red", highlightthickness=2)
            frame1.configure(bg='black',
                             highlightbackground="red", highlightthickness=2)
            thblack.configure(background='yellow')

        elif bordercolor == 3:

            frame1.configure(bg='black',
                             highlightbackground="#000080", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="#000080", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="#000080", highlightthickness=2)
            frame4.configure(bg='black',
                             highlightbackground="#000080", highlightthickness=2)
            thblack.configure(background='black')

        elif bordercolor == 0:
            frame4.configure(bg='black',
                             highlightbackground="green", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="green", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="green", highlightthickness=2)
            frame1.configure(bg='black',
                             highlightbackground="green", highlightthickness=2)
            thblack.configure(background='red')

        elif bordercolor == 5:
            frame1.configure(bg='black',
                             highlightbackground="white", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="white", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="white", highlightthickness=2)
            frame4.configure(bg='black',
                             highlightbackground="white", highlightthickness=2)
            thblack.configure(background='green')

        elif bordercolor == 4:
            frame4.configure(bg='black',
                             highlightbackground="black", highlightthickness=2)
            frame3.configure(bg='black',
                             highlightbackground="black", highlightthickness=2)
            frame2.configure(bg='black',
                             highlightbackground="black", highlightthickness=2)
            frame1.configure(bg='black',
                             highlightbackground="black", highlightthickness=2)
            thblack.configure(background='white')

        def bgco():
            global bgcolor
            if bgcolor == 0:
                bgcolor = 1
                window.configure(background='white')
                frame1.configure(bg='white')
                frame2.configure(bg='white')
                frame3.configure(bg='white')
                frame4.configure(bg='white')
                frame5.configure(bg='white')
                thblackbg.configure(bg='green')
            elif bgcolor == 1:
                bgcolor = 2
                window.configure(background='green')
                frame1.configure(bg='green')
                frame2.configure(bg='green')
                frame5.configure(bg='green')
                frame3.configure(bg='green')
                thblackbg.configure(bg='yellow')
                frame4.configure(bg='green')
            elif bgcolor == 2:
                bgcolor = 3
                window.configure(background='yellow')
                frame1.configure(bg='yellow')
                frame2.configure(bg='yellow')
                frame3.configure(bg='yellow')
                frame5.configure(bg='yellow')
                thblackbg.configure(bg='#000080')
                frame4.configure(bg='yellow')
            elif bgcolor == 3:
                bgcolor = 4
                window.configure(background='#000080')
                frame1.configure(bg='#000080')
                frame5.configure(bg='#000080')
                frame2.configure(bg='#000080')
                frame3.configure(bg='#000080')
                thblackbg.configure(bg='red')
                frame4.configure(bg='#000080')
            elif bgcolor == 4:
                bgcolor = 5
                window.configure(background='red')
                frame1.configure(bg='red')
                frame2.configure(bg='red')
                frame5.configure(bg='red')
                frame3.configure(bg='red')
                thblackbg.configure(bg='black')
                frame4.configure(bg='red')
            elif bgcolor == 5:
                bgcolor = 0
                window.configure(background='black')
                frame1.configure(bg='black')
                frame2.configure(bg='black')
                frame3.configure(bg='black')
                frame5.configure(bg='black')
                frame4.configure(bg='black')
                thblackbg.configure(bg='white')

            return

        if bgcolor == 0:
            window.configure(background='black')
            frame1.configure(bg='black')
            frame2.configure(bg='black')
            frame3.configure(bg='black')
            frame5.configure(bg='black')
            frame4.configure(bg='black')
            thblackbg.configure(background='white')

        if bgcolor == 1:
            window.configure(background='white')
            frame1.configure(bg='white')
            frame2.configure(bg='white')
            frame3.configure(bg='white')
            frame5.configure(bg='white')
            frame4.configure(bg='white')
            thblackbg.configure(background='green')

        if bgcolor == 2:
            window.configure(background='green')
            frame1.configure(bg='green')
            frame2.configure(bg='green')
            frame3.configure(bg='green')
            frame5.configure(bg='green')
            frame4.configure(bg='green')
            thblackbg.configure(background='yellow')

        if bgcolor == 3:
            window.configure(background='yellow')
            frame1.configure(bg='yellow')
            frame2.configure(bg='yellow')
            frame3.configure(bg='yellow')
            frame5.configure(bg='yellow')
            frame4.configure(bg='yellow')
            thblackbg.configure(background='#000080')

        if bgcolor == 4:
            window.configure(background='#000080')
            frame1.configure(bg='#000080')
            frame2.configure(bg='#000080')
            frame3.configure(bg='#000080')
            frame5.configure(bg='#000080')
            frame4.configure(bg='#000080')
            thblackbg.configure(background='red')

        if bgcolor == 5:
            window.configure(background='red')
            frame1.configure(bg='red')
            frame2.configure(bg='red')
            frame3.configure(bg='red')
            frame5.configure(bg='red')
            frame4.configure(bg='red')
            thblackbg.configure(background='black')

        def borco():
            global bordercolor
            if bordercolor == 0:
                frame1.configure(
                    highlightbackground="red", highlightthickness=2)
                frame2.configure(
                    highlightbackground="red", highlightthickness=2)
                frame3.configure(
                    highlightbackground="red", highlightthickness=2)
                frame4.configure(
                    highlightbackground="red", highlightthickness=2)
                thblack.configure(background='yellow')
                bordercolor = 1
            elif bordercolor == 2:
                bordercolor = 3
                frame1.configure(
                    highlightbackground="#000080", highlightthickness=2)
                frame2.configure(
                    highlightbackground="#000080", highlightthickness=2)
                frame4.configure(
                    highlightbackground="#000080", highlightthickness=2)
                frame3.configure(
                    highlightbackground="#000080", highlightthickness=2)
                thblack.configure(background='black')
            elif bordercolor == 1:
                bordercolor = 2
                frame1.configure(
                    highlightbackground="yellow", highlightthickness=2)
                frame2.configure(
                    highlightbackground="yellow", highlightthickness=2)
                frame3.configure(
                    highlightbackground="yellow", highlightthickness=2)
                frame4.configure(
                    highlightbackground="yellow", highlightthickness=2)
                thblack.configure(background='#000080')
            elif bordercolor == 3:
                bordercolor = 4
                frame1.configure(
                    highlightbackground="black", highlightthickness=2)
                frame2.configure(
                    highlightbackground="black", highlightthickness=2)
                frame3.configure(
                    highlightbackground="black", highlightthickness=2)
                frame4.configure(
                    highlightbackground="black", highlightthickness=2)
                thblack.configure(background='white')
            elif bordercolor == 4:
                bordercolor = 5
                frame1.configure(
                    highlightbackground="white", highlightthickness=2)
                frame2.configure(
                    highlightbackground="white", highlightthickness=2)
                frame3.configure(
                    highlightbackground="white", highlightthickness=2)
                frame4.configure(
                    highlightbackground="white", highlightthickness=2)
                thblack.configure(background='green')
            elif bordercolor == 5:
                bordercolor = 0
                frame1.configure(
                    highlightbackground="green", highlightthickness=2)
                frame2.configure(
                    highlightbackground="green", highlightthickness=2)
                frame3.configure(
                    highlightbackground="green", highlightthickness=2)
                frame4.configure(
                    highlightbackground="green", highlightthickness=2)
                thblack.configure(background='red')

        # ==== PDFViewer=====

        def Umnia():
            global v2
            global RBUmnia
            RBUmnia = filedialog.askopenfilename(initialdir=os.getcwd(),
                                                title='Select Umnia pdf file',
                                                filetypes=(("PDF File", ".pdf"), ("PDF File", ".PDF"), ("All file", ".txt")))
            if type(RBUmnia) == tuple or RBUmnia == '':
                acceuil()
                return
            
            v1 = pdf.ShowPdf()
            v2 = v1.pdf_view(frame2, pdf_location=open(RBUmnia, 'r'), width=75)
            v2.pack(fill='y', side='left', expand=True)

        def AWB():
            global v2
            global RBAwb
            RBAwb = filedialog.askopenfilename(initialdir=os.getcwd(),
                                            title='Select AWB pdf file',
                                            filetypes=(("PDF File", ".pdf"), ("PDF File", ".PDF"), ("All file", ".txt")))
            if type(RBAwb) == tuple or RBAwb == '':
                acceuil()
                return 
            if v2 : 
                v2.pack_forget()
            v1 = pdf.ShowPdf()
            v2 = v1.pdf_view(frame2, pdf_location=open(RBAwb, 'r'), width=75)
            v2.pack(fill='y', side='right', expand=True)


        def Cmi():
            global v6
            global RBCmi
            RBCmi = filedialog.askopenfilename(initialdir=os.getcwd(),
                                                title='Select excel CMI file',
                                                filetypes=(("Fichiers Excel CMI", "*.xls"),("Fichiers Excel CMI", "*.xlsx"), ("All file", ".txt")))
            if type(RBCmi) == tuple or RBCmi == '':
                acceuil()
                return

        def afficher_selection():
            global RBUmnia
            global RBAwb
            global RBCmi
            global selectionsRB
            global v2
            global v4
            global v6
            global charger
            global lo20
            global traité
            for i, var in enumerate(checkbox_vars):
                if var.get() == 1:
                    selectionsRB.append(banques[i])
            topRBMENU.destroy()
            if selectionsRB == []:
                frame1.destroy()
                frame2.destroy()
                frame3.destroy()
                traité = 0
                acceuil()
            b.place_forget()
            a.place_forget()
            for element in selectionsRB:
                if element == "UmniaBank":
                    Umnia()
                if element == "AWB":
                    AWB()
                if element == "CMI":
                    Cmi()
            charger = Button(frame1, text='Charger les données',
                             command=traitement)
            charger.pack(pady=20)
            lo20 = Label(frame1, text='Entrée', fg='white',
                         background='black', font=('Times', 10))
            lo20.pack()

        topRBMENU = Toplevel(window)
        window_width = window.winfo_reqwidth()
        window_height = window.winfo_reqheight()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = int((screen_width / 2) - (window_width / 2))
        y = int((screen_height / 2) - (window_height / 2))
        topRBMENU.geometry(f"+{x}+{y}")
        topRBMENU.overrideredirect(True)

        banques = ["UmniaBank", "AWB", "CMI"]
        checkbox_vars = []

        for i, banque in enumerate(banques):
            var = IntVar()
            checkbox = Checkbutton(topRBMENU, text=banque, variable=var)
            checkbox.pack(anchor=W)
            checkbox_vars.append(var)

        btn_valider = Button(topRBMENU, text="Valider",
                             command=afficher_selection)
        btn_valider.pack()

        def traitement():
            global traité
            global Dattta
            global laDate
            global v2
            global RBUmnia
            global RBAwb
            global RBCmi
            global selectionsRB
            global NumAtt

            charger.config(bg='gray')
            if traité == 0:
                Dattta = []
                global année
                charger.config(text='Chargement...')
                année = simpledialog.askinteger(
                    "Année", "Merci de Saisir l'année", initialvalue=2023, parent=None, maxvalue=9999, minvalue=1000)
                traité = 1
                if année == None:
                    charger.config(text='Charger les données')
                    traité = 0
                    return
                year = année-2000
                charger.pack_forget()
                lo20.pack_forget()
                try : 
                    v2.pack_forget()
                except : 
                    donothing = 0
                try : 
                    if RBUmnia != '' or type(RBUmnia) == tuple :
                        tables = tabula.read_pdf(RBUmnia, pages='all')
                        for i in range(len(tables)):
                            csv_file = os.path.join(os.getcwd(), f"table{i}.csv")
                            tables[i].to_csv(csv_file)
                        for j in range(len(tables)):
                            csv_file = os.path.join(os.getcwd(), f"table{j}.csv")
                            with open(csv_file, 'r') as file:
                                reader = csv.reader(file)
                                for each in reader:
                                    if (each[1] != ''):
                                        if (each[len(each)-1] != ''):
                                            if each.count("Unnamed: 2") == 0:
                                                if each[4] == '':
                                                    if each.count("Page N°") == 0:
                                                        for i in each:
                                                            if i == '':
                                                                each.remove(i)
                                                                for i in each:
                                                                    if i == '':
                                                                        each.remove(i)
                                                        tableUmnia.append(each)

                        i = 1
                        for each in tableUmnia:
                            type = each[2][0:4]
                            if type != 'CRED':
                                each[1] = each[1][0:5]+'/'+str(year)
                                each[0] = 'D-' + str(year) + '-'+each[1][3:5]+'-'+str(i)
                                i = i+1
                            if type == 'VIRE':
                                if each[2].count('DE') == 2:
                                    each.insert(3, each[2][29:])
                                if each[2].count('DE') < 2:
                                    each.insert(3, each[2][24:])
                                each[2] = 'VIREMENT PERMANENT'
                            if type == 'VIR ':
                                each.insert(3, each[2][26:])
                                each[2] = 'VIREMENT'
                            if type == 'VERS':
                                each.insert(3, each[2][13:])
                                each[2] = 'VERSEMENT'
                            if type == 'ENCA':
                                each.insert(3, "")
                            if type == 'OBJE':
                                each.insert(3, "")
                        tableUmnia.pop(len(tableUmnia)-1)
                        for j in range(len(tables)):
                            os.remove(f"table{j}.csv")
                except : 
                        donothing=0
                
                try : 
                    if RBAwb != '' or type(RBAwb) == tuple :
                        algo = 1
                        tables2= tabula.read_pdf(RBAwb, pages='all')
                        for i in range(len(tables2)):
                            csv_file = os.path.join(os.getcwd(), f"table{i}.csv")
                            tables2[i].to_csv(csv_file)
                            # print(tables2[i])
                        if tableUmnia != [] :
                            iidd = int(tableUmnia[len(tableUmnia)-1][0][8:])+1
                        else : 
                            iidd = 1
                        for j in range(len(tables2)):
                            csv_file = os.path.join(os.getcwd(), f"table{j}.csv")
                            with open(csv_file, 'r') as file : 
                                reader = csv.reader(file) 
                                for i, each in enumerate(reader): 
                                    print(each)
                                    if len(each) == 4 : 
                                        algo= 1
                                    elif len(each) == 5 : 
                                        algo= 2
                                    elif len(each) == 6 : 
                                        algo= 3
                                    if algo == 1  : 
                                        if " RECU " in each[1] or " RE CU " in each[1] or " A ENC " in each[1] or "VERSEMENT" in each[1] or "V ER S EM ENT" in each[1] or "RECU " in each[1] or " RECU" in each[1]  : 
                                            print('ressu',each)
                                            if each[2] == '' or each[2] == 'NaN' : 
                                                if any(c.isalpha() for c in each[1][-10:]):
                                                    ddaattee = '01 01 99'
                                                else : 
                                                    ddaattee = each[1][-10:]
                                            else : 
                                                ddaattee = each[2]
                                            if "VIR." in each[1] or "VIREMENT" in each[1] or "VIR " in each[1]: 
                                                nnoomm = re.search(r" DE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                                ttyyppee = "VIREMENT"
                                            elif "VERSEMENT" in each[1]  or "V ER S EM ENT" in each[1] :
                                                ttyyppee = "VERSEMENT"
                                                nnoomm = re.search(r" DEPLACE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                            elif "CHEQUE A ENC" in each[1] :
                                                ttyyppee = 'chèque'.upper()
                                                nnoomm1 = ""
                                            ddaattee = ddaattee.replace(" ","/") 
                                            ddaattee = ddaattee.replace(ddaattee[6:8],"")
                                            kok = 0
                                            x = ""
                                            each[3] = each[3].replace(" ","")
                                            if len(each[3]) > 4:
                                                bima = each[3][len(each[3])-3: len(each[3])]
                                                each[3] = each[3][0:len(each[3])-3]
                                                for i in range(len(each[3])-1, -1, -1):
                                                    if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21 or kok == 24 or kok == 27 or kok == 30 or kok == 33 or kok == 36 or kok == 39 or kok == 42:
                                                        x = x+'.'
                                                        x = x+each[3][i]
                                                    else:
                                                        x = x+each[3][i]
                                                    kok = kok+1
                                            if kok != 0:
                                                each[3] = ''
                                                for i in range(len(x)-1, -1, -1):
                                                    each[3] = each[3]+x[i]
                                                each[3] = each[3]+bima
                                            tableAWB.append(['D-' +str(year) + '-'+ddaattee[3:5]+'-'+str(iidd),ddaattee,ttyyppee, nnoomm1, each[3] ])
                                            iidd = iidd+1
                                    if algo == 3  : 
                                        if " RECU " in each[1] or " RE CU " in each[1] or " A ENC " in each[1] or "VERSEMENT" in each[1] or "RECU " in each[1]  or "V ER S EM ENT" in each[1] or " RECU" in each[1] : 
                                            print('ressu',each)
                                            if each[2] == '' or each[2] == 'NaN' : 
                                                if any(c.isalpha() for c in each[1][-10:]):
                                                    ddaattee = '01 01 99'
                                                else : 
                                                    ddaattee = each[1][-10:]
                                            else : 
                                                ddaattee = each[2]
                                            if "VIR." in each[1] or "VIREMENT" in each[1] or "VIR " in each[1] : 
                                                nnoomm = re.search(r" DE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                                ttyyppee = "VIREMENT"
                                            elif "VERSEMENT" in each[1] or "V ER S EM ENT" in each[1] :
                                                ttyyppee = "VERSEMENT"
                                                nnoomm = re.search(r" DEPLACE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                            elif "CHEQUE A ENC" in each[1] :
                                                ttyyppee = 'chèque'.upper()
                                                nnoomm1 = ""
                                            ddaattee = ddaattee.replace(" ","/") 
                                            ddaattee = ddaattee.replace(ddaattee[6:8],"")
                                            kok = 0
                                            x = ""
                                            each[5] = each[5].replace(" ","")
                                            if len(each[5]) > 4:
                                                bima = each[5][len(each[5])-3: len(each[5])]
                                                each[5] = each[5][0:len(each[5])-3]
                                                for i in range(len(each[5])-1, -1, -1):
                                                    if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21 or kok == 24 or kok == 27 or kok == 30 or kok == 33 or kok == 36 or kok == 39 or kok == 42:
                                                        x = x+'.'
                                                        x = x+each[5][i]
                                                    else:
                                                        x = x+each[5][i]
                                                    kok = kok+1
                                            if kok != 0:
                                                each[5] = ''
                                                for i in range(len(x)-1, -1, -1):
                                                    each[5] = each[5]+x[i]
                                                each[5] = each[5]+bima
                                            tableAWB.append(['D-' + str(year)  + '-'+ddaattee[3:5]+'-'+str(iidd),ddaattee,ttyyppee, nnoomm1, each[5]])
                                            iidd = iidd+1
                                    if algo == 2 : 
                                        if " RECU " in each[1] or " RE CU " in each[1]  or " A ENC " in each[1] or "VERSEMENT" in each[1] or "RECU " in each[1] or "V ER S EM ENT" in each[1]  or " RECU" in each[1] : 
                                            print('ressu',each)
                                            if each[2] == '' or each[2] == 'NaN' : 
                                                if any(c.isalpha() for c in each[1][-10:]):
                                                    ddaattee = '01 01 99'
                                                else : 
                                                    ddaattee = each[1][-10:]
                                            else : 
                                                ddaattee = each[2]
                                            if "VIR." in each[1] or "VIREMENT" in each[1] or "VIR " in each[1]: 
                                                nnoomm = re.search(r" DE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                                ttyyppee = "VIREMENT"
                                            elif "VERSEMENT" in each[1]  or "V ER S EM ENT" in each[1]:
                                                ttyyppee = "VERSEMENT"
                                                nnoomm = re.search(r" DEPLACE ([^0-9]+|$)", each[1])
                                                try : 
                                                    nnoomm1 = nnoomm.group(1).strip()
                                                except : 
                                                    nnoomm1 = ""
                                            elif "CHEQUE A ENC" in each[1] :
                                                ttyyppee = 'chèque'.upper()
                                                nnoomm1 = ""
                                            ddaattee = ddaattee.replace(" ","/") 
                                            ddaattee = ddaattee.replace(ddaattee[6:8],"")
                                            kok = 0
                                            x = ""
                                            each[4] = each[4].replace(" ","")
                                            if len(each[4]) > 4:
                                                bima = each[4][len(each[4])-3: len(each[4])]
                                                each[4] = each[4][0:len(each[4])-3]
                                                for i in range(len(each[4])-1, -1, -1):
                                                    if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21 or kok == 24 or kok == 27 or kok == 30 or kok == 33 or kok == 36 or kok == 39 or kok == 42:
                                                        x = x+'.'
                                                        x = x+each[4][i]
                                                    else:
                                                        x = x+each[4][i]
                                                    kok = kok+1
                                            if kok != 0:
                                                each[4] = ''
                                                for i in range(len(x)-1, -1, -1):
                                                    each[4] = each[4]+x[i]
                                                each[4] = each[4]+bima
                                            tableAWB.append(['D-' + str(year)  + '-'+ddaattee[3:5]+'-'+str(iidd),ddaattee,ttyyppee, nnoomm1, each[4]])
                                            iidd = iidd+1
                        # for i in range(len(tableAWB)):
                        #     print(tableAWB[i])

                        for j in range(len(tables2)):
                            os.remove(f"table{j}.csv")                
                except : 
                        donothing=0

                # for i in range(len(tableAWB)):
                #     print(tableAWB[i])

                try : 
                    if RBCmi != '' or type(RBCmi) == tuple :
                        def read_data_from_excel_xls(input_file_path):
                            workbook = xlrd.open_workbook(input_file_path)
                            sheet = workbook.sheet_by_index(0)

                            data = []
                            for row_idx in range(sheet.nrows):
                                row_data = sheet.row_values(row_idx)
                                data.append(row_data)

                            return data
                        
                        if tableAWB != [] :
                            iidd = int(tableAWB[len(tableAWB)-1][0][8:])+1
                        else : 
                            if tableUmnia != [] :
                                iidd = int(tableUmnia[len(tableUmnia)-1][0][8:])+1
                            else : 
                                iidd=1
                        dataCMI = read_data_from_excel_xls(RBCmi)
                        for i in range(len(dataCMI)):
                            if i != 0:
                                number = dataCMI[i][4]
                                formatted_number = "{:.2f}".format(number).replace('.', ',')
                                def format_number_with_dots(number):
                                    # Convertir le nombre en chaîne de caractères
                                    num_str = str(number)
                                    
                                    # Diviser la partie entière et la partie décimale
                                    if '.' in num_str:
                                        integer_part, decimal_part = num_str.split('.')
                                    else:
                                        integer_part = num_str
                                        decimal_part = ''
                                    # Formatter la partie entière avec des points
                                    formatted_integer = ''
                                    count = 0
                                    for digit in reversed(integer_part):
                                        formatted_integer = digit + formatted_integer
                                        count += 1
                                        if count % 3 == 0 and count != len(integer_part):
                                            formatted_integer = '.' + formatted_integer
                                    
                                    # Recomposer le nombre avec la partie décimale
                                    formatted_number = formatted_integer
                                    if decimal_part:
                                        formatted_number += ',' + decimal_part
                                    formatted_number = formatted_number[:len(formatted_number)-4] + formatted_number[len(formatted_number)-3:]
                                    return formatted_number
                                if dataCMI[i][6] == 'Authentifié' :
                                    date_input = dataCMI[i][2].split()[0][2:].replace('-', '/')
                                    date_parts = date_input.split('/')
                                    formatted_date = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"
                                    tableCMI.append(['D' + '-' + str(year) + "-" + dataCMI[i][2][5:7] + '-'+str(iidd),formatted_date, "CARTE BANCAIRE" ,dataCMI[i][7],format_number_with_dots(formatted_number) ])
                                    iidd = iidd+1
                except Exception as e:
                    print("Une erreur s'est produite :", e)
                table =  tableUmnia+tableAWB +tableCMI
                style = ttk.Style()
                style.theme_use('default')
                style.configure("Treeview", foreground="black",
                                fieldbackground="silver", rowheight=25)
                scrolly = ttk.Scrollbar(frame2, orient=VERTICAL)
                my_tree = ttk.Treeview(
                    frame2, height=37, yscrollcommand=scrolly.set)

                my_tree.tag_configure('gray', background='gray')
                my_tree.tag_configure('normal', background='white')
                my_tree.tag_configure('blue', background='lightblue')
                my_tree.tag_configure('green', background='lightgreen')
                my_tree.tag_configure('red', background='red')
                my_tree['columns'] = ("Date", "Mois", "Type", "NOM Donneur d'ordre",
                                      "Montant", "Détail", "Montant en lettre", "N° Attestation")

                my_tree.column("#0", width=0, stretch=NO)
                my_tree.column("Date", width=80, anchor=CENTER, minwidth=25)
                my_tree.column("Mois", width=90, anchor=CENTER, minwidth=25)
                my_tree.column("Type", width=280, anchor=CENTER, minwidth=25)
                my_tree.column("NOM Donneur d'ordre", width=395,
                               anchor=CENTER, minwidth=25)
                my_tree.column("Montant", width=170,
                               anchor=CENTER, minwidth=25)
                my_tree.column("Détail", width=80, anchor=CENTER, minwidth=25)
                my_tree.column("Montant en lettre", width=405,
                               anchor=CENTER, minwidth=25)
                my_tree.column("N° Attestation", width=120,
                               anchor=CENTER, minwidth=25)

                my_tree.heading("#0", text="", anchor=W)
                my_tree.heading("Date", text="Date", anchor=CENTER)
                my_tree.heading("Mois", text="Mois", anchor=CENTER)
                my_tree.heading("Type", text="Type", anchor=CENTER)
                my_tree.heading("NOM Donneur d'ordre",
                                text="NOM Donneur d'ordre", anchor=CENTER)
                my_tree.heading("Montant", text="Montant", anchor=CENTER)
                my_tree.heading("Détail", text="Détail", anchor=CENTER)
                my_tree.heading("Montant en lettre",
                                text="Montant en lettre", anchor=CENTER)
                my_tree.heading("N° Attestation",
                                text="N° Attestation", anchor=CENTER)

                my_tree.tag_configure("pink", background="purple")
                my_tree.pack(padx=5, fill='y')
                scrolly.configure(command=my_tree.yview)
                scrolly.place(y=70, height=860, x=1655)
                laDate = table[1][1]
                for i in table:
                    l = len(i[2])
                    if i[2][0:l] != 'CREDIT COMMERCANT':
                        if i[1][3:5] == '01':
                            mois = 'JANVIER'
                        if i[1][3:5] == '02':
                            mois = 'FEVRIER'
                        if i[1][3:5] == '03':
                            mois = 'MARS'
                        if i[1][3:5] == '04':
                            mois = 'AVRIL'
                        if i[1][3:5] == '05':
                            mois = 'MAI'
                        if i[1][3:5] == '06':
                            mois = 'JUIN'
                        if i[1][3:5] == '07':
                            mois = 'JUILLET'
                        if i[1][3:5] == '08':
                            mois = 'AOUT'
                        if i[1][3:5] == '09':
                            mois = 'SEPTEMBRE'
                        if i[1][3:5] == '10':
                            mois = 'OCTOBRE'
                        if i[1][3:5] == '11':
                            mois = 'NOVEMBRE'
                        if i[1][3:5] == '12':
                            mois = 'DECEMBRE'

                        def chiftolett(value, skip=-1):
                            if int(value) < len(units) and units[int(value)]:
                                return [] if int(value) <= skip else [units[int(value)]]
                            for name, v in bases:
                                if int(value) >= v:
                                    return chiftolett(int(int(value)/v), 1 if v <= 1000 else -1) + [name] + chiftolett(int(int(value) % v), 0)
                        lenght = len(i[4])
                        po = i[4][lenght-2]+i[4][lenght-1]
                        mytag = 'normal'
                        if i[2][0:6] == 'ENCAIS':
                            i[2] = 'chèque'.upper()
                        if i[3] == '' or  i[3] == ' ':
                            mytag = 'red'
                        if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[1]) : 
                            mytag = 'pink'
                        if po == '00':
                            my_tree.insert(parent='', index='end', iid=i, text='', values=(i[1], mois, i[2].replace('CREDIT COMMERCANT', 'CARTE BANCAIRE').replace('ENCAISSEMENT CHEQUE NUM - ', ''), i[3], i[4], "", "".join(
                                chiftolett(int(int(i[4].replace(',', '').replace('.', ''))/100))), i[0]), tags=(mytag))
                            mytag = 'normal'
                        else:
                            ntl = "".join(chiftolett(
                                int(int(i[4].replace(',', '').replace('.', ''))/100)))
                            ntl += 'VIRGULE '
                            ntl += "".join(chiftolett(int(po)))
                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                i[1], mois, i[2].replace('CREDIT COMMERCANT', 'CARTE BANCAIRE'), i[3], i[4], "", ntl, i[0]), tags=(mytag))
                            mytag = 'normal'
                children = my_tree.get_children()
                for parent in my_tree.get_children():
                    Dattta.append(my_tree.item(parent)["values"])
                pdff = Button(frame1, text='PDF',
                              command=lambda: showPDF())
                pdff.pack(padx=20)
                lo = Label(frame1, text='Ctrl-P', fg='white',
                           background='black', font=('Times', 10))
                lo.pack()
                frame3.pack(fill='y', padx=10, pady=10)

                Date = StringVar()
                Mois = StringVar()
                Type = StringVar()
                Nom = StringVar()
                Montant = StringVar()
                détail = StringVar()
                Montant_en_lettre = StringVar()
                NAttestation = StringVar()

                def showPDF():
                    global v2
                    global showpdf
                    if showpdf == 0:
                        my_tree.pack_forget()
                        v2.pack(fill='y', side='left', expand=True)
                        pdff.configure(text='DATA')
                        showpdf = 1
                        scrolly.place_forget()
                        frame5.place_forget()
                        delete.pack_forget()
                        edite.pack_forget()
                        Attestation.pack_forget()
                        lo2.pack_forget()
                        lo4.pack_forget()
                        lo5.pack_forget()
                        rowscolor.place_forget()
                        AttTypeD.place_forget()
                        AttTypeS.place_forget()
                        AttTypeP.place_forget()
                        AttTypeC.place_forget()
                        Mme.place_forget()
                        Mlle.place_forget()
                        Mr.place_forget()
                        Dr.place_forget()
                        organise.place_forget()

                    else:
                        add.pack_forget()
                        lo3.pack_forget()
                        pdff.configure(text='PDF')
                        showpdf = 0
                        v2.pack_forget()
                        my_tree.pack()
                        scrolly.place(y=70, height=860, x=1655)
                        frame5.place(x=5, y=202)
                        delete.pack(pady=20, padx=20)
                        lo2.pack()
                        add.pack(pady=20, padx=20)
                        lo3.pack()
                        edite.pack(pady=20, padx=20)
                        lo4.pack()
                        Attestation.pack(pady=20, padx=20)
                        lo5.pack()
                        rowscolor.place(x=75, y=5)
                        AttTypeD.place(x=1525, y=10)
                        AttTypeS.place(x=1560, y=10)
                        AttTypeP.place(x=1595, y=10)
                        AttTypeC.place(x=1630, y=10)
                        Dr.place(x=545,y=10)
                        Mme.place(x=600, y=10)
                        Mlle.place(x=670, y=10)
                        Mr.place(x=740, y=10)
                        organise.place(x=1660, y = 50)


                def editTypeAttD(tree):
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    global ope
                    global Organ 
                    Organ = 5
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][7][0] == "D":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'D-', 'D-')
                                            if Dattta[j][7][0] == "S":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'S-', 'D-')
                                            if Dattta[j][7][0] == "P":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'P-', 'D-')
                                            if Dattta[j][7][0] == "C":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'C-', 'D-')
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0

                def editTypeAttS(tree):
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    global ope
                    global Organ 
                    Organ = 5
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][7][0] == "D":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'D-', 'S-')
                                            if Dattta[j][7][0] == "S":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'S-', 'S-')
                                            if Dattta[j][7][0] == "P":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'P-', 'S-')
                                            if Dattta[j][7][0] == "C":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'C-', 'S-')
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0

                def editTypeAttP(tree):
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    global ope
                    global Organ 
                    Organ = 5
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][7][0] == "D":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'D-', 'P-')
                                            if Dattta[j][7][0] == "S":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'S-', 'P-')
                                            if Dattta[j][7][0] == "P":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'P-', 'P-')
                                            if Dattta[j][7][0] == "C":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'C-', 'P-')
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0
                def organiser(tree) : 
                    global Dattta
                    global Organ 
                    global bach
                    global changerowcolo
                    global kn9lb3la
                    global reloulo 
                    global ope
                    changerowcolo = 0
                    if ope == 0:
                        ope = 1
                        for parent in my_tree.get_children():
                            my_tree.delete(parent)
                        my_tree.tag_configure('yellow', background = 'yellow')
                        my_tree.tag_configure('orange', background = 'orange')
                        my_tree.tag_configure('brown', background='brown')
                        my_tree.tag_configure("pink", background="purple")
                        mytag = 'normal'
                        if reloulo == 0:
                            if Organ == 1 : 
                                Organ = 2 
                                AttTypeD.place_forget()
                                AttTypeS.place_forget()
                                AttTypeP.place_forget()
                                AttTypeC.place_forget()
                                AttTypeC.place(x=1525, y=10)
                                AttTypeD.place(x=1560, y=10)
                                AttTypeS.place(x=1595, y=10)
                                AttTypeP.place(x=1630, y=10)
                                for i in Dattta:
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            elif Organ == 2 : 
                                Organ = 3 
                                AttTypeD.place_forget()
                                AttTypeS.place_forget()
                                AttTypeP.place_forget()
                                AttTypeC.place_forget()
                                AttTypeP.place(x=1525, y=10)
                                AttTypeC.place(x=1560, y=10)
                                AttTypeD.place(x=1595, y=10)
                                AttTypeS.place(x=1630, y=10)
                                for i in Dattta:
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            elif Organ == 0 or Organ == 5 : 
                                AttTypeD.place_forget()
                                AttTypeS.place_forget()
                                AttTypeP.place_forget()
                                AttTypeC.place_forget()
                                AttTypeD.place(x=1525, y=10)
                                AttTypeS.place(x=1560, y=10)
                                AttTypeP.place(x=1595, y=10)
                                AttTypeC.place(x=1630, y=10)
                                Organ = 1 
                                for i in Dattta:
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            elif Organ == 3 : 
                                Organ = 0 
                                AttTypeD.place_forget()
                                AttTypeS.place_forget()
                                AttTypeP.place_forget()
                                AttTypeC.place_forget()
                                AttTypeS.place(x=1525, y=10)
                                AttTypeP.place(x=1560, y=10)
                                AttTypeC.place(x=1595, y=10)
                                AttTypeD.place(x=1630, y=10)
                                for i in Dattta:
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in Dattta:
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            Dattta.clear()
                            for parent in my_tree.get_children():
                                Dattta.append(my_tree.item(parent)["values"])
                        elif reloulo == 1:
                            try:
                                if bach == 'date':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    
                                    elif Organ == 0 or Organ == 5 : 
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[0]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'mois':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5  : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[1]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                    i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'type':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5 : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[2]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'nom':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5 : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[3]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'montant':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5  : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[4]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'détail':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5  : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[5]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'mol':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5  : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[6]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'att':
                                    if Organ == 1 : 
                                        Organ = 2 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeC.place(x=1525, y=10)
                                        AttTypeD.place(x=1560, y=10)
                                        AttTypeS.place(x=1595, y=10)
                                        AttTypeP.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 2 : 
                                        Organ = 3 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeP.place(x=1525, y=10)
                                        AttTypeC.place(x=1560, y=10)
                                        AttTypeD.place(x=1595, y=10)
                                        AttTypeS.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif  Organ == 0 or Organ == 5  : 
                                        
                                        Organ = 1 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeD.place(x=1525, y=10)
                                        AttTypeS.place(x=1560, y=10)
                                        AttTypeP.place(x=1595, y=10)
                                        AttTypeC.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    elif Organ == 3 : 
                                        Organ = 0 
                                        AttTypeD.place_forget()
                                        AttTypeS.place_forget()
                                        AttTypeP.place_forget()
                                        AttTypeC.place_forget()
                                        AttTypeS.place(x=1525, y=10)
                                        AttTypeP.place(x=1560, y=10)
                                        AttTypeC.place(x=1595, y=10)
                                        AttTypeD.place(x=1630, y=10)
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'S':
                                                    mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'P':
                                                    mytag = 'orange'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'C':
                                                    mytag = 'brown'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        for i in Dattta:
                                            if kn9lb3la in i[7]:
                                                if i[7][0] == 'D':
                                                    mytag = 'normal'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            except:
                                donothing = 0
                                children = my_tree.get_children()
                                if children:
                                    my_tree.focus(children[0])
                                    my_tree.selection_set(children[0])
                                    my_tree.selection_add(children[0])
                        ope = 0
                def editTypeAttC(tree):
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    global ope
                    global Organ 
                    Organ = 5
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][7][0] == "D":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'D-', 'C-')
                                            if Dattta[j][7][0] == "S":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'S-', 'C-')
                                            if Dattta[j][7][0] == "P":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'P-', 'C-')
                                            if Dattta[j][7][0] == "C":
                                                Dattta[j][7] = Dattta[j][7].replace(
                                                    'C-', 'C-')
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0

                def editData(tree):
                    global ope
                    if ope == 0:
                        ope = 1
                        curItem = tree.focus()
                        values = tree.item(curItem, "values")
                        indexkhdr = tree.index(curItem)
                        if len(tree.selection()) < 2:
                            if len(curItem) > 0:
                                topEdit = Toplevel()
                                topEdit.title("Modifier")
                                topEdit.geometry("500x430")
                                icon = PhotoImage(file='logo-light.png')
                                window.tk.call(
                                    'wm', 'iconphoto', topEdit._w, icon)
                                topEdit.resizable(width=0, height=0)

                                def idcode(e):
                                    global Dattta
                                    compte7taline = 0 
                                    compteC = 0 
                                    compteD = 0 
                                    compteS = 0
                                    compteP = 0
                                    if values[7][0] == 'D' : 
                                        compteC = 1
                                        compteD = 0 
                                        compteS = 1
                                        compteP = 1
                                    if values[7][0] == 'P' : 
                                        compteC = 1
                                        compteD = 1 
                                        compteS = 1
                                        compteP = 0
                                    if values[7][0] == 'C' : 
                                        compteC = 0
                                        compteD = 1 
                                        compteS = 1
                                        compteP = 1
                                    if values[7][0] == 'S' : 
                                        compteC = 1
                                        compteD = 1
                                        compteS = 0
                                        compteP = 1
                                    for i in range(len(Dattta)):
                                        if Dattta[i][7] == values[7]  : 
                                            compte7taline =  compte7taline + 1 
                                            break
                                        else : 
                                            compte7taline =  compte7taline + 1 
                                    id = values[7][8:len(values[7])]
                                    if e == "C":
                                        C.configure(
                                            font=('Times', 11, 'bold'), bg='green', fg='white')
                                        P.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        S.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        D.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    if e == "S":
                                        S.configure(
                                            font=('Times', 11, 'bold'), bg='green', fg='white')
                                        P.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        C.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        D.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    if e == "D":
                                        D.configure(
                                            font=('Times', 11, 'bold'), bg='green', fg='white')
                                        C.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        S.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        P.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    if e == "P":
                                        P.configure(
                                            font=('Times', 11, 'bold'), bg='green', fg='white')
                                        C.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        S.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                                        D.configure(
                                            font=('Times', 11, 'bold'), bg='blue', fg='white')

                                    if D.cget('bg') == 'blue' and C.cget('bg') == 'blue' and P.cget('bg') == 'blue' and S.cget('bg') == 'blue':
                                        for li in range(len(Dattta)+1):
                                            if li == compte7taline : 
                                                break
                                            if Dattta[li][7][0] == "D":
                                                compteD = compteD + 1
                                        if compteD == 0 : 
                                            compteD = 1  
                                        e16.config(state='normal')
                                        e16.delete(0, 'end')
                                        e16.insert(
                                            'end', 'D-' + e9.get()[6: len(e9.get())] + '-' + e9.get()[3:5] + '-' + str(compteD))
                                        if e9.get() == '':
                                            e16.delete(0, 'end')
                                            e16.insert('end', 'D-AA-MM-'+ str(compteD))
                                        e16.config(state='disabled')

                                    if D.cget('bg') == 'green':
                                        for li in range(len(Dattta)+1):
                                            if li == compte7taline : 
                                                break
                                            if Dattta[li][7][0] == "D":
                                                compteD = compteD + 1 
                                        if compteD == 0 : 
                                            compteD = 1 
                                        e16.config(state='normal')
                                        e16.delete(0, 'end')
                                        e16.insert(
                                            'end', 'D-' + e9.get()[6: len(e9.get())] + '-' + e9.get()[3:5] + '-'+ str(compteD))
                                        if e9.get() == '':
                                            e16.delete(0, 'end')
                                            e16.insert('end', 'D-AA-MM-'+ str(compteD))
                                        e16.config(state='disabled')

                                    if C.cget('bg') == 'green':
                                        for li in range(len(Dattta)+1):
                                            if li == compte7taline : 
                                                break
                                            if Dattta[li][7][0] == "C":
                                                compteC = compteC + 1 
                                        if compteC == 0 : 
                                            compteC = 1 
                                        e16.config(state='normal')
                                        e16.delete(0, 'end')
                                        e16.insert(
                                            'end', 'C-' + e9.get()[6: len(e9.get())] + '-' + e9.get()[3:5] + '-'+ str(compteC))
                                        if e9.get() == '':
                                            e16.delete(0, 'end')
                                            e16.insert('end', 'C-AA-MM-'+ str(compteC))
                                        e16.config(state='disabled')

                                    if P.cget('bg') == 'green':
                                        for li in range(len(Dattta)+1):
                                            if li == compte7taline : 
                                                break
                                            if Dattta[li][7][0] == "P":
                                                compteP = compteP + 1
                                        if compteP == 0 : 
                                            compteP = 1 
                                        e16.config(state='normal')
                                        e16.delete(0, 'end')
                                        e16.insert(
                                            'end', 'P-' + e9.get()[6: len(e9.get())] + '-' + e9.get()[3:5] + '-' + str(compteP))
                                        if e9.get() == '':
                                            e16.delete(0, 'end')
                                            e16.insert('end', 'P-AA-MM-'+ str(compteP))
                                        e16.config(state='disabled')

                                    if S.cget('bg') == 'green':
                                        for li in range(len(Dattta)+1):
                                            if li == compte7taline : 
                                                break
                                            if Dattta[li][7][0] == "S":
                                                compteS = compteS + 1 
                                        if compteS == 0 : 
                                            compteS = 1 
                                        e16.config(state='normal')
                                        e16.delete(0, 'end')
                                        e16.insert(
                                            'end', 'S-' + e9.get()[6: len(e9.get())] + '-' + e9.get()[3:5] + '-' + str(compteS))
                                        if e9.get() == '':
                                            e16.delete(0, 'end')
                                            e16.insert('end', 'S-AA-MM-'+ str(compteS))
                                        e16.config(state='disabled')
                                D = Button(
                                    topEdit, text='D', command=lambda: idcode("D"))
                                D.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                D.place(x=200, y=20)
                                S = Button(
                                    topEdit, text='S', command=lambda: idcode("S"))
                                S.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                S.place(x=230, y=20)
                                P = Button(
                                    topEdit, text='P', command=lambda: idcode("P"))
                                P.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                P.place(x=260, y=20)
                                C = Button(
                                    topEdit, text='C', command=lambda: idcode("C"))
                                C.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                C.place(x=290, y=20)
                                l9 = Label(topEdit, text="Date", width=20,
                                           font=('Times', 11, 'bold'))
                                e9 = Entry(
                                    topEdit, textvariable=Date, width=25)
                                l9.place(x=50, y=60)
                                e9.place(x=200, y=60)

                                l10 = Label(topEdit, text="Mois", width=20,
                                            font=('Times', 11, 'bold'))
                                e10 = Entry(
                                    topEdit, textvariable=Mois, width=25)
                                l10.place(x=50, y=100)
                                e10.place(x=200, y=100)

                                l11 = Label(topEdit, text="Type", width=20,
                                            font=('Times', 11, 'bold'))

                                options = [
                                    "VIREMENT PERMANENT",
                                    "VIREMENT",
                                    "chèque".upper(),
                                    "CARTE BANCAIRE",
                                    "espèces".upper(),
                                ]

                                # datatype of menu text
                                clicked = StringVar()

                                # initial menu text
                                clicked.set(values[2])

                                # Create Dropdown menu
                                drop = OptionMenu(topEdit, clicked, *options)
                                drop.place(x=200, y=135)
                                l11.place(x=50, y=140)

                                l12 = Label(topEdit, text="Nom donneur d'ordre",
                                            width=20, font=('Times', 11, 'bold'))
                                e12 = Entry(
                                    topEdit, textvariable=Nom, width=25)
                                l12.place(x=50, y=180)
                                e12.place(x=200, y=180)

                                l13 = Label(topEdit, text="Montant",
                                            width=20, font=('Times', 11, 'bold'))
                                e13 = Entry(
                                    topEdit, textvariable=Montant, width=25)
                                l13.place(x=50, y=220)
                                e13.place(x=200, y=220)

                                l14 = Label(topEdit, text="Détail",
                                            width=20, font=('Times', 11, 'bold'))
                                e14 = Entry(
                                    topEdit, textvariable=détail, width=25)
                                l14.place(x=50, y=260)
                                e14.place(x=200, y=260)

                                l15 = Label(topEdit, text="Montant en lettre",
                                            width=20, font=('Times', 11, 'bold'))
                                e15 = Entry(
                                    topEdit, textvariable=Montant_en_lettre, width=25)
                                l15.place(x=50, y=300)
                                e15.place(x=200, y=300)

                                l16 = Label(topEdit, text="N° Attestation",
                                            width=20, font=('Times', 11, 'bold'))
                                e16 = Entry(
                                    topEdit, textvariable=NAttestation, width=25)
                                l16.place(x=50, y=340)
                                e16.place(x=200, y=340)
                                e9.focus()
                                if values[7][0] == 'D' : 
                                    D.configure(
                                        font=('Times', 11, 'bold'), bg='green', fg='white')
                                    C.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    P.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    S.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                elif values[7][0] == 'S' : 
                                    D.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    C.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    P.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    S.configure(
                                        font=('Times', 11, 'bold'), bg='green', fg='white')
                                elif values[7][0] == 'C' : 
                                    D.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    C.configure(
                                        font=('Times', 11, 'bold'), bg='green', fg='white')
                                    P.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    S.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                elif values[7][0] == 'P' : 
                                    D.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    C.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')
                                    P.configure(
                                        font=('Times', 11, 'bold'), bg='green', fg='white')
                                    S.configure(
                                        font=('Times', 11, 'bold'), bg='blue', fg='white')

                                e9.bind("<KeyRelease>", idcode)

                                def insertData(tree):
                                    global ope
                                    nonlocal e9, e10, e12, e13, e14, e15, e16
                                    global Dattta
                                    global Organ 
                                    Organ = 5
                                    da = Date.get().strip()
                                    mo = Mois.get().strip()
                                    ty = clicked.get().upper()
                                    no = Nom.get().strip().upper()
                                    mon = Montant.get().strip()
                                    dé = détail.get().strip()
                                    mol = Montant_en_lettre.get().strip()
                                    att = NAttestation.get().strip()
                                    x = ""
                                    kok = 0

                                    if mon.count(',') == 1:
                                        if len(mon)-mon.index(',') != 3:
                                            messagebox.showinfo(
                                                title='Erreur !!', message="Montant invalide, merci de saisir deux chiffres après la virgule.", parent=topEdit)
                                            ope = 0
                                            return
                                        else:
                                            for i in mon:
                                                if i == ',' or i == '.':
                                                    pass
                                                elif not i.isnumeric():
                                                    messagebox.showinfo(
                                                        title='Erreur !!', message="Montant invalide", parent=topEdit)
                                                    ope = 0
                                                    return
                                            po = mon[len(mon)-2] + \
                                                mon[len(mon)-1]
                                            if po == '00':
                                                mol = "".join(chiftolett(
                                                    int(int(mon.replace(',', '').replace('.', ''))/100)))
                                            else:
                                                mol = "".join(chiftolett(
                                                    int(int(mon.replace(',', '').replace('.', ''))/100)))
                                                mol += 'VIRGULE '
                                                if po == "01" or po == "02" or po == "03" or po == "04" or po == "05" or po == "06" or po == "07" or po == "08" or po == "09":
                                                    mol += "ZERO "
                                                mol += "".join(chiftolett(int(po)))
                                    else:
                                        for i in mon:
                                            if i == '.':
                                                pass
                                            elif not i.isnumeric():
                                                messagebox.showinfo(
                                                    title='Erreur !!', message="Montant invalide", parent=topEdit)
                                                ope = 0
                                                return
                                        mol = "".join(chiftolett(
                                            int(int(mon.replace(',', '').replace('.', '')))))
                                        mon = "".join((mon, ',00'))
                                    if da[0:2].isnumeric() and da[3:5].isnumeric() and da[6:8].isnumeric() and len(da) == 8:
                                        if int(da[3:5]) > 12 or int(da[0:2]) > 31:
                                            messagebox.showinfo(
                                                title='Erreur !!', message="* Le nombre de jour ne doit pas être superieur à 31. \n\n* Le nombre de mois ne doit pas être superieur à 12.", parent=topEdit)
                                            ope = 0
                                            return
                                    else:
                                        messagebox.showinfo(
                                            title='Erreur !!', message="La date doit être sous la forme suivante :\n\n            'JJxMMxAA'", parent=topEdit)
                                        ope = 0
                                        return

                                    mon = mon.replace('.', '')
                                    if len(mon) > 4:
                                        bima = mon[len(mon)-3: len(mon)]
                                        mon = mon[0:len(mon)-3]
                                        for i in range(len(mon)-1, -1, -1):
                                            if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21 or kok == 24 or kok == 27 or kok == 30 or kok == 33 or kok == 36 or kok == 39 or kok == 42:
                                                x = x+'.'
                                                x = x+mon[i]
                                            else:
                                                x = x+mon[i]
                                            kok = kok+1
                                    if kok != 0:
                                        mon = ''
                                        for i in range(len(x)-1, -1, -1):
                                            mon = mon+x[i]
                                        mon = mon+bima

                                    hoho = list(da)
                                    hoho[2] = '/'
                                    hoho[5] = '/'
                                    da = ''.join(hoho)
                                    if da != values[0]:
                                        if da[3:5] == '01':
                                            mo = 'JANVIER'
                                        if da[3:5] == '02':
                                            mo = 'FEVRIER'
                                        if da[3:5] == '03':
                                            mo = 'MARS'
                                        if da[3:5] == '04':
                                            mo = 'AVRIL'
                                        if da[3:5] == '05':
                                            mo = 'MAI'
                                        if da[3:5] == '06':
                                            mo = 'JUIN'
                                        if da[3:5] == '07':
                                            mo = 'JUILLET'
                                        if da[3:5] == '08':
                                            mo = 'AOUT'
                                        if da[3:5] == '09':
                                            mo = 'SEPTEMBRE'
                                        if da[3:5] == '10':
                                            mo = 'OCTOBRE'
                                        if da[3:5] == '11':
                                            mo = 'NOVEMBRE'
                                        if da[3:5] == '12':
                                            mo = 'DECEMBRE'
                                    hadighirxhaha = tree.item(
                                        curItem, 'values')[7]
                                    tree.item(curItem, values=(
                                        da, mo, ty, no, mon, dé, mol, att.replace(att[2:4], da[6:8]).replace(att[5:7], da[3:5])))
                                    e9.delete(0, END)
                                    e10.config(state='normal')
                                    e10.delete(0, END)
                                    e12.delete(0, END)
                                    e13.delete(0, END)
                                    e14.delete(0, END)
                                    e15.config(state='normal')
                                    e16.config(state='normal')
                                    e15.delete(0, END)
                                    e16.delete(0, END)
                                    topEdit.destroy()
                                    if values[0].strip() != str(da) or values[2].strip() != str(ty) or values[3].strip() != str(no) or values[4].strip() != str(mon) or values[5].strip() != str(dé) or values[7].strip() != str(att):
                                        tree.item(curItem, tags='green')
                                        j = 0
                                        for fi in Dattta:
                                            if hadighirxhaha == fi[7]:
                                                Dattta[j] = [tree.item(curItem, 'values')[0], tree.item(curItem, 'values')[1], tree.item(curItem, 'values')[2], tree.item(curItem, 'values')[
                                                    3], tree.item(curItem, 'values')[4], tree.item(curItem, 'values')[5], tree.item(curItem, 'values')[6], att.replace(att[2:4], da[6:8]).replace(att[5:7], da[3:5])]
                                            j += 1
                                        stflhzak(my_tree)

                                        def change_row_color_by_index(index):
                                            if 0 <= index < len(tree.get_children()):
                                                item_id = tree.get_children()[index]
                                                tree.tag_configure("custom_tag", background="lightgreen")
                                                tree.item(item_id, tags=("custom_tag",))
                                        change_row_color_by_index(indexkhdr)

                                        messagebox.showinfo(
                                            title='Enregistrer', message='Donnée(s) modifiée(s) !')
                                    ope = 0
                                    return
                                e9.insert(0, values[0])
                                e10.insert(0, values[1])
                                e12.insert(0, values[3])
                                e13.insert(0, values[4])
                                e14.insert(0, values[5])
                                e15.insert(0, values[6])
                                e16.insert(0, values[7])

                                def hantaTchouf():
                                    global ope
                                    ope = 0
                                    e9.delete(0, END)
                                    e10.config(state='normal')
                                    e10.delete(0, END)
                                    e12.delete(0, END)
                                    e13.delete(0, END)
                                    e14.delete(0, END)
                                    e15.config(state='normal')
                                    e16.config(state='normal')
                                    e15.delete(0, END)
                                    e16.delete(0, END)
                                    topEdit.destroy()
                                e16.config(state='disabled')
                                e15.config(state='disabled')
                                e10.config(state='disabled')
                                submitbutton = Button(
                                    topEdit, text='Enregistrer', command=lambda: insertData(my_tree))
                                submitbutton.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                submitbutton.place(x=300, y=380)
                                cancelbutton = Button(
                                    topEdit, text='Annulé', command=lambda: hantaTchouf())
                                cancelbutton.configure(
                                    font=('Times', 11, 'bold'), bg='red', fg='white')
                                cancelbutton.place(x=200, y=380)
                                topEdit.protocol(
                                    "WM_DELETE_WINDOW", hantaTchouf)
                                topEdit.bind(
                                    "<Return>", lambda e: insertData(my_tree))
                                topEdit.bind(
                                    "<Escape>", lambda e: hantaTchouf())
                            else:
                                messagebox.showinfo(
                                    title='!!', message='Merci de selectionner une ligne')
                                ope = 0
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner une seule ligne')
                            ope = 0

                def addData(tree):
                    global Dattta
                    global reloulo
                    global ope
                    if ope == 0:
                        ope = 1

                        def idcode(e):
                            global Dattta
                            if e == "C":
                                C.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                P.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                S.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                D.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                            if e == "S":
                                S.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                P.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                C.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                D.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                            if e == "D":
                                D.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                C.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                S.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                P.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                            if e == "P":
                                P.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                C.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                S.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                                D.configure(
                                    font=('Times', 11, 'bold'), bg='blue', fg='white')
                            compteD = 0 
                            compteS = 0
                            compteC=0
                            compteP=0
                            for lpo in range(len(Dattta)):
                                if Dattta[lpo][7][0] == "D":
                                    compteD = compteD + 1 
                                if Dattta[lpo][7][0] == "S":
                                    compteS = compteS + 1 
                                if Dattta[lpo][7][0] == "P":
                                    compteP = compteP + 1 
                                if Dattta[lpo][7][0] == "C":
                                    compteC = compteC + 1 
                            if D.cget('bg') == 'blue' and C.cget('bg') == 'blue' and P.cget('bg') == 'blue' and S.cget('bg') == 'blue':
                                l27.configure(
                                    text='D-'+e20.get()[6: len(e20.get())]+'-'+e20.get()[3:5]+'-'+ str(compteD+1))
                                if e20.get() == '':
                                    l27.configure(text='D-AA-MM-'+ str(compteD+1))
                            if D.cget('bg') == 'green':
                                l27.configure(
                                    text='D-'+e20.get()[6: len(e20.get())]+'-'+e20.get()[3:5]+'-'+ str(compteD+1))
                                if e20.get() == '':
                                    l27.configure(text='D-AA-MM-'+ str(compteD+1))
                            if C.cget('bg') == 'green':
                                l27.configure(
                                    text='C-'+e20.get()[6: len(e20.get())]+'-'+e20.get()[3:5]+'-'+ str(compteC+1))
                                if e20.get() == '':
                                    l27.configure(text='C-AA-MM-'+ str(compteC+1))
                            if P.cget('bg') == 'green':
                                l27.configure(
                                    text='P-'+e20.get()[6: len(e20.get())]+'-'+e20.get()[3:5]+'-'+ str(compteP+1))
                                if e20.get() == '':
                                    l27.configure(text='P-AA-MM-'+ str(compteP+1))
                            if S.cget('bg') == 'green':
                                l27.configure(
                                    text='S-'+e20.get()[6: len(e20.get())]+'-'+e20.get()[3:5]+'-'+ str(compteS+1))
                                if e20.get() == '':
                                    l27.configure(text='S-AA-MM-'+ str(compteS+1))

                        topAdd = Toplevel()
                        topAdd.title("Ajouter")
                        topAdd.geometry("500x430")
                        topAdd.resizable(width=0, height=0)
                        icon = PhotoImage(file='logo-light.png')
                        window.tk.call('wm', 'iconphoto', topAdd._w, icon)

                        D = Button(
                            topAdd, text='D', command=lambda: idcode("D"))
                        D.configure(
                            font=('Times', 11, 'bold'), bg='green', fg='white')
                        D.place(x=200, y=20)
                        S = Button(
                            topAdd, text='S', command=lambda: idcode("S"))
                        S.configure(
                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                        S.place(x=230, y=20)
                        P = Button(
                            topAdd, text='P', command=lambda: idcode("P"))
                        P.configure(
                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                        P.place(x=260, y=20)
                        C = Button(
                            topAdd, text='C', command=lambda: idcode("C"))
                        C.configure(
                            font=('Times', 11, 'bold'), bg='blue', fg='white')
                        C.place(x=290, y=20)

                        l20 = Label(topAdd, text="Date", width=20,
                                    font=('Times', 11, 'bold'))
                        e20 = Entry(
                            topAdd, textvariable=Date, width=25)
                        l20.place(x=50, y=60)
                        e20.place(x=200, y=60)
                        l19 = Label(topAdd, text="Mois", width=20,
                                    font=('Times', 11, 'bold'))
                        e19 = Entry(
                            topAdd, textvariable=Mois, width=25)
                        l19.place(x=50, y=100)
                        e19.place(x=200, y=100)
                        l21 = Label(topAdd, text="Type", width=20,
                                    font=('Times', 11, 'bold'))

                        options = [
                            "VIREMENT PERMANENT",
                            "VIREMENT",
                            "chèque".upper(),
                            "CARTE BANCAIRE",
                            "espèces".upper(),
                        ]

                        # datatype of menu text
                        clicked = StringVar()

                        # initial menu text
                        clicked.set("VIREMENT")

                        # Create Dropdown menu
                        drop = OptionMenu(topAdd, clicked, *options)
                        drop.place(x=200, y=135)
                        l21.place(x=50, y=140)
                        l22 = Label(topAdd, text="Nom donneur d'ordre",
                                    width=20, font=('Times', 11, 'bold'))
                        e22 = Entry(
                            topAdd, textvariable=Nom, width=25)
                        l22.place(x=50, y=180)
                        e22.place(x=200, y=180)
                        l23 = Label(topAdd, text="Montant",
                                    width=20, font=('Times', 11, 'bold'))
                        e23 = Entry(
                            topAdd, textvariable=Montant, width=25)
                        l23.place(x=50, y=220)
                        e23.place(x=200, y=220)
                        l24 = Label(topAdd, text="Détail",
                                    width=20, font=('Times', 11, 'bold'))
                        e24 = Entry(
                            topAdd, textvariable=détail, width=25)
                        l24.place(x=50, y=260)
                        e24.place(x=200, y=260)
                        l25 = Label(topAdd, text="Montant en lettre",
                                    width=20, font=('Times', 11, 'bold'))
                        e25 = Entry(
                            topAdd, textvariable=Montant_en_lettre, width=25)
                        l25.place(x=50, y=300)
                        e25.place(x=200, y=300)
                        l26 = Label(topAdd, text="N° Attestation",
                                    width=20, font=('Times', 11, 'bold'))
                        compteD = 0 
                        for lpo in range(len(Dattta)):
                            if Dattta[lpo][7][0] == "D":
                                compteD = compteD + 1 
                        l27 = Label(topAdd, text="D-AA-MM-"+ str(compteD+1),
                                    width=20, font=('Times', 11, 'bold'))
                        l26.place(x=50, y=340)
                        l27.place(x=170, y=345)
                        e20.focus()
                        e20.bind("<KeyRelease>", idcode)
                        e19.config(state='disabled')
                        e25.config(state='disabled')

                        def adddData(tree):
                            global ope
                            nonlocal e20, e19, e22, e23, e24, e25, l27
                            global Dattta
                            global kn9lb3la
                            global bach
                            global changerowcolo
                            global Organ 
                            Organ = 5
                            compteX = 0 
                            kok = 0
                            da = Date.get().strip()
                            mo = Mois.get().strip()
                            ty = clicked.get().upper()
                            no = Nom.get().strip().upper()
                            mon = Montant.get().strip()
                            dé = détail.get().strip()
                            mol = Montant_en_lettre.get().strip()
                            typeatt = l27.cget('text')
                            if typeatt[0] == "D":
                                att = "".join(
                                    "D-"+da[6:8]+'-'+da[3:5]+'-'+NAttestation.get().strip())
                            if typeatt[0] == "P":
                                att = "".join(
                                    "P-"+da[6:8]+'-'+da[3:5]+'-'+NAttestation.get().strip())
                            if typeatt[0] == "S":
                                att = "".join(
                                    "S-"+da[6:8]+'-'+da[3:5]+'-'+NAttestation.get().strip())
                            if typeatt[0] == "C":
                                att = "".join(
                                    "C-"+da[6:8]+'-'+da[3:5]+'-'+NAttestation.get().strip())
                            if mon == '' or da == '':
                                messagebox.showinfo(
                                    title='Erreur !!', message="la date et le montant sont obligatoire pour enregistrer un paiement.", parent=topAdd)
                                ope = 0
                                return
                            if mon.count(',') == 1:
                                if len(mon)-mon.index(',') != 3:
                                    messagebox.showinfo(
                                        title='Erreur !!', message="Montant invalide, merci de saisir deux chiffres après la virgule.", parent=topAdd)
                                    ope = 0
                                    return
                                else:
                                    for i in mon:
                                        if i == ',' or i == '.':
                                            pass
                                        elif not i.isnumeric():
                                            messagebox.showinfo(
                                                title='Erreur !!', message="Montant invalide", parent=topAdd)
                                            ope = 0
                                            return
                                    po = mon[len(mon)-2] + \
                                        mon[len(mon)-1]
                                    if po == '00':
                                        mol = "".join(chiftolett(
                                            int(int(mon.replace(',', '').replace('.', ''))/100)))
                                    else:
                                        mol = "".join(chiftolett(
                                            int(int(mon.replace(',', '').replace('.', ''))/100)))
                                        mol += 'VIRGULE '
                                        if po == "01" or po == "02" or po == "03" or po == "04" or po == "05" or po == "06" or po == "07" or po == "08" or po == "09":
                                            mol += "ZERO "
                                        mol += "".join(chiftolett(int(po)))
                            else:
                                for i in mon:
                                    if i == '.':
                                        pass
                                    elif not i.isnumeric():
                                        messagebox.showinfo(
                                            title='Erreur !!', message="Montant invalide", parent=topAdd)
                                        ope = 0
                                        return
                                mol = "".join(chiftolett(
                                    int(int(mon.replace(',', '').replace('.', '')))))
                                mon = "".join((mon, ',00'))
                            if da[0:2].isnumeric() and da[3:5].isnumeric() and da[6:8].isnumeric() and len(da) == 8:
                                if int(da[3:5]) > 12 or int(da[0:2]) > 31:
                                    messagebox.showinfo(
                                        title='Erreur !!', message="* Le nombre de jour ne doit pas être superieur à 31. \n\n* Le nombre de mois ne doit pas être superieur à 12.", parent=topAdd)
                                    ope = 0
                                    return
                            else:
                                messagebox.showinfo(
                                    title='Erreur !!', message="La date doit être sous la forme suivante :\n\n            'JJxMMxAA'", parent=topAdd)
                                ope = 0
                                return
                            x = ''
                            mon = mon.replace('.', '')
                            if len(mon) > 4:
                                bima = mon[len(mon)-3: len(mon)]
                                mon = mon[0:len(mon)-3]
                                for i in range(len(mon)-1, -1, -1):
                                    if kok == 3 or kok == 6 or kok == 9 or kok == 12 or kok == 15 or kok == 18 or kok == 21 or kok == 24 or kok == 27 or kok == 30 or kok == 33 or kok == 36  or kok == 39 or kok == 42:
                                        x = x+'.'
                                        x = x+mon[i]
                                    else:
                                        x = x+mon[i]
                                    kok = kok+1
                            if kok != 0:
                                mon = ''
                                for i in range(len(x)-1, -1, -1):
                                    mon = mon+x[i]
                                mon = mon+bima
                            hoho = list(da)
                            hoho[2] = '/'
                            hoho[5] = '/'
                            da = ''.join(hoho)
                            if da[3:5] == '01':
                                mo = 'JANVIER'
                            if da[3:5] == '02':
                                mo = 'FEVRIER'
                            if da[3:5] == '03':
                                mo = 'MARS'
                            if da[3:5] == '04':
                                mo = 'AVRIL'
                            if da[3:5] == '05':
                                mo = 'MAI'
                            if da[3:5] == '06':
                                mo = 'JUIN'
                            if da[3:5] == '07':
                                mo = 'JUILLET'
                            if da[3:5] == '08':
                                mo = 'AOUT'
                            if da[3:5] == '09':
                                mo = 'SEPTEMBRE'
                            if da[3:5] == '10':
                                mo = 'OCTOBRE'
                            if da[3:5] == '11':
                                mo = 'NOVEMBRE'
                            if da[3:5] == '12':
                                mo = 'DECEMBRE'
                            hahia = Dattta[len(
                                Dattta)-1][7][8:len(Dattta[len(Dattta)-1][7])]
                            hahoua = '%0d' % (int(hahia)+1)
                            j = 0
                            try:
                                compteD = 0 
                                compteS = 0
                                compteC=0
                                compteP=0
                                compteX=0
                                for li in range(len(Dattta)+1):
                                    if Dattta[li][7][8: len(Dattta[li][7])] == NAttestation.get().strip():
                                        j += 1
                                    if j == 1:
                                        try:
                                            Dattta.insert(
                                                li, [da, mo, ty, no, mon, dé, mol, att])
                                        except:
                                            donothing = 0
                                    if j != 0 and j != 1:
                                        try:
                                            if Dattta[li][7][0] == "D":
                                                compteD = compteD + 1 
                                                Dattta[li][7] = 'D-'+Dattta[li][0][3:5]+'-' + \
                                                    Dattta[li][0][6:len(
                                                        Dattta[li][0])]+'-'+str(compteD)
                                            if Dattta[li][7][0] == "S":
                                                compteS = compteS + 1 
                                                Dattta[li][7] = 'S-'+Dattta[li][0][3:5]+'-' + \
                                                    Dattta[li][0][6:len(
                                                        Dattta[li][0])]+'-'+str(compteS)
                                            if Dattta[li][7][0] == "P":
                                                compteP = compteP + 1 
                                                Dattta[li][7] = 'P-'+Dattta[li][0][3:5]+'-' + \
                                                    Dattta[li][0][6:len(
                                                        Dattta[li][0])]+'-'+str(compteP)
                                            if Dattta[li][7][0] == "C":
                                                compteC = compteC + 1 
                                                Dattta[li][7] = 'C-'+Dattta[li][0][3:5]+'-' + \
                                                    Dattta[li][0][6:len(
                                                        Dattta[li][0])]+'-'+str(compteC)

                                        except:
                                            donothing = 0
                            except:
                                donothing = 0
                            if j != 0:
                                for parent in tree.get_children():
                                    tree.delete(parent)
                                if reloulo == 0:
                                    for i in Dattta:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))

                                else:
                                    for parent in tree.get_children():
                                        tree.delete(parent)
                                    try:
                                        if bach == 'date':
                                            for i in Dattta:

                                                if kn9lb3la in i[0]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'mois':
                                            for i in Dattta:
                                                if kn9lb3la in i[1]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'type':
                                            for i in Dattta:
                                                if kn9lb3la in i[2]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'nom':
                                            for i in Dattta:
                                                if kn9lb3la in i[3]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'montant':
                                            for i in Dattta:
                                                if kn9lb3la in i[4]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'détail':
                                            for i in Dattta:
                                                if kn9lb3la in i[5]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'mol':
                                            for i in Dattta:
                                                if kn9lb3la in i[6]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'att':
                                            for i in Dattta:
                                                if kn9lb3la in i[7]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    except:
                                        donothing = 0
                                stflhzak(my_tree)
                                if showpdf == 1:
                                    showPDF()
                                e20.delete(0, END)
                                e22.delete(0, END)
                                e23.delete(0, END)
                                e24.delete(0, END)
                                e25.delete(0, END)
                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                                topAdd.destroy()
                                ope = 0
                                if my_tree.get_children():
                                    for o in range(len(my_tree.get_children())):
                                        if my_tree.item(my_tree.get_children()[o])["values"][7] == att:
                                            my_tree.focus_set()
                                            my_tree.focus(
                                                my_tree.get_children()[o])
                                            my_tree.selection_add(
                                                my_tree.get_children())
                                            my_tree.selection_set(
                                                my_tree.get_children()[o])
                                stflhzak(my_tree)
                                messagebox.showinfo(
                                    title='Enregistrer', message='Donnée(s) ajoutée(s) !')
                                return
                            if j == 0:
                                for i in range(len(Dattta)) : 
                                    if Dattta[i][7][0] == typeatt[0]:
                                        compteX = compteX + 1 
                                if typeatt[0] == "D":
                                    hana = [da, mo, ty, no, mon, dé, mol, "".join(
                                        "D-"+da[6:8]+'-'+da[3:5]+'-'+str(compteX+1))]
                                if typeatt[0] == "P":
                                    hana = [da, mo, ty, no, mon, dé, mol, "".join(
                                        "P-"+da[6:8]+'-'+da[3:5]+'-'+str(compteX+1))]
                                if typeatt[0] == "S":
                                    hana = [da, mo, ty, no, mon, dé, mol, "".join(
                                        "S-"+da[6:8]+'-'+da[3:5]+'-'+str(compteX+1))]
                                if typeatt[0] == "C":
                                    hana = [da, mo, ty, no, mon, dé, mol, "".join(
                                        "C-"+da[6:8]+'-'+da[3:5]+'-'+str(compteX+1))]
                                Dattta.append(hana)
                                for parent in tree.get_children():
                                    tree.delete(parent)
                                if reloulo != 1:
                                    for i in Dattta:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        mytag = 'normal'
                                else:
                                    try:
                                        if bach == 'date':
                                            for i in Dattta:
                                                if kn9lb3la in i[0]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'mois':
                                            for i in Dattta:
                                                if kn9lb3la in i[1]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'type':
                                            for i in Dattta:
                                                if kn9lb3la in i[2]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'nom':
                                            for i in Dattta:
                                                if kn9lb3la in i[3]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'montant':
                                            for i in Dattta:
                                                if kn9lb3la in i[4]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'détail':
                                            for i in Dattta:
                                                if kn9lb3la in i[5]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'mol':
                                            for i in Dattta:
                                                if kn9lb3la in i[6]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        if bach == 'att':
                                            for i in Dattta:
                                                if kn9lb3la in i[7]:
                                                    mytag = 'normal'
                                                    if i[7][0] == 'P':
                                                        mytag = 'orange'
                                                    if i[7][0] == 'C':
                                                        mytag = 'brown'
                                                    if i[7][0] == 'D':
                                                        mytag = 'normal'
                                                    if i[7][0] == 'S':
                                                        mytag = 'yellow'
                                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    except:
                                        donothing = 0
                            if showpdf == 1:
                                showPDF()
                            e20.delete(0, END)
                            e22.delete(0, END)
                            e23.delete(0, END)
                            e24.delete(0, END)
                            e25.delete(0, END)
                            if changerowcolo == 1:
                                changerowcolo = 0
                                changerowcolor()
                            else:
                                changerowcolo = 1
                                changerowcolor()
                            topAdd.destroy()
                            ope = 0
                            if my_tree.get_children():
                                for o in range(len(my_tree.get_children())):
                                    if my_tree.item(my_tree.get_children()[o])["values"][7] == "".join("D-"+da[6:8]+'-'+da[3:5]+'-'+hahoua):
                                        my_tree.focus_set()
                                        my_tree.focus(
                                            my_tree.get_children()[o])
                                        my_tree.selection_add(
                                            my_tree.get_children())
                                        my_tree.selection_set(
                                            my_tree.get_children()[o])
                            stflhzak(my_tree)
                            messagebox.showinfo(
                                title='Enregistrer', message='Donnée(s) ajoutée(s) !')
                            return

                        def hantaTchouf():
                            global ope
                            ope = 0
                            e20.delete(0, END)
                            e22.delete(0, END)
                            e23.delete(0, END)
                            e24.delete(0, END)
                            e25.delete(0, END)
                            topAdd.destroy()
                        submitbutton = Button(
                            topAdd, text='Enregistrer', command=lambda: adddData(my_tree))
                        submitbutton.configure(
                            font=('Times', 11, 'bold'), bg='green', fg='white')
                        submitbutton.place(x=300, y=380)
                        cancelbutton = Button(
                            topAdd, text='Annulé', command=lambda: hantaTchouf())
                        cancelbutton.configure(
                            font=('Times', 11, 'bold'), bg='red', fg='white')
                        cancelbutton.place(x=200, y=380)
                        topAdd.protocol(
                            "WM_DELETE_WINDOW", hantaTchouf)
                        topAdd.bind(
                            "<Return>", lambda e: adddData(my_tree))
                        topAdd.bind(
                            "<Escape>", lambda e: hantaTchouf())
                    else:
                        return

                    return

                def stflhzak(tree):
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    lololo = tree.selection()
                    for i in lololo:
                        compteD = 0 
                        compteS = 0
                        compteC=0
                        compteP=0
                        for lpo in range(len(Dattta)):
                            try:
                                if Dattta[lpo][7][0] == "D":
                                    compteD = compteD + 1 
                                    Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                        Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                if Dattta[lpo][7][0] == "S":
                                    compteS = compteS + 1
                                    Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                        Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                if Dattta[lpo][7][0] == "P":
                                    compteP = compteP + 1
                                    Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                        Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                if Dattta[lpo][7][0] == "C":
                                    compteC = compteC + 1
                                    Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                        Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                            except:
                                donothing = 0
                    for parent in tree.get_children():
                        tree.delete(parent)
                    if changerowcolo == 1:
                        changerowcolo = 0
                        changerowcolor()
                    else:
                        changerowcolo = 1
                        changerowcolor()

                def deleteData(tree):
                    global ope
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach

                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            sur = messagebox.askquestion(
                                'les données seront perdues !', "êtes-vous sûr de vouloir supprimer la (les) ligne(s) séléctionnée(s)?", icon='warning')
                            if sur == 'yes':
                                lololo = tree.selection()
                                for i in lololo:
                                    j = 0
                                    l9it = 0
                                    for y in Dattta:
                                        if l9it == 1:
                                            break
                                        if l9it == 0:
                                            if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                                tree.delete(tree.selection()[
                                                            tree.selection().index(i)])
                                                Dattta.remove(Dattta[j])
                                                l9it = 1
                                        j += 1
                            
                                for i in lololo:
                                    compteD = 0 
                                    compteS = 0
                                    compteC=0
                                    compteP=0
                                    for lpo in range(len(Dattta)):
                                        try:
                                            if Dattta[lpo][7][0] == "D":
                                                compteD = compteD + 1 
                                                Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                                    Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                            if Dattta[lpo][7][0] == "S":
                                                compteS = compteS + 1
                                                Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                                    Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                            if Dattta[lpo][7][0] == "P":
                                                compteP = compteP + 1
                                                Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                                    Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                            if Dattta[lpo][7][0] == "C":
                                                compteC = compteC + 1
                                                Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                                    Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                                        except:
                                            donothing = 0
                                for parent in tree.get_children():
                                    tree.delete(parent)
                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à supprimer')
                        ope = 0

                def menuRecherch():
                    global rech
                    if rech == 0:
                        rech = 1
                        global deb

                        def check1():
                            global deb
                            deb = 1
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            c1.select()
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='normal')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check2():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            deb = 2
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=True)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='normal')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check3():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=True)
                            checkVar4 = BooleanVar(value=False)
                            deb = 3
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)

                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='normal')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check4():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            deb = 4
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=True)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='normal')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check5():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            deb = 5
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=True)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='normal')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check6():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            deb = 6
                            checkVar6 = BooleanVar(value=True)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='normal')
                            e7.config(state='disabled')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check7():
                            global deb
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            deb = 7
                            checkVar7 = BooleanVar(value=True)
                            checkVar8 = BooleanVar(value=False)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='normal')
                            e8.config(state='disabled')
                            topSearch.destroy()
                            global rech
                            rech = 0
                            menuRecherch()
                            return

                        def check8():
                            global deb
                            global rech
                            deb = 0
                            rech = 0
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=True)
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.config(state='disabled')
                            e2.config(state='disabled')
                            e3.config(state='disabled')
                            e4.config(state='disabled')
                            e5.config(state='disabled')
                            e6.config(state='disabled')
                            e7.config(state='disabled')
                            e8.config(state='normal')
                            topSearch.destroy()
                            menuRecherch()
                            return
                        topSearch = Toplevel()
                        topSearch.title("Recherche")
                        topSearch.geometry("1000x400")
                        icon = PhotoImage(file='logo-light.png')
                        window.tk.call('wm', 'iconphoto', topSearch._w, icon)
                        loubil = Listbox(topSearch, bg='white',
                                         activestyle='dotbox', justify="center")
                        louou = Label(topSearch, text='',  font=('Times', 18))
                        loubil.place(width=550, height=300, x=430, y=30)
                        topSearch.resizable(width=0, height=0)
                        loubil.delete(0, END)
                        c1 = Checkbutton(topSearch, text='   Date                                          ', variable=checkVar1,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check1())
                        c1.place(x=20, y=30)
                        e1 = Entry(
                            topSearch, textvariable=Date, width=25)
                        e1.place(x=200, y=30)
                        e1.config(state='disabled')
                        if deb == 1:
                            louou.configure(text='DATES')
                            louou.place(width=550, height=20, x=430, y=10)
                            e1.config(state='normal')
                            e1.focus()
                            c1.select()
                            topSearch.title("Recherche par date")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][0]
                                loubil.insert(END, x)
                        c2 = Checkbutton(topSearch, text='   Mois                                          ', variable=checkVar2,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check2())
                        c2.place(x=20, y=70)
                        e2 = Entry(
                            topSearch, textvariable=Mois, width=25)
                        e2.place(x=200, y=70)
                        e2.config(state='disabled')
                        if deb == 2:
                            louou.configure(text='MOIS')
                            louou.place(width=550, height=20, x=430, y=10)
                            e2.config(state='normal')
                            e2.focus()
                            c2.select()
                            topSearch.title("Recherche par mois")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][1]
                                loubil.insert(END, x)
                        c3 = Checkbutton(topSearch, text='   Type                                        ', variable=checkVar3,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check3())
                        c3.place(x=20, y=110)
                        e3 = Entry(
                            topSearch, textvariable=Type, width=25)
                        e3.place(x=200, y=110)
                        e3.config(state='disabled')
                        if deb == 3:
                            louou.configure(text='TYPES')
                            louou.place(width=550, height=20, x=430, y=10)
                            e3.config(state='normal')
                            e3.focus()
                            c3.select()
                            topSearch.title("Recherche par type")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][2]
                                loubil.insert(END, x)
                        c4 = Checkbutton(topSearch, text="   Nom donneur d'ordre                  ", variable=checkVar4,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check4())
                        c4.place(x=20, y=150)
                        e4 = Entry(topSearch, textvariable=Nom, width=25)
                        e4.place(x=200, y=150)
                        e4.config(state='disabled')
                        if deb == 4:
                            louou.configure(text='NOMS')
                            louou.place(width=550, height=20, x=430, y=10)
                            e4.config(state='normal')
                            e4.focus()
                            c4.select()
                            topSearch.title("Recherche par nom")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][3]
                                loubil.insert(END, x)
                        c5 = Checkbutton(topSearch, text="   Montant                                  ", variable=checkVar5,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check5())
                        c5.place(x=20, y=190)
                        e5 = Entry(
                            topSearch, textvariable=Montant, width=25)
                        e5.place(x=200, y=190)
                        e5.config(state='disabled')
                        if deb == 5:
                            louou.configure(text='MONTANTS')
                            louou.place(width=550, height=20, x=430, y=10)
                            e5.config(state='normal')
                            e5.focus()
                            c5.select()
                            topSearch.title("Recherche par montant")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][4]
                                loubil.insert(END, x)
                        c6 = Checkbutton(topSearch, text="   Détail                                   ", variable=checkVar6,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check6())
                        c6.place(x=20, y=230)
                        e6 = Entry(
                            topSearch, textvariable=détail, width=25)
                        e6.place(x=200, y=230)
                        e6.config(state='disabled')
                        if deb == 6:
                            louou.configure(text='DETAIL')
                            louou.place(width=550, height=20, x=430, y=10)
                            e6.config(state='normal')
                            e6.focus()
                            c6.select()
                            topSearch.title("Recherche par détail")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][5]
                                loubil.insert(END, x)
                        c7 = Checkbutton(topSearch, text="   Montant en lettre              ", variable=checkVar7,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check7())
                        c7.place(x=20, y=270)
                        e7 = Entry(
                            topSearch, textvariable=Montant_en_lettre, width=25)
                        e7.place(x=200, y=270)
                        e7.config(state='disabled')
                        if deb == 7:
                            louou.configure(text='MONTANTS EN LETTRE')
                            louou.place(width=550, height=20, x=430, y=10)
                            e7.config(state='normal')
                            e7.focus()
                            c7.select()
                            topSearch.title("Recherche par montent en lettre")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][6]
                                loubil.insert(END, x)
                        c8 = Checkbutton(topSearch, text="   N° Attestation                  ", variable=checkVar8,
                                         font=('Times', 11, 'bold'),
                                         onvalue=1, offvalue=0, command=lambda: check8())
                        c8.place(x=20, y=310)
                        e8 = Entry(
                            topSearch, textvariable=NAttestation, width=25)
                        e8.place(x=200, y=310)
                        e8.config(state='disabled')
                        if deb == 0:
                            louou.configure(text="NUMEROS D'ATTESTATIONS")
                            louou.place(width=550, height=20, x=430, y=10)
                            e8.config(state='normal')
                            e8.focus()
                            c8.select()
                            topSearch.title(
                                "Recherche par numero d'attestation")
                            for parent in my_tree.get_children():
                                x = my_tree.item(parent)["values"][7]
                                loubil.insert(END, x)

                        def stopSearch():
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            global deb
                            deb = 0
                            global rech
                            rech = 0
                            checkVar1 = BooleanVar(value=False)
                            checkVar2 = BooleanVar(value=False)
                            checkVar3 = BooleanVar(value=False)
                            checkVar4 = BooleanVar(value=False)
                            checkVar5 = BooleanVar(value=False)
                            checkVar6 = BooleanVar(value=False)
                            checkVar7 = BooleanVar(value=False)
                            checkVar8 = BooleanVar(value=True)
                            topSearch.destroy()

                        def autofill(e):
                            e1.delete(0, END)
                            e2.delete(0, END)
                            e3.delete(0, END)
                            e4.delete(0, END)
                            e5.delete(0, END)
                            e6.delete(0, END)
                            e7.delete(0, END)
                            e8.delete(0, END)
                            e1.insert(0, loubil.get(loubil.curselection()))
                            e2.insert(0, loubil.get(loubil.curselection()))
                            e3.insert(0, loubil.get(loubil.curselection()))
                            e4.insert(0, loubil.get(loubil.curselection()))
                            e5.insert(0, loubil.get(loubil.curselection()))
                            e6.insert(0, loubil.get(loubil.curselection()))
                            e7.insert(0, loubil.get(loubil.curselection()))
                            e8.insert(0, loubil.get(loubil.curselection()))

                        def tl31(e):
                            global Dattta
                            typed = e1.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[0])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[0].upper():
                                        loubil.insert(END, parent[0])

                        def tl32(e):
                            typed = e2.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[1])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[1].upper():
                                        loubil.insert(END, parent[1])

                        def tl33(e):
                            typed = e3.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[2])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[2].upper():
                                        loubil.insert(END, parent[2])

                        def tl34(e):
                            typed = e4.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[3])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[3].upper():
                                        loubil.insert(END, parent[3])

                        def tl35(e):
                            typed = e5.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[4])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[4].upper():
                                        loubil.insert(END, parent[4])

                        def tl36(e):
                            typed = e6.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[5])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[5].upper():
                                        loubil.insert(END, parent[5])

                        def tl37(e):
                            typed = e7.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[6])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[6].upper():
                                        loubil.insert(END, parent[6])

                        def tl38(e):
                            typed = e8.get()
                            if typed == '':
                                for parent in Dattta:
                                    loubil.insert(END, parent[7])
                            else:
                                loubil.delete(0, END)
                                for parent in Dattta:
                                    if typed.upper() in parent[7].upper():
                                        loubil.insert(END, parent[7])

                        loubil.bind("<<ListboxSelect>>", autofill)
                        e1.bind("<KeyRelease>", tl31)
                        e2.bind("<KeyRelease>", tl32)
                        e3.bind("<KeyRelease>", tl33)
                        e4.bind("<KeyRelease>", tl34)
                        e5.bind("<KeyRelease>", tl35)
                        e6.bind("<KeyRelease>", tl36)
                        e7.bind("<KeyRelease>", tl37)
                        e8.bind("<KeyRelease>", tl38)

                        def searchData(tree):
                            global rech
                            global deb
                            global Dattta
                            global ka
                            global reloulo
                            global reloulou
                            global bach
                            global kn9lb3la
                            global rj3Data
                            global changerowcolo
                            global checkVar1, checkVar2, checkVar3, checkVar4, checkVar5, checkVar6, checkVar7, checkVar8
                            movedown.config(state='disabled')
                            moveup.config(state='disabled')

                            def rj3Data(tree):
                                global showpdf
                                global Dattta
                                global ka
                                global reloulo
                                global kn9lb3la
                                global bach
                                global changerowcolo
                                global reloulou
                                global Organ
                                movedown.config(state='normal')
                                moveup.config(state='normal')
                                ka = 0
                                reloulo = 0
                                reloulou.place_forget()
                                for parent in tree.get_children():
                                    tree.delete(parent)
                                for i in Dattta:
                                    mytag = 'normal'
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    mytag = 'normal'

                                if showpdf == 1:
                                    showPDF()
                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                                if Organ == 1 : 
                                    Organ = 0
                                elif Organ == 2 : 
                                    Organ = 1
                                elif Organ == 3 : 
                                    Organ = 2
                                elif Organ == 0 : 
                                    Organ = 3
                                organiser(my_tree)
                            if deb == 1:
                                if showpdf == 1:
                                    showPDF()
                                ha3lachkt9lbe = e1.get().upper()
                                kn9lb3la = ha3lachkt9lbe
                                e1.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[0]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'date'
                                if ka == 1:
                                    stopSearch()

                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1

                                else:
                                    rech = 1
                                    c1.select()
                                    e1.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 2:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e2.get().upper()
                                e2.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[1]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'mois'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c2.select()
                                    e2.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 3:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e3.get().upper()
                                e3.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[2]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'type'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c3.select()
                                    e3.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 4:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e4.get().upper()
                                e4.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[3]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'nom'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c4.select()
                                    e4.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 5:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e5.get().upper()
                                e5.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[4]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'montant'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c5.select()
                                    e5.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 6:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e6.get().upper()
                                e6.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[5]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'détail'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c6.select()
                                    e6.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 7:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e7.get().upper()
                                e7.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[6]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'mol'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c7.select()
                                    e7.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            if deb == 0:
                                if showpdf == 1:
                                    showPDF()
                                kn9lb3la = e8.get().upper()
                                e8.delete(0, END)
                                deb = 0
                                rech = 0
                                checkVar1 = BooleanVar(value=False)
                                checkVar8 = BooleanVar(value=True)
                                for parent in tree.get_children():
                                    tree.delete(parent)

                                for i in Dattta:
                                    if kn9lb3la in i[7]:
                                        mytag = 'normal'
                                        if i[7][0] == 'P':
                                            mytag = 'orange'
                                        if i[7][0] == 'C':
                                            mytag = 'brown'
                                        if i[7][0] == 'D':
                                            mytag = 'normal'
                                        if i[7][0] == 'S':
                                            mytag = 'yellow'
                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                        ka = 1
                                        bach = 'att'
                                if ka == 1:
                                    stopSearch()
                                    if reloulo == 0:
                                        reloulou = Button(
                                            frame1, text='<-', command=lambda: rj3Data(my_tree))
                                        reloulou.place(x=0, y=137)
                                        reloulo = 1
                                else:
                                    rech = 1
                                    c8.select()
                                    e8.focus()
                                    messagebox.showinfo(
                                        title=kn9lb3la, message="Pas de données !", parent=topSearch)

                                if changerowcolo == 1:
                                    changerowcolo = 0
                                    changerowcolor()
                                else:
                                    changerowcolo = 1
                                    changerowcolor()
                            return
                        ch = Button(
                            topSearch, text='Chercher', command=lambda: searchData(my_tree))
                        ch.configure(
                            font=('Times', 11, 'bold'), bg='green', fg='white')
                        ch.place(x=300, y=350)
                        anul = Button(
                            topSearch, text='Annulé')
                        anul.configure(
                            font=('Times', 11, 'bold'), bg='red', fg='white')
                        anul.place(x=200, y=350)
                        topSearch.protocol("WM_DELETE_WINDOW", stopSearch)
                        topSearch.bind(
                            "<Return>", lambda e: searchData(my_tree))
                        topSearch.bind(
                            "<Escape>", lambda e: stopSearch())
                    return

                def changerowcolor():
                    global mytag
                    global bach
                    global reloulo
                    global changerowcolo
                    global Organ
                    if changerowcolo == 0:
                        for parent in my_tree.get_children():
                            my_tree.delete(parent)
                        if reloulo == 0:
                            for i in Dattta:
                                if mytag == 'normal':
                                    mytag = 'gray'
                                else:
                                    mytag = 'normal'
                                if i[3] == '' or i[3] == ' ':
                                    mytag = 'red'
                                if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                    mytag = 'pink'
                                my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                    i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                        else:
                            try:
                                if bach == 'date':
                                    for i in Dattta:
                                        if kn9lb3la in i[0]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'mois':
                                    for i in Dattta:
                                        if kn9lb3la in i[1]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'type':
                                    for i in Dattta:
                                        if kn9lb3la in i[2]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'nom':
                                    for i in Dattta:
                                        if kn9lb3la in i[3]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'montant':
                                    for i in Dattta:
                                        if kn9lb3la in i[4]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'détail':
                                    for i in Dattta:
                                        if kn9lb3la in i[5]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'mol':
                                    for i in Dattta:
                                        if kn9lb3la in i[6]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                if bach == 'att':
                                    for i in Dattta:
                                        if kn9lb3la in i[7]:
                                            if mytag == 'normal':
                                                mytag = 'gray'
                                            else:
                                                mytag = 'normal'
                                            if i[3] == '' or i[3] == ' ':
                                                mytag = 'red'
                                            if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                                mytag = 'pink'
                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                            except:
                                donothing = 0
                        changerowcolo = 1

                    else:
                        for parent in my_tree.get_children():
                            my_tree.delete(parent)
                        if reloulo == 0:
                            my_tree.tag_configure('yellow', background = 'yellow')
                            my_tree.tag_configure('orange', background = 'orange')
                            my_tree.tag_configure('brown', background='brown')
                            my_tree.tag_configure('purple', background='pink')
                            for i in Dattta:
                                mytag = 'normal'
                                if i[7][0] == 'P':
                                    mytag = 'orange'
                                if i[7][0] == 'C':
                                    mytag = 'brown'
                                if i[7][0] == 'D':
                                    mytag = 'normal'
                                if i[7][0] == 'S':
                                    mytag = 'yellow'
                                my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                    i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                        else:
                            if Organ == 1 : 
                                Organ = 0
                            elif Organ == 2 : 
                                Organ = 1
                            elif Organ == 3 : 
                                Organ = 2
                            elif Organ == 0 : 
                                Organ = 3
                            organiser(my_tree)
                        changerowcolo = 0
                    return

                def selectibabahom():
                    global mselec
                    if mselec == 0:
                        if children:
                            my_tree.selection_set(my_tree.get_children()[0])
                            my_tree.focus_set()
                            my_tree.focus(my_tree.get_children()[0])
                            my_tree.selection_add(my_tree.get_children())
                            mselec = 1
                    else:
                        for item in my_tree.selection():
                            my_tree.selection_remove(item)
                        mselec = 0

                def up():
                    global changerowcolo
                    global ope
                    global Dattta
                    global Organ 
                    Organ = 5

                    if ope == 0:
                        ope = 1
                        Dattta = []
                        x = []
                        rows = my_tree.selection()
                        if len(my_tree.selection()) < 1:
                            messagebox.showinfo(
                                title='Erreur !!', message='Merci de selectionner une ou plusieurs lignes. ')
                        else:
                            if changerowcolo == 0:
                                for row in rows:
                                    my_tree.move(row, my_tree.parent(
                                        row), my_tree.index(row)-1)
                                    x.append(my_tree.index(row)-1)
                                for parent in my_tree.get_children():
                                    Dattta.append(my_tree.item(parent)["values"])
                                compteD = 0 
                                compteS = 0
                                compteC=0
                                compteP=0
                                for lpo in range(len(Dattta)):
                                    try:
                                        if Dattta[lpo][7][0] == "D":
                                            compteD = compteD + 1 
                                            Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                        if Dattta[lpo][7][0] == "S":
                                            compteS = compteS + 1 
                                            Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                        if Dattta[lpo][7][0] == "P":
                                            compteP = compteP + 1 
                                            Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                        if Dattta[lpo][7][0] == "C":
                                            compteC = compteC + 1 
                                            Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                                    except:
                                        donothing = 0
                                for parent in my_tree.get_children():
                                    my_tree.delete(parent)
                                for i in Dattta:
                                    mytag = 'normal'
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    mytag = 'normal'
                                for i in x:
                                    if children:
                                        my_tree.selection_add(
                                            my_tree.get_children()[i+1])
                                        my_tree.focus_set()
                                        my_tree.focus(my_tree.get_children()[i+1])
                            else:
                                for row in rows:
                                    my_tree.move(row, my_tree.parent(
                                        row), my_tree.index(row)-1)
                                    x.append(my_tree.index(row)-1)
                                for parent in my_tree.get_children():
                                    Dattta.append(my_tree.item(parent)["values"])
                                compteD = 0 
                                compteS = 0
                                compteC=0
                                compteP=0
                                for lpo in range(len(Dattta)):
                                    try:
                                        if Dattta[lpo][7][0] == "D":
                                            compteD = compteD + 1 
                                            Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                        if Dattta[lpo][7][0] == "S":
                                            compteS = compteS + 1 
                                            Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                        if Dattta[lpo][7][0] == "P":
                                            compteP = compteP + 1 
                                            Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                        if Dattta[lpo][7][0] == "C":
                                            compteC = compteC + 1 
                                            Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                                    except:
                                        donothing = 0
                                for parent in my_tree.get_children():
                                    my_tree.delete(parent)
                                mytag = 'normal'
                                for i in Dattta:
                                    if mytag == 'normal':
                                        mytag = 'gray'
                                    else:
                                        mytag = 'normal'
                                    if  i[3] == ' ' or i[3] == '':
                                        mytag = 'red'
                                    if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                        mytag = 'pink'
                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in x:
                                    if children:
                                        my_tree.selection_add(
                                            my_tree.get_children()[i+1])
                                        my_tree.focus_set()
                                        my_tree.focus(my_tree.get_children()[i+1])
                        ope = 0
                    return

                def down():
                    global changerowcolo
                    global ope
                    global Dattta
                    global Organ 
                    Organ = 5
                    if ope == 0:
                        ope = 1
                        Dattta = []
                        x = []
                        rows = my_tree.selection()
                        if len(my_tree.selection()) < 1:
                            messagebox.showinfo(
                                title='Erreur !!', message='Merci de selectionner une ou plusieurs lignes. ')
                        else:
                            if changerowcolo == 0:
                                for row in reversed(rows):
                                    my_tree.move(row, my_tree.parent(
                                        row), my_tree.index(row)+1)
                                    x.append(my_tree.index(row)+1)
                                for parent in my_tree.get_children():
                                    Dattta.append(my_tree.item(parent)["values"])
                                compteD = 0 
                                compteS = 0
                                compteC=0
                                compteP=0
                                for lpo in range(len(Dattta)):
                                    try:
                                        if Dattta[lpo][7][0] == "D":
                                            compteD = compteD + 1 
                                            Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                        if Dattta[lpo][7][0] == "S":
                                            compteS = compteS + 1 
                                            Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                        if Dattta[lpo][7][0] == "P":
                                            compteP = compteP + 1 
                                            Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                        if Dattta[lpo][7][0] == "C":
                                            compteC = compteC + 1 
                                            Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                                    except:
                                        donothing = 0
                                for parent in my_tree.get_children():
                                    my_tree.delete(parent)
                                for i in Dattta:
                                    mytag = 'normal'
                                    if i[7][0] == 'P':
                                        mytag = 'orange'
                                    if i[7][0] == 'C':
                                        mytag = 'brown'
                                    if i[7][0] == 'D':
                                        mytag = 'normal'
                                    if i[7][0] == 'S':
                                        mytag = 'yellow'
                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                    mytag = 'normal'
                                for i in x:
                                    if children:
                                        my_tree.selection_add(
                                            my_tree.get_children()[i-1])
                                        my_tree.focus_set()
                                        my_tree.focus(my_tree.get_children()[i-1])
                            else : 
                                for row in reversed(rows):
                                    my_tree.move(row, my_tree.parent(
                                        row), my_tree.index(row)+1)
                                    x.append(my_tree.index(row)+1)
                                for parent in my_tree.get_children():
                                    Dattta.append(my_tree.item(parent)["values"])
                                for parent in my_tree.get_children():
                                    my_tree.delete(parent)
                                compteD = 0 
                                compteS = 0
                                compteC=0
                                compteP=0
                                for lpo in range(len(Dattta)):
                                    try:
                                        if Dattta[lpo][7][0] == "D":
                                            compteD = compteD + 1 
                                            Dattta[lpo][7] = 'D-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteD)
                                        if Dattta[lpo][7][0] == "S":
                                            compteS = compteS + 1 
                                            Dattta[lpo][7] = 'S-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteS)
                                        if Dattta[lpo][7][0] == "P":
                                            compteP = compteP + 1 
                                            Dattta[lpo][7] = 'P-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteP)
                                        if Dattta[lpo][7][0] == "C":
                                            compteC = compteC + 1 
                                            Dattta[lpo][7] = 'C-'+Dattta[lpo][0][6:len(
                                                Dattta[lpo][0])]+'-' + Dattta[lpo][0][3:5]+'-'+str(compteC)
                                    except:
                                        donothing = 0
                                mytag = 'normal'
                                for i in Dattta:
                                    if mytag == 'normal':
                                        mytag = 'gray'
                                    else:
                                        mytag = 'normal'
                                    if  i[3] == ' ' or i[3] == '':
                                        mytag = 'red'
                                    if not re.match(r'^\d{2}/\d{2}/\d{2}$', i[0]) : 
                                        mytag = 'pink'
                                    my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                        i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                for i in x:
                                    if children:
                                        my_tree.selection_add(
                                            my_tree.get_children()[i-1])
                                        my_tree.focus_set()
                                        my_tree.focus(my_tree.get_children()[i-1])
                        ope = 0
                    
                    return

                def idin():
                    global ope
                    if ope == 0:
                        ope = 1
                        curItem = my_tree.focus()
                        values = my_tree.item(curItem, "values")
                        if len(my_tree.selection()) < 2:
                            if len(curItem) > 0:
                                if values[7][0] == "D":
                                    Natt = "D-" + \
                                        values[0][6:len(values[0])] + \
                                        '-'+values[0][3:5]+'-'
                                if values[7][0] == "P":
                                    Natt = "P-" + \
                                        values[0][6:len(values[0])] + \
                                        '-'+values[0][3:5]+'-'
                                if values[7][0] == "C":
                                    Natt = "C-" + \
                                        values[0][6:len(values[0])] + \
                                        '-'+values[0][3:5]+'-'
                                if values[7][0] == "S":
                                    Natt = "S-" + \
                                        values[0][6:len(values[0])] + \
                                        '-'+values[0][3:5]+'-'
                                TopIdin = Toplevel()
                                TopIdin.title("Modification de l'ID")
                                TopIdin.geometry("300x150")
                                icon = PhotoImage(file='logo-light.png')
                                window.tk.call(
                                    'wm', 'iconphoto', TopIdin._w, icon)
                                TopIdin.resizable(width=0, height=0)
                                l60 = Label(TopIdin, text="Merci de saisir le nouveau numéro \nd'attestation pour ce paiement.",
                                            font=('Times', 11, 'bold'))
                                l61 = Label(TopIdin, text=Natt,
                                            font=('Times', 11, 'bold'))
                                e60 = Entry(
                                    TopIdin, textvariable=Date, width=5)
                                l60.pack(pady=10)
                                e60.pack()
                                l61.place(x=64, y=58)
                                e60.focus()

                                def hantaTchouf():
                                    global ope
                                    ope = 0
                                    e60.delete(0, END)
                                    TopIdin.destroy()

                                def changeID():
                                    compteX = 0
                                    global changerowcolo
                                    global ope
                                    global Organ 
                                    Organ = 5
                                    j = 0
                                    e600 = e60.get()
                                    try:
                                        for i in range(len(Dattta)) :
                                            if Dattta[i][7][0] == values[7][0] : 
                                                compteX=compteX+1
                                        
                                        if compteX >0:
                                            if int(e600) <= compteX:
                                                try:
                                                    for li in range(len(Dattta)):
                                                        if j == 1:
                                                            break
                                                        if Dattta[li][7][8: len(Dattta[li][7])] == e60.get() and Dattta[li][7][0]== values[7][0]:
                                                            j += 1
                                                            try:
                                                                x = [Dattta[li][0], Dattta[li][1], Dattta[li][2], Dattta[li][3],
                                                                    Dattta[li][4], Dattta[li][5], Dattta[li][6], Dattta[li][7]]
                                                                del Dattta[li]
                                                                Dattta.insert(
                                                                    li, [values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7][0:8]+e60.get()])
                                                            except:
                                                                donothing = 0
                                                    for z in range(len(Dattta)):
                                                        if values[7] == Dattta[z][7]:
                                                            del Dattta[z]
                                                            Dattta.insert(
                                                                z, [x[0], x[1], x[2], x[3], x[4], x[5], x[6], x[7][0:8]+values[7][8:len(values[7])]])
                                                            break

                                                except:
                                                    donothing = 0
                                                if j != 0:
                                                    for parent in my_tree.get_children():
                                                        my_tree.delete(parent)
                                                    if reloulo == 0:
                                                        for i in Dattta:
                                                            mytag = 'normal'
                                                            if i[7][0] == 'P':
                                                                mytag = 'orange'
                                                            if i[7][0] == 'C':
                                                                mytag = 'brown'
                                                            if i[7][0] == 'D':
                                                                mytag = 'normal'
                                                            if i[7][0] == 'S':
                                                                mytag = 'yellow'
                                                            my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                            mytag = 'normal'
                                                        for u in my_tree.get_children():
                                                            if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                my_tree.focus(u)
                                                                my_tree.selection_set(
                                                                    u)
                                                        hantaTchouf()
                                                        
                                                    else:
                                                        for parent in my_tree.get_children():
                                                            my_tree.delete(parent)
                                                        try:
                                                            if bach == 'date':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[0]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'mois':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[1]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'type':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[2]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'nom':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[3]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'montant':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[4]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'détail':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[5]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'mol':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[6]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            if bach == 'att':
                                                                for i in Dattta:
                                                                    if kn9lb3la in i[7]:
                                                                        mytag = 'normal'
                                                                        if i[7][0] == 'P':
                                                                            mytag = 'orange'
                                                                        if i[7][0] == 'C':
                                                                            mytag = 'brown'
                                                                        if i[7][0] == 'D':
                                                                            mytag = 'normal'
                                                                        if i[7][0] == 'S':
                                                                            mytag = 'yellow'
                                                                        my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                                                                            i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]), tags=(mytag))
                                                                for u in my_tree.get_children():
                                                                    if my_tree.item(u)["values"][7][8:len(my_tree.item(u)["values"][7])] == e60.get():
                                                                        my_tree.focus(
                                                                            u)
                                                                        my_tree.selection_set(
                                                                            u)
                                                                hantaTchouf()
                                                            
                                                        except:
                                                            donothing = 0
                                                stflhzak(my_tree)
                                            else : 
                                                messagebox.showinfo(
                                                    title='Invalide !!', message='Pour les paiements de type "' + values[7][0] +'" merci de saisir un numéro entre 1'+' et '+ str(compteX)+'.', parent=TopIdin)
                                                ope = 1
                                                e60.delete(0, END)
                                        else:
                                            messagebox.showinfo(
                                                title='Invalide !!', message='Pour les paiements de type "' + values[7][0] +'" merci de saisir un numéro entre 1'+' et '+ str(compteX)+'.', parent=TopIdin)
                                            ope = 1
                                            e60.delete(0, END)
                                    except:
                                        messagebox.showinfo(
                                            title='Invalide !!', message='Pour les paiements de type "' + values[7][0] +'" merci de saisir un numéro entre 1'+' et '+ str(compteX)+'.', parent=TopIdin)
                                        ope = 1
                                        e60.delete(0, END)
                                    
                                    if changerowcolo == 1:
                                        changerowcolo = 0
                                        changerowcolor()
                                    else:
                                        changerowcolo = 1
                                        changerowcolor()

                                def hantaTchouf():
                                    global ope
                                    ope = 0
                                    e60.delete(0, END)
                                    TopIdin.destroy()
                                submitbutton = Button(
                                    TopIdin, text='Enregistrer', command=lambda: changeID())
                                submitbutton.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                submitbutton.place(x=150, y=100)
                                cancelbutton = Button(
                                    TopIdin, text='Annulé', command=lambda: hantaTchouf())
                                cancelbutton.configure(
                                    font=('Times', 11, 'bold'), bg='red', fg='white')
                                cancelbutton.place(x=60, y=100)
                                TopIdin.protocol(
                                    "WM_DELETE_WINDOW", hantaTchouf)
                                TopIdin.bind(
                                    "<Escape>", lambda e: hantaTchouf())
                                TopIdin.bind(
                                    "<Return>", lambda e: changeID())
                            else:
                                messagebox.showinfo(
                                    title='!!', message='Merci de selectionner une ligne')
                                ope = 0
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner une seule ligne')
                            ope = 0
                    return

                def telechDossCom(tree):
                    my_tree.selection_set(my_tree.get_children()[0])
                    my_tree.focus_set()
                    my_tree.focus(my_tree.get_children()[0])
                    my_tree.selection_add(my_tree.get_children())
                    telechCSV(tree)
                    att(tree)
                    return

                def telechCSV(tree):
                    global laDate
                    global Dattta
                    global année
                    global télo
                    x = laDate
                    laDate = str(laDate[6:len(laDate)]) + \
                        '-' + str(laDate[3:5])

                    # Créer le nom du dossier basé sur le nom du fichier Excel
                    nom_dossier = os.path.splitext(str(laDate) + ".xlsx")[0]

                    # Créer le dossier s'il n'existe pas déjà
                    if not os.path.exists(nom_dossier):
                        os.makedirs(nom_dossier)

                    try:
                        wb = openpyxl.load_workbook(os.path.join(
                            nom_dossier, str(laDate) + ".xlsx"), read_only=True)
                        if laDate in wb.sheetnames:
                            télo = 1
                    except:
                        télo = 0

                    excel = xlsxwriter.Workbook(os.path.join(
                        nom_dossier, str(laDate) + ".xlsx"))

                    fiche = excel.add_worksheet(laDate)
                    fiche.set_row(0, 30)
                    fiche.autofilter('A1:H11')
                    format1 = excel.add_format({'align': 'center', 'valign': 'vcenter'})
                    format1.set_bg_color('#00B0F0')
                    format1.set_border()
                    format1.set_border_color('#000000')
                    format1.set_bold()
                    format1.set_center_across()
                    format1.set_shrink()
                    format1.set_font_color('#44546A')
                    format1.set_font_size(10)

                    format2 = excel.add_format()
                    format2.set_bg_color('#FFFFFF')
                    format2.set_border()
                    format2.set_border_color('#000000')
                    format2.set_center_across()
                    format2.set_shrink()
                    format2.set_font_size(9)

                    format3 = excel.add_format()
                    format3.set_bg_color('#FFFFFF')
                    format3.set_border()
                    format3.set_border_color('#000000')
                    format3.set_align('right')
                    format3.set_shrink()
                    format3.set_font_size(9)

                    fiche.write(0, 0, 'Dates', format1)
                    fiche.write(0, 1, 'Mois', format1)
                    fiche.write(0, 2, 'Type', format1)
                    fiche.write(0, 3, "NOM Donneur d'ordre", format1)
                    fiche.write(0, 4, 'Montant', format1)
                    fiche.write(0, 5, 'Détail', format1)
                    fiche.write(0, 6, 'Montant en lettre', format1)
                    fiche.write(0, 7, 'N° Attestation', format1)
                    if len(Dattta) > 1:
                        for i in range(len(Dattta)):
                            try:
                                if Dattta[i][7][8:len(Dattta[i][7])] == '1':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'01'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '2':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'02'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '3':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'03'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '4':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'04'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '5':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'05'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '6':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'06'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '7':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'07'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '8':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'08'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '9':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'09'
                                if Dattta[i][3] == '':
                                    Dattta[i][3] = ' '
                                if Dattta[i][2] == '':
                                    Dattta[i][2] = ' '
                                if Dattta[i][1] == '':
                                    Dattta[i][1] = ' '
                                if Dattta[i][0] == '':
                                    Dattta[i][0] = ' '
                                if Dattta[i][4] == '':
                                    Dattta[i][4] = ' '
                                if Dattta[i][5] == '':
                                    Dattta[i][5] = ' '
                                if Dattta[i][6] == '':
                                    Dattta[i][6] = ' '
                                if Dattta[i][7] == '':
                                    Dattta[i][7] = ' '
                                Dattta[i][0] = Dattta[i][0][0:6] + \
                                    "20"+str(année)[2:len(str(année))]
                            except:
                                donothing = 0
                            fiche.write(i+1, 0, Dattta[i][0], format2)
                            fiche.write(i+1, 1, Dattta[i][1], format2)
                            fiche.write(i+1, 2, Dattta[i][2], format2)
                            fiche.write(i+1, 3, Dattta[i][3], format2)
                            fiche.write(i+1, 4, Dattta[i][4], format3)
                            fiche.write(i+1, 5, Dattta[i][5], format2)
                            fiche.write(i+1, 6, Dattta[i][6], format2)
                            fiche.write(i+1, 7, Dattta[i][7], format2)
                            fiche.set_column(0, 0, 13)
                            fiche.set_column(1, 1, 13)
                            fiche.set_column(2, 2, 22)
                            fiche.set_column(3, 3, 32)
                            fiche.set_column(4, 4, 12)
                            fiche.set_column(5, 5, 12)
                            fiche.set_column(6, 6, 60)
                            fiche.set_column(7, 7, 15)
                            fiche.set_row(i, 12)
                        excel.close()
                        for i in Dattta:
                            i[0] = i[0][0:5]+'/'+i[0][8:len(i[0])]
                        for i in range(len(Dattta)):
                            try:
                                if Dattta[i][3] == ' ':
                                    Dattta[i][3] = ''
                                if Dattta[i][2] == ' ':
                                    Dattta[i][2] = ''
                                if Dattta[i][1] == ' ':
                                    Dattta[i][1] = ''
                                if Dattta[i][0] == ' ':
                                    Dattta[i][0] = ''
                                if Dattta[i][4] == ' ':
                                    Dattta[i][4] = ''
                                if Dattta[i][5] == ' ':
                                    Dattta[i][5] = ''
                                if Dattta[i][6] == ' ':
                                    Dattta[i][6] = ''
                                if Dattta[i][7] == ' ':
                                    Dattta[i][7] = ''
                                if Dattta[i][7][8:len(Dattta[i][7])] == '01':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'1'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '02':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'2'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '03':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'3'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '04':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'4'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '05':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'5'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '06':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'6'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '07':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'7'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '08':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'8'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '09':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'9'
                            except:
                                donothing = 0
                        if télo == 1:
                            messagebox.showinfo(
                                title='fichier XLSX enregistré', message='Fichier Excel "'+str(laDate)+'.xlsx" mis à jour.')
                        else:
                            messagebox.showinfo(
                                title='fichier XLSX téléchargé', message='Fichier Excel "'+str(laDate)+'.xlsx" créé.')
                        laDate = x
                    else:
                        messagebox.showinfo(
                            title='!!', message='NO DATA !!')
                    return

                def att(tree):

                    citrop = 0
                    global laDate
                    x = laDate
                    laDate = str(laDate[6:len(laDate)]) + \
                        '-' + str(laDate[3:5])

                    # Créer le nom du dossier basé sur le nom du fichier Excel
                    nom_dossier = os.path.splitext(str(laDate) + ".xlsx")[0]

                    # Créer le dossier s'il n'existe pas déjà
                    if not os.path.exists(nom_dossier):
                        os.makedirs(nom_dossier)

                    if len(tree.selection()) < 1:
                        messagebox.showinfo(
                            title='!!', message='Merci de selectionner une ou plusieurs lignes')
                    elif len(tree.selection()) == len(Dattta):
                        for i in range(len(Dattta)):
                            try:
                                if Dattta[i][7][8:len(Dattta[i][7])] == '1':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'01'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '2':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'02'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '3':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'03'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '4':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'04'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '5':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'05'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '6':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'06'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '7':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'07'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '8':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'08'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '9':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'09'
                                if Dattta[i][3] == '':
                                    Dattta[i][3] = ' '
                                if Dattta[i][2] == '':
                                    Dattta[i][2] = ' '
                                if Dattta[i][1] == '':
                                    Dattta[i][1] = ' '
                                if Dattta[i][0] == '':
                                    Dattta[i][0] = ' '
                                if Dattta[i][4] == '':
                                    Dattta[i][4] = ' '
                                if Dattta[i][5] == '':
                                    Dattta[i][5] = ' '
                                if Dattta[i][6] == '':
                                    Dattta[i][6] = ' '
                                if Dattta[i][7] == '':
                                    Dattta[i][7] = ' '

                            except:
                                donotin = 0
                        for i in range(len(Dattta)):
                            D_22_09_10 = Dattta[i][7]
                            Fait_le = Dattta[i][0]
                            Montant = Dattta[i][4]
                            Montant_en_lettre = Dattta[i][6]
                            Nom = Dattta[i][3]
                            Type = Dattta[i][2]
                            mois = ''
                            if Fait_le[3:5] == '01':
                                mois = 'Janvier'
                            if Fait_le[3:5] == '02':
                                mois = 'Février'
                            if Fait_le[3:5] == '03':
                                mois = 'Mars'
                            if Fait_le[3:5] == '04':
                                mois = 'Avril'
                            if Fait_le[3:5] == '05':
                                mois = 'Mai'
                            if Fait_le[3:5] == '06':
                                mois = 'Juin'
                            if Fait_le[3:5] == '07':
                                mois = 'Juillet'
                            if Fait_le[3:5] == '08':
                                mois = 'Août'
                            if Fait_le[3:5] == '09':
                                mois = 'Septembre'
                            if Fait_le[3:5] == '10':
                                mois = 'Octobre'
                            if Fait_le[3:5] == '11':
                                mois = 'Novembre'
                            if Fait_le[3:5] == '12':
                                mois = 'Décembre'
                            pdf = FPDF(orientation='P', format='A4')
                            pdf.add_page()

                            if D_22_09_10[0] == "D":
                                pdf.set_xy(90, 65)
                                pdf.set_font("times", size=21, style='BU')
                                pdf.cell(txt='Attestation de Don',
                                         w=30, align='C')
                                pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                pdf.set_xy(91.5, 75)
                                pdf.cell(txt=D_22_09_10, w=28, align='C')
                                text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                                text2 = Montant + \
                                    " dirhams ("+Montant_en_lettre+" dirhams)"
                                if Type.lower() == "espèce" or Type.lower() == "espece":
                                    text3 = " en "
                                else:
                                    text3 = " par "
                                text4 = Type+" "
                                text5 = "de "
                                text6 = Nom+"."
                                text7 = "La contribution de "
                                text8 = Nom+" "
                                text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                                text10 = Nom+" "
                                text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                                text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                text14 = "Fait à Casablanca, le "
                                text15 = Fait_le[0:2]+" "+mois + \
                                    " "+"20"+Fait_le[6:len(Fait_le)]

                                text16 = "Bouchra OUTAGHANI"

                                pdf.set_auto_page_break("ON", margin=0.0)
                                pdf.set_font("times", size=12)
                                pdf.set_xy(20, 105)
                                pdf.multi_cell(w=170, h=5, txt=text1+"**"+text2+"**"+text3+"**"+text4+"**"+text5+"**"+text6+"**"+"\n\n"+text7+"**"+text8+"**"+text9+"\n\n"+"**"+text10+"**"+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                               align='L')

                                pdf.set_font("times", size=11)
                                pdf.set_xy(100, 200)
                                pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                               "\n\n"+"**"+text16+"**", markdown=True, align='R')
                                pdf.set_xy(100, 215)
                                pdf.multi_cell(
                                    w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                                pdf.set_font("times", size=9)
                                pdf.set_xy(100, 220)
                                pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                               markdown=True, align='R')
                                pdf.set_xy(100, 225)
                                pdf.multi_cell(
                                    w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                pdf.set_xy(100, 230)
                                pdf.multi_cell(
                                    w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                pdf.set_fill_color(193, 153, 9)
                                pdf.set_xy(8, 275)
                                pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                pdf.set_text_color(45, 82, 158)
                                pdf.set_font("times", size=14, style="B")
                                pdf.set_xy(8, 280)
                                pdf.multi_cell(
                                    w=0, h=5, txt="JADARA Foundation")

                                pdf.set_text_color(193, 153, 9)
                                pdf.set_font("times", size=7.5, style="")
                                pdf.set_xy(8, 285)
                                pdf.multi_cell(
                                    w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                pdf.set_text_color(0, 0, 0)
                                pdf.set_font("times", size=7.5, style="")
                                pdf.set_xy(8, 289)
                                pdf.multi_cell(
                                    w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                pdf.set_font("times", size=8, style="")
                                pdf.set_xy(107, 279)
                                pdf.multi_cell(
                                    w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                pdf.set_font("times", size=8, style="")
                                pdf.set_xy(158, 283)
                                pdf.multi_cell(
                                    w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                pdf.set_xy(152, 275)
                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                pdf.set_xy(102, 275)
                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                nom_fichier_pdf = os.path.join(
                                    nom_dossier, str(D_22_09_10) + ".pdf")

                                pdf.output(nom_fichier_pdf)
                            if D_22_09_10[0] == "S":
                                def swbatts() : 
                                    pdf.set_xy(90, 65)
                                    pdf.set_font("times", size=21, style='BU')
                                    pdf.cell(txt='Attestation de Don',
                                            w=30, align='C')
                                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                    pdf.set_xy(91.5, 75)
                                    pdf.cell(txt=D_22_09_10, w=28, align='C')
                                    nameevent = e808.get()
                                    text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                                    text2 = "**"+Montant + \
                                        " dirhams ("+Montant_en_lettre+" dirhams)"+"**"
                                    if Type.lower() == "espèce" or Type.lower() == "espece":
                                        text3 = " en "
                                    else:
                                        text3 = " par "
                                    text4 = "**"+Type+" "+"**"
                                    text5 = "de"
                                    text6 = "**"+" "+Nom+"**"+"."
                                    text7 = "La contribution de "
                                    text8 = "**"+Nom+" "+"**"
                                    text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                                    textX = "Cette contribution participera au financement de l'évènement :"
                                    textY = "**"+" "+nameevent+"**"
                                    text10 = "**"+" "+Nom+" "+"**"
                                    text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                                    text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                    text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                    text14 = "Fait à Casablanca, le "
                                    text15 = Fait_le[0:2]+" "+mois + \
                                        " "+"20"+Fait_le[6:len(Fait_le)]

                                    text16 = "Bouchra OUTAGHANI"

                                    pdf.set_auto_page_break("ON", margin=0.0)
                                    pdf.set_font("times", size=12)
                                    pdf.set_xy(20, 105)
                                    pdf.multi_cell(w=170, h=5, txt=text1+text2+text3+text4+text5+text6+"\n\n"+text7+text8+text9+"\n\n"+textX+textY+"\n\n"+text10+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                                align='L')

                                    pdf.set_font("times", size=11)
                                    pdf.set_xy(100, 200)
                                    pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                                "\n\n"+"**"+text16+"**", markdown=True, align='R')
                                    pdf.set_xy(100, 215)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                                    pdf.set_font("times", size=9)
                                    pdf.set_xy(100, 220)
                                    pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                                markdown=True, align='R')
                                    pdf.set_xy(100, 225)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                    pdf.set_xy(100, 230)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                    pdf.set_fill_color(193, 153, 9)
                                    pdf.set_xy(8, 275)
                                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                    pdf.set_text_color(45, 82, 158)
                                    pdf.set_font("times", size=14, style="B")
                                    pdf.set_xy(8, 280)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="JADARA Foundation")

                                    pdf.set_text_color(193, 153, 9)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 285)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                    pdf.set_text_color(0, 0, 0)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 289)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(107, 279)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(158, 283)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                    pdf.set_xy(152, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                    pdf.set_xy(102, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                    e808.delete(0, END)
                                    topatts.destroy()

                                    nom_fichier_pdf = os.path.join(
                                        nom_dossier, str(D_22_09_10) + ".pdf")

                                    pdf.output(nom_fichier_pdf)
                                topatts = Toplevel()
                                topatts.title(D_22_09_10)
                                topatts.geometry("500x200")
                                topatts.resizable(width=0, height=0)
                                icon = PhotoImage(file='logo-light.png')
                                window.tk.call('wm', 'iconphoto', topatts._w, icon)
                                l800 = Label(topatts, text=Type+" ("+Montant+" DH) de "+Nom,
                                            font=('Times', 11, 'bold'))
                                l800.place(x=50, y=25)
                                l808 = Label(topatts, text="Cette contribution participera au financement de l'évènement :",
                                            font=('Times', 11, 'bold'))
                                e808 = Entry(
                                    topatts,  width=25)
                                l808.place(x=50, y=72)
                                e808.place(x=150, y=95)
                                def hantaTchoufatts():
                                    e808.delete(0, END)
                                    topatts.destroy()
                                submitbuttonatts = Button(
                                    topatts, text='Enregistrer', command=lambda: swbatts())
                                submitbuttonatts.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                submitbuttonatts.place(x=250, y=150)
                                cancelbuttonatts = Button(
                                    topatts, text='Annulé', command=lambda: hantaTchoufatts())
                                cancelbuttonatts.configure(
                                    font=('Times', 11, 'bold'), bg='red', fg='white')
                                cancelbuttonatts.place(x=150, y=150)
                                topatts.protocol(
                                    "WM_DELETE_WINDOW", hantaTchoufatts)
                                topatts.bind(
                                    "<Return>", lambda e: swbatts())
                                topatts.bind(
                                    "<Escape>", lambda e: hantaTchoufatts())
                                topatts.wait_window()
                            if D_22_09_10[0] == "P":
                                def swbattp():
                                    pdf.set_xy(90, 65)
                                    pdf.set_font("times", size=21, style='BU')
                                    pdf.cell(txt='Attestation de Don en nature',
                                            w=30, align='C')
                                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                    pdf.set_xy(91.5, 75)
                                    pdf.cell(txt=D_22_09_10+"-"+e802.get()+"DN", w=30, align='C')
                                    text10 = "Je soussignée, Mme Bochra CHABBOUBA ELIDRISSI, Responsable Administrative et Financière de JADARA FOUBDATION, atteste par la présente que l'association a bénéficié au titre de l'année scolaire"
                                    text11 = "**"+" "+e801.get()+"**"
                                    text12 = " d'un don en nature de la part de :"
                                    text13 = "**"+" "+e808.get()+"**"
                                    text14 = "Ce don est sous forme d'une place pédagogique gracieusement offerte au profit du boursier inscrit régulièrement au titre de l'année universitaire"
                                    text15 = "**"+"Nom : "+e803.get()+"**"
                                    text16 = "**"+"CIN : "+e804.get()+"**"
                                    text17 = "**"+"Etudiant en : "+e805.get()+"**"
                                    text18 = "Ce don est valorisé dans les livres comptables de notre association au titre de l'exercice"+"**"+" "+e802.get()+"**"
                                    text19 ="**"+ Montant + \
                                        " dirhams ("+Montant_en_lettre+" dirhams)."+"**"
                                    text199="Cette contribution participe au financement de la mission de JADARA FOUNDATION dont l'objet est de financer des bourses d'études supérieures pour les bacheliers méritants issus de milieux défavorisés."
                                    text20 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                    text21 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                    text22 = "Fait à Casablanca, le "
                                    text23 = Fait_le[0:2]+" "+mois + \
                                        " "+"20"+Fait_le[6:len(Fait_le)]


                                    pdf.set_auto_page_break("ON", margin=0.0)
                                    pdf.set_font("times", size=12)
                                    pdf.set_xy(20, 100)
                                    pdf.multi_cell(w=170, h=5, txt=text10+text11+text12+"\n\n"+"                                               "+text13+"\n\n"+text14+text11+" :"+"\n\n"+"       "+text15+"\n\n"+"       "+text16+"\n\n"+"       "+text17+"\n\n"+text18+" à hauteur de "+text19+"\n\n"+text199+"\n\n"+text20+"\n\n"+text21, markdown=True,
                                                align='L')

                                    pdf.set_font("times", size=11)
                                    pdf.set_xy(100, 240)
                                    pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                                    pdf.set_xy(100, 250)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                    pdf.set_xy(100, 255)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                    pdf.set_fill_color(193, 153, 9)
                                    pdf.set_xy(8, 275)
                                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                    pdf.set_text_color(45, 82, 158)
                                    pdf.set_font("times", size=14, style="B")
                                    pdf.set_xy(8, 280)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="JADARA Foundation")

                                    pdf.set_text_color(193, 153, 9)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 285)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                    pdf.set_text_color(0, 0, 0)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 289)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(107, 279)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(158, 283)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                    pdf.set_xy(152, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                    pdf.set_xy(102, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                    e808.delete(0, END)
                                    e801.delete(0, END)
                                    e802.delete(0, END)
                                    e803.delete(0, END)
                                    e804.delete(0, END)
                                    e805.delete(0, END)
                                    topattp.destroy()

                                    nom_fichier_pdf = os.path.join(
                                        nom_dossier, str(D_22_09_10) + ".pdf")
                                    pdf.output(nom_fichier_pdf)
                                topattp = Toplevel()
                                topattp.title(D_22_09_10)
                                topattp.geometry("500x360")
                                topattp.resizable(width=0, height=0)
                                icon = PhotoImage(file='logo-light.png')
                                window.tk.call('wm', 'iconphoto', topattp._w, icon)
                                l800 = Label(topattp, text=Type+" ("+Montant+" DH) de "+Nom,
                                            font=('Times', 11, 'bold'))
                                l800.place(x=50, y=25)
                                l808 = Label(topattp, text="De la part de : ",
                                            font=('Times', 11, 'bold'))
                                e808 = Entry(
                                    topattp,  width=25)
                                l808.place(x=50, y=72)
                                e808.place(x=200, y=70)
                                l802 = Label(topattp, text="Exercice : ",
                                            font=('Times', 11, 'bold'))
                                e802 = Entry(
                                    topattp,  width=25)
                                l802.place(x=50, y=112)
                                e802.place(x=200, y=110)
                                l801 = Label(topattp, text="Année scolaire : ",
                                            font=('Times', 11, 'bold'))
                                e801 = Entry(
                                    topattp,  width=25)
                                l801.place(x=50, y=152)
                                e801.place(x=200, y=150)
                                l803 = Label(topattp, text="Nom étudiant : ",
                                            font=('Times', 11, 'bold'))
                                e803 = Entry(
                                    topattp,  width=25)
                                l803.place(x=50, y=192)
                                e803.place(x=200, y=190)
                                l804 = Label(topattp, text="CIN étudiant : ",
                                            font=('Times', 11, 'bold'))
                                e804 = Entry(
                                    topattp,  width=25)
                                l804.place(x=50, y=232)
                                e804.place(x=200, y=230)
                                l805 = Label(topattp, text="Etudiant en : ",
                                            font=('Times', 11, 'bold'))
                                e805 = Entry(
                                    topattp,  width=25)
                                l805.place(x=50, y=272)
                                e805.place(x=200, y=270)
                                def hantaTchoufattp():
                                    e808.delete(0, END)
                                    e801.delete(0, END)
                                    e802.delete(0, END)
                                    e803.delete(0, END)
                                    e804.delete(0, END)
                                    e805.delete(0, END)
                                    topattp.destroy()
                                submitbuttonattp = Button(
                                    topattp, text='Enregistrer', command=lambda: swbattp())
                                submitbuttonattp.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                submitbuttonattp.place(x=300, y=320)
                                cancelbuttonattp = Button(
                                    topattp, text='Annulé', command=lambda: hantaTchoufattp())
                                cancelbuttonattp.configure(
                                    font=('Times', 11, 'bold'), bg='red', fg='white')
                                cancelbuttonattp.place(x=200, y=320)
                                topattp.protocol(
                                    "WM_DELETE_WINDOW", hantaTchoufattp)
                                topattp.bind(
                                    "<Return>", lambda e: swbattp())
                                topattp.bind(
                                    "<Escape>", lambda e: hantaTchoufattp())
                                topattp.wait_window()
                            if D_22_09_10[0] == "C":
                                def swbattc() : 
                                    pdf.set_xy(90, 65)
                                    pdf.set_font("times", size=21, style='BU')
                                    pdf.cell(txt='Attestation de Cotisation',
                                            w=30, align='C')
                                    pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                    pdf.set_xy(91.5, 75)
                                    pdf.cell(txt=D_22_09_10, w=28, align='C')
                                    text1 = "Nous, JADARA FOUNDATION, attestons par la présente avoir reçu la somme de "
                                    text2 = "**"+Montant + \
                                        " dirhams ("+Montant_en_lettre+" dirhams) "+"**"
                                    text5 = "de "
                                    text6 ="**"+Nom+" "+"**"
                                    text7 = labelc.cget('text')
                                    text8 = "**"+" "+e908.get()+"**"+"."
                                    text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                    text22 = "Fait à Casablanca, le "
                                    text23 = Fait_le[0:2]+" "+mois + \
                                        " "+"20"+Fait_le[6:len(Fait_le)]

                                    pdf.set_auto_page_break("ON", margin=0.0)
                                    pdf.set_font("times", size=13)
                                    pdf.set_xy(20, 120)
                                    pdf.multi_cell(w=170, h=5, txt=text1+text2+text5+text6+text7+text8+"\n\n"+text13, markdown=True,
                                                align='L')
                                    pdf.set_font("times", size=11)
                                    pdf.set_xy(100, 200)
                                    pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                                    pdf.set_xy(100, 210)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                    pdf.set_xy(100, 215)
                                    pdf.multi_cell(
                                        w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                    pdf.set_fill_color(193, 153, 9)
                                    pdf.set_xy(8, 275)
                                    pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                    pdf.set_text_color(45, 82, 158)
                                    pdf.set_font("times", size=14, style="B")
                                    pdf.set_xy(8, 280)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="JADARA Foundation")

                                    pdf.set_text_color(193, 153, 9)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 285)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                    pdf.set_text_color(0, 0, 0)
                                    pdf.set_font("times", size=7.5, style="")
                                    pdf.set_xy(8, 289)
                                    pdf.multi_cell(
                                        w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(107, 279)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                    pdf.set_font("times", size=8, style="")
                                    pdf.set_xy(158, 283)
                                    pdf.multi_cell(
                                        w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                    pdf.set_xy(152, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                    pdf.set_xy(102, 275)
                                    pdf.multi_cell(w=0.5, h=20, txt="", fill=True)
                                    
                                    e908.delete(0, END)
                                    topattc.destroy()
                                    nom_fichier_pdf = os.path.join(
                                        nom_dossier, str(D_22_09_10) + ".pdf")

                                    pdf.output(nom_fichier_pdf)

                                topattc = Toplevel()
                                topattc.title(D_22_09_10)
                                topattc.geometry("500x220")
                                topattc.resizable(width=0, height=0)
                                icon = PhotoImage(file='logo-light.png')
                                window.tk.call('wm', 'iconphoto', topattc._w, icon)
                                l900 = Label(topattc, text=Type+" ("+Montant+" DH) de "+Nom,
                                            font=('Times', 11, 'bold'))
                                l900.place(x=50, y=25)
                                def update_label():
                                    if var.get() == 1:
                                        labelc.config(text="en tant que membre de l'association au titre de l'année")
                                    else:
                                        labelc.config(text="en tant que membre de l'association au titre des années :")

                                var = IntVar()
                                var.set(1)  # Coche le premier checkbutton par défaut

                                checkbuttonc1 = Checkbutton(topattc, text="Une année", variable=var, onvalue=1, offvalue=0, command=update_label)
                                checkbuttonc1.place(x=90, y=60)

                                checkbuttonc2 = Checkbutton(topattc, text="Plusieurs années", variable=var, onvalue=0, offvalue=1, command=update_label)
                                checkbuttonc2.place(x=240, y=60)

                                labelc = Label(topattc, text="en tant que membre de l'association au titre de l'année")
                                labelc.place(x=50, y=90)

                                e908 = Entry(topattc, width=25)
                                e908.pack()
                                e908.place(x=150, y=120)
                                def hantaTchoufattc():
                                    e908.delete(0, END)
                                    topattc.destroy()
                                submitbuttonattc = Button(
                                    topattc, text='Enregistrer', command=lambda: swbattc())
                                submitbuttonattc.configure(
                                    font=('Times', 11, 'bold'), bg='green', fg='white')
                                submitbuttonattc.place(x=250, y=170)
                                cancelbuttonattc = Button(
                                    topattc, text='Annulé', command=lambda: hantaTchoufattc())
                                cancelbuttonattc.configure(
                                    font=('Times', 11, 'bold'), bg='red', fg='white')
                                cancelbuttonattc.place(x=150, y=170)
                                topattc.protocol(
                                    "WM_DELETE_WINDOW", hantaTchoufattc)
                                topattc.bind(
                                    "<Return>", lambda e: swbattc())
                                topattc.bind(
                                    "<Escape>", lambda e: hantaTchoufattc())
                                topattc.wait_window()
                            laDate = x
                            citrop = 1
                        for i in range(len(Dattta)):
                            try:
                                if Dattta[i][3] == ' ':
                                    Dattta[i][3] = ''
                                if Dattta[i][2] == ' ':
                                    Dattta[i][2] = ''
                                if Dattta[i][1] == ' ':
                                    Dattta[i][1] = ''
                                if Dattta[i][0] == ' ':
                                    Dattta[i][0] = ''
                                if Dattta[i][4] == ' ':
                                    Dattta[i][4] = ''
                                if Dattta[i][5] == ' ':
                                    Dattta[i][5] = ''
                                if Dattta[i][6] == ' ':
                                    Dattta[i][6] = ''
                                if Dattta[i][7] == ' ':
                                    Dattta[i][7] = ''
                                if Dattta[i][7][8:len(Dattta[i][7])] == '01':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'1'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '02':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'2'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '03':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'3'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '04':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'4'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '05':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'5'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '06':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'6'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '07':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'7'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '08':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'8'
                                if Dattta[i][7][8:len(Dattta[i][7])] == '09':
                                    Dattta[i][7] = Dattta[i][7][0:8]+'9'
                            except:
                                donothing = 0
                    else:

                        lololo = tree.selection()

                        for i in lololo:
                            j = 0
                            l9it = 0
                            for y in Dattta:
                                if l9it == 1:
                                    break
                                if l9it == 0:
                                    if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                        try:
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '1':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'01'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '2':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'02'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '3':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'03'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '4':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'04'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '5':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'05'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '6':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'06'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '7':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'07'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '8':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'08'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '9':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'09'
                                            if Dattta[j][3] == '':
                                                Dattta[j][3] = ' '
                                            if Dattta[j][2] == '':
                                                Dattta[j][2] = ' '
                                            if Dattta[j][1] == '':
                                                Dattta[j][1] = ' '
                                            if Dattta[j][0] == '':
                                                Dattta[j][0] = ' '
                                            if Dattta[j][4] == '':
                                                Dattta[j][4] = ' '
                                            if Dattta[j][5] == '':
                                                Dattta[j][5] = ' '
                                            if Dattta[j][6] == '':
                                                Dattta[j][6] = ' '
                                            if Dattta[j][7] == '':
                                                Dattta[j][7] = ' '
                                        except:
                                            donotin = 0
                                        D_22_09_10 = Dattta[j][7]
                                        Fait_le = Dattta[j][0]
                                        Montant = Dattta[j][4]
                                        Montant_en_lettre = Dattta[j][6]
                                        Nom = Dattta[j][3]
                                        if Nom == "":
                                            Nom = " "
                                        Type = Dattta[j][2]
                                        mois = ''
                                        if Fait_le[3:5] == '01':
                                            mois = 'Janvier'
                                        if Fait_le[3:5] == '02':
                                            mois = 'Février'
                                        if Fait_le[3:5] == '03':
                                            mois = 'Mars'
                                        if Fait_le[3:5] == '04':
                                            mois = 'Avril'
                                        if Fait_le[3:5] == '05':
                                            mois = 'Mai'
                                        if Fait_le[3:5] == '06':
                                            mois = 'Juin'
                                        if Fait_le[3:5] == '07':
                                            mois = 'Juillet'
                                        if Fait_le[3:5] == '08':
                                            mois = 'Août'
                                        if Fait_le[3:5] == '09':
                                            mois = 'Septembre'
                                        if Fait_le[3:5] == '10':
                                            mois = 'Octobre'
                                        if Fait_le[3:5] == '11':
                                            mois = 'Novembre'
                                        if Fait_le[3:5] == '12':
                                            mois = 'Décembre'
                                        pdf = FPDF(
                                            orientation='P', format='A4')
                                        pdf.add_page()
                                        if D_22_09_10[0] == "D":
                                            pdf.set_xy(90, 65)
                                            pdf.set_font("times", size=21, style='BU')
                                            pdf.cell(txt='Attestation de Don',
                                                    w=30, align='C')
                                            pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                            pdf.set_xy(91.5, 75)
                                            pdf.cell(txt=D_22_09_10, w=28, align='C')
                                            text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                                            text2 = Montant + \
                                                " dirhams ("+Montant_en_lettre+" dirhams)"
                                            if Type.lower() == "espèce" or Type.lower() == "espece":
                                                text3 = " en "
                                            else:
                                                text3 = " par "
                                            text4 = Type+" "
                                            text5 = "de "
                                            text6 = Nom+"."
                                            text7 = "La contribution de "
                                            text8 = Nom+" "
                                            text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                                            text10 = Nom+" "
                                            text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                                            text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                            text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                            text14 = "Fait à Casablanca, le "
                                            text15 = Fait_le[0:2]+" "+mois + \
                                                " "+"20"+Fait_le[6:len(Fait_le)]

                                            text16 = "Bouchra OUTAGHANI"

                                            pdf.set_auto_page_break("ON", margin=0.0)
                                            pdf.set_font("times", size=12)
                                            pdf.set_xy(20, 105)
                                            pdf.multi_cell(w=170, h=5, txt=text1+"**"+text2+"**"+text3+"**"+text4+"**"+text5+"**"+text6+"**"+"\n\n"+text7+"**"+text8+"**"+text9+"\n\n"+"**"+text10+"**"+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                                        align='L')

                                            pdf.set_font("times", size=11)
                                            pdf.set_xy(100, 200)
                                            pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                                        "\n\n"+"**"+text16+"**", markdown=True, align='R')
                                            pdf.set_xy(100, 215)
                                            pdf.multi_cell(
                                                w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                                            pdf.set_font("times", size=9)
                                            pdf.set_xy(100, 220)
                                            pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                                        markdown=True, align='R')
                                            pdf.set_xy(100, 225)
                                            pdf.multi_cell(
                                                w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                            pdf.set_xy(100, 230)
                                            pdf.multi_cell(
                                                w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                            pdf.set_fill_color(193, 153, 9)
                                            pdf.set_xy(8, 275)
                                            pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                            pdf.set_text_color(45, 82, 158)
                                            pdf.set_font("times", size=14, style="B")
                                            pdf.set_xy(8, 280)
                                            pdf.multi_cell(
                                                w=0, h=5, txt="JADARA Foundation")

                                            pdf.set_text_color(193, 153, 9)
                                            pdf.set_font("times", size=7.5, style="")
                                            pdf.set_xy(8, 285)
                                            pdf.multi_cell(
                                                w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                            pdf.set_text_color(0, 0, 0)
                                            pdf.set_font("times", size=7.5, style="")
                                            pdf.set_xy(8, 289)
                                            pdf.multi_cell(
                                                w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                            pdf.set_font("times", size=8, style="")
                                            pdf.set_xy(107, 279)
                                            pdf.multi_cell(
                                                w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                            pdf.set_font("times", size=8, style="")
                                            pdf.set_xy(158, 283)
                                            pdf.multi_cell(
                                                w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                            pdf.set_xy(152, 275)
                                            pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                            pdf.set_xy(102, 275)
                                            pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                            nom_fichier_pdf = os.path.join(
                                                nom_dossier, str(D_22_09_10) + ".pdf")

                                            pdf.output(nom_fichier_pdf)
                                        if D_22_09_10[0] == "S":
                                            def swbatts() : 
                                                pdf.set_xy(90, 65)
                                                pdf.set_font("times", size=21, style='BU')
                                                pdf.cell(txt='Attestation de Don',
                                                        w=30, align='C')
                                                pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                                pdf.set_xy(91.5, 75)
                                                pdf.cell(txt=D_22_09_10, w=28, align='C')
                                                nameevent = e808.get()
                                                text1 = "Je soussignée, Mme Bouchra OUTAGHANI, Trésorière Générale de JADARA FOUNDATION, atteste par la présente avoir reçu la somme de "
                                                text2 = "**"+Montant + \
                                                    " dirhams ("+Montant_en_lettre+" dirhams)"+"**"
                                                if Type.lower() == "espèce" or Type.lower() == "espece":
                                                    text3 = " en "
                                                else:
                                                    text3 = " par "
                                                text4 = "**"+Type+" "+"**"
                                                text5 = "de"
                                                text6 = "**"+" "+Nom+"**"+"."
                                                text7 = "La contribution de "
                                                text8 = "**"+Nom+" "+"**"
                                                text9 = "participera au financement de la mission de JADARA Foudation telle que prévue par ses status accessibles sur son site web www.jadara.foundation."
                                                textX = "Cette contribution participera au financement de l'évènement :"
                                                textY = "**"+" "+nameevent+"**"
                                                text10 = "**"+" "+Nom+" "+"**"
                                                text11 = "peut accéder au rapport morale et financier de JADARA Foudation sur son site web www.jadara.foudation."
                                                text12 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                                text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                                text14 = "Fait à Casablanca, le "
                                                text15 = Fait_le[0:2]+" "+mois + \
                                                    " "+"20"+Fait_le[6:len(Fait_le)]

                                                text16 = "Bouchra OUTAGHANI"

                                                pdf.set_auto_page_break("ON", margin=0.0)
                                                pdf.set_font("times", size=12)
                                                pdf.set_xy(20, 105)
                                                pdf.multi_cell(w=170, h=5, txt=text1+text2+text3+text4+text5+text6+"\n\n"+text7+text8+text9+"\n\n"+textX+textY+"\n\n"+text10+text11+"\n\n"+text12+"\n\n"+text13, markdown=True,
                                                            align='L')

                                                pdf.set_font("times", size=11)
                                                pdf.set_xy(100, 200)
                                                pdf.multi_cell(w=90, h=5, txt=text14+"**"+text15+"**" +
                                                            "\n\n"+"**"+text16+"**", markdown=True, align='R')
                                                pdf.set_xy(100, 215)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Trésorière Générale**", markdown=True, align='R')
                                                pdf.set_font("times", size=9)
                                                pdf.set_xy(100, 220)
                                                pdf.multi_cell(w=90, h=5, txt="**P.O**",
                                                            markdown=True, align='R')
                                                pdf.set_xy(100, 225)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                                pdf.set_xy(100, 230)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                                pdf.set_fill_color(193, 153, 9)
                                                pdf.set_xy(8, 275)
                                                pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                                pdf.set_text_color(45, 82, 158)
                                                pdf.set_font("times", size=14, style="B")
                                                pdf.set_xy(8, 280)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="JADARA Foundation")

                                                pdf.set_text_color(193, 153, 9)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 285)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                                pdf.set_text_color(0, 0, 0)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 289)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(107, 279)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(158, 283)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                                pdf.set_xy(152, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                                pdf.set_xy(102, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                                e808.delete(0, END)
                                                topatts.destroy()

                                                nom_fichier_pdf = os.path.join(
                                                    nom_dossier, str(D_22_09_10) + ".pdf")

                                                pdf.output(nom_fichier_pdf)
                                            topatts = Toplevel()
                                            topatts.title(D_22_09_10)
                                            topatts.geometry("500x200")
                                            topatts.resizable(width=0, height=0)
                                            icon = PhotoImage(file='logo-light.png')
                                            window.tk.call('wm', 'iconphoto', topatts._w, icon)
                                            l800 = Label(topatts, text=Type+" ("+Montant+" DH) de "+Nom,
                                                        font=('Times', 11, 'bold'))
                                            l800.place(x=50, y=25)
                                            l808 = Label(topatts, text="Cette contribution participera au financement de l'évènement :",
                                                        font=('Times', 11, 'bold'))
                                            e808 = Entry(
                                                topatts,  width=25)
                                            l808.place(x=50, y=72)
                                            e808.place(x=150, y=95)
                                            def hantaTchoufatts():
                                                e808.delete(0, END)
                                                topatts.destroy()
                                            submitbuttonatts = Button(
                                                topatts, text='Enregistrer', command=lambda: swbatts())
                                            submitbuttonatts.configure(
                                                font=('Times', 11, 'bold'), bg='green', fg='white')
                                            submitbuttonatts.place(x=250, y=150)
                                            cancelbuttonatts = Button(
                                                topatts, text='Annulé', command=lambda: hantaTchoufatts())
                                            cancelbuttonatts.configure(
                                                font=('Times', 11, 'bold'), bg='red', fg='white')
                                            cancelbuttonatts.place(x=150, y=150)
                                            topatts.protocol(
                                                "WM_DELETE_WINDOW", hantaTchoufatts)
                                            topatts.bind(
                                                "<Return>", lambda e: swbatts())
                                            topatts.bind(
                                                "<Escape>", lambda e: hantaTchoufatts())
                                            topatts.wait_window()
                                        if D_22_09_10[0] == "P":
                                            def swbattp():
                                                pdf.set_xy(90, 65)
                                                pdf.set_font("times", size=21, style='BU')
                                                pdf.cell(txt='Attestation de Don en nature',
                                                        w=30, align='C')
                                                pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                                pdf.set_xy(91.5, 75)
                                                pdf.cell(txt=D_22_09_10+"-"+e802.get()+"-DN", w=30, align='C')
                                                text10 = "Je soussignée, Mme Bochra CHABBOUBA ELIDRISSI, Responsable Administrative et Financière de JADARA FOUBDATION, atteste par la présente que l'association a bénéficié au titre de l'année scolaire"
                                                text11 = "**"+" "+e801.get()+"**"
                                                text12 = " d'un don en nature de la part de :"
                                                text13 = "**"+" "+e808.get()+"**"
                                                text14 = "Ce don est sous forme d'une place pédagogique gracieusement offerte au profit du boursier inscrit régulièrement au titre de l'année universitaire"
                                                text15 = "**"+"Nom : "+e803.get()+"**"
                                                text16 = "**"+"CIN : "+e804.get()+"**"
                                                text17 = "**"+"Etudiant en : "+e805.get()+"**"
                                                text18 = "Ce don est valorisé dans les livres comptables de notre association au titre de l'exercice"+"**"+" "+e802.get()+"**"
                                                text19 ="**"+ Montant + \
                                                    " dirhams ("+Montant_en_lettre+" dirhams)."+"**"
                                                text199="Cette contribution participe au financement de la mission de JADARA FOUNDATION dont l'objet est de financer des bourses d'études supérieures pour les bacheliers méritants issus de milieux défavorisés."
                                                text20 = "Jadara Foudation est une association reconnue d'utilité publique par le Décret N°2-13-339 du 29 Avril 2013 tel que modifié par le Décret N°2.22.834 du 19 octobre 2022."
                                                text21 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                                text22 = "Fait à Casablanca, le "
                                                text23 = Fait_le[0:2]+" "+mois + \
                                                    " "+"20"+Fait_le[6:len(Fait_le)]


                                                pdf.set_auto_page_break("ON", margin=0.0)
                                                pdf.set_font("times", size=12)
                                                pdf.set_xy(20, 100)
                                                pdf.multi_cell(w=170, h=5, txt=text10+text11+text12+"\n\n"+"                                               "+text13+"\n\n"+text14+text11+" :"+"\n\n"+"       "+text15+"\n\n"+"       "+text16+"\n\n"+"       "+text17+"\n\n"+text18+" à hauteur de "+text19+"\n\n"+text199+"\n\n"+text20+"\n\n"+text21, markdown=True,
                                                            align='L')

                                                pdf.set_font("times", size=11)
                                                pdf.set_xy(100, 240)
                                                pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                                                pdf.set_xy(100, 250)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                                pdf.set_xy(100, 255)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                                pdf.set_fill_color(193, 153, 9)
                                                pdf.set_xy(8, 275)
                                                pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                                pdf.set_text_color(45, 82, 158)
                                                pdf.set_font("times", size=14, style="B")
                                                pdf.set_xy(8, 280)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="JADARA Foundation")

                                                pdf.set_text_color(193, 153, 9)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 285)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                                pdf.set_text_color(0, 0, 0)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 289)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(107, 279)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(158, 283)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                                pdf.set_xy(152, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                                pdf.set_xy(102, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                                e808.delete(0, END)
                                                e801.delete(0, END)
                                                e802.delete(0, END)
                                                e803.delete(0, END)
                                                e804.delete(0, END)
                                                e805.delete(0, END)
                                                topattp.destroy()

                                                nom_fichier_pdf = os.path.join(
                                                    nom_dossier, str(D_22_09_10) + ".pdf")
                                                pdf.output(nom_fichier_pdf)
                                            topattp = Toplevel()
                                            topattp.title(D_22_09_10)
                                            topattp.geometry("500x360")
                                            topattp.resizable(width=0, height=0)
                                            icon = PhotoImage(file='logo-light.png')
                                            window.tk.call('wm', 'iconphoto', topattp._w, icon)
                                            l800 = Label(topattp, text=Type+" ("+Montant+" DH) de "+Nom,
                                                        font=('Times', 11, 'bold'))
                                            l800.place(x=50, y=25)
                                            l808 = Label(topattp, text="De la part de : ",
                                                        font=('Times', 11, 'bold'))
                                            e808 = Entry(
                                                topattp,  width=25)
                                            l808.place(x=50, y=72)
                                            e808.place(x=200, y=70)
                                            l802 = Label(topattp, text="Exercice : ",
                                                        font=('Times', 11, 'bold'))
                                            e802 = Entry(
                                                topattp,  width=25)
                                            l802.place(x=50, y=112)
                                            e802.place(x=200, y=110)
                                            l801 = Label(topattp, text="Année scolaire : ",
                                                        font=('Times', 11, 'bold'))
                                            e801 = Entry(
                                                topattp,  width=25)
                                            l801.place(x=50, y=152)
                                            e801.place(x=200, y=150)
                                            l803 = Label(topattp, text="Nom étudiant : ",
                                                        font=('Times', 11, 'bold'))
                                            e803 = Entry(
                                                topattp,  width=25)
                                            l803.place(x=50, y=192)
                                            e803.place(x=200, y=190)
                                            l804 = Label(topattp, text="CIN étudiant : ",
                                                        font=('Times', 11, 'bold'))
                                            e804 = Entry(
                                                topattp,  width=25)
                                            l804.place(x=50, y=232)
                                            e804.place(x=200, y=230)
                                            l805 = Label(topattp, text="Etudiant en : ",
                                                        font=('Times', 11, 'bold'))
                                            e805 = Entry(
                                                topattp,  width=25)
                                            l805.place(x=50, y=272)
                                            e805.place(x=200, y=270)
                                            def hantaTchoufattp():
                                                e808.delete(0, END)
                                                e801.delete(0, END)
                                                e802.delete(0, END)
                                                e803.delete(0, END)
                                                e804.delete(0, END)
                                                e805.delete(0, END)
                                                topattp.destroy()
                                            submitbuttonattp = Button(
                                                topattp, text='Enregistrer', command=lambda: swbattp())
                                            submitbuttonattp.configure(
                                                font=('Times', 11, 'bold'), bg='green', fg='white')
                                            submitbuttonattp.place(x=300, y=320)
                                            cancelbuttonattp = Button(
                                                topattp, text='Annulé', command=lambda: hantaTchoufattp())
                                            cancelbuttonattp.configure(
                                                font=('Times', 11, 'bold'), bg='red', fg='white')
                                            cancelbuttonattp.place(x=200, y=320)
                                            topattp.protocol(
                                                "WM_DELETE_WINDOW", hantaTchoufattp)
                                            topattp.bind(
                                                "<Return>", lambda e: swbattp())
                                            topattp.bind(
                                                "<Escape>", lambda e: hantaTchoufattp())
                                            topattp.wait_window()
                                        if D_22_09_10[0] == "C":
                                            def swbattc() : 
                                                pdf.set_xy(90, 65)
                                                pdf.set_font("times", size=21, style='BU')
                                                pdf.cell(txt='Attestation de Cotisation',
                                                        w=30, align='C')
                                                pdf.image('logo-att.png', 80, 10, w=50, h=40)
                                                pdf.set_xy(91.5, 75)
                                                pdf.cell(txt=D_22_09_10, w=28, align='C')
                                                text1 = "Nous, JADARA FOUNDATION, attestons par la présente avoir reçu la somme de "
                                                text2 = "**"+Montant + \
                                                    " dirhams ("+Montant_en_lettre+" dirhams) "+"**"
                                                text5 = "de "
                                                text6 ="**"+Nom+" "+"**"
                                                text7 = labelc.cget('text')
                                                text8 = "**"+" "+e908.get()+"**"+"."
                                                text13 = "Cette attestation est délivrée pour servir et valoir ce que de droit."

                                                text22 = "Fait à Casablanca, le "
                                                text23 = Fait_le[0:2]+" "+mois + \
                                                    " "+"20"+Fait_le[6:len(Fait_le)]

                                                pdf.set_auto_page_break("ON", margin=0.0)
                                                pdf.set_font("times", size=13)
                                                pdf.set_xy(20, 120)
                                                pdf.multi_cell(w=170, h=5, txt=text1+text2+text5+text6+text7+text8+"\n\n"+text13, markdown=True,
                                                            align='L')

                                                pdf.set_font("times", size=11)
                                                pdf.set_xy(100, 200)
                                                pdf.multi_cell(w=90, h=5, txt=text22+"**"+text23+"**", markdown=True, align='R')
                                                pdf.set_xy(100, 210)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Bochra CHABBOUBA ELIDRISSI**", markdown=True, align='R')
                                                pdf.set_xy(100, 215)
                                                pdf.multi_cell(
                                                    w=90, h=5, txt="**Responsable Administrative et Financière**", markdown=True, align='R')

                                                pdf.set_fill_color(193, 153, 9)
                                                pdf.set_xy(8, 275)
                                                pdf.multi_cell(w=0, h=0.5, txt="", fill=True)

                                                pdf.set_text_color(45, 82, 158)
                                                pdf.set_font("times", size=14, style="B")
                                                pdf.set_xy(8, 280)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="JADARA Foundation")

                                                pdf.set_text_color(193, 153, 9)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 285)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="Jadara Foundation, association reconnue d'utilité publique selon de décret N°2.22.834")

                                                pdf.set_text_color(0, 0, 0)
                                                pdf.set_font("times", size=7.5, style="")
                                                pdf.set_xy(8, 289)
                                                pdf.multi_cell(
                                                    w=0, h=5, txt="**Adresse**: 295, Bd Abdelmoumen C24, angle Rue la Percée, 20100, Casablanca", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(107, 279)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**T**: +212 522 861 880\n**F**: +212 522 864 178\n**Mail**: contact@jadara.foundation", markdown=True)

                                                pdf.set_font("times", size=8, style="")
                                                pdf.set_xy(158, 283)
                                                pdf.multi_cell(
                                                    w=40, h=5, txt="**www.jadara.foundation**", markdown=True)

                                                pdf.set_xy(152, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)

                                                pdf.set_xy(102, 275)
                                                pdf.multi_cell(w=0.5, h=20, txt="", fill=True)
                                                
                                                e908.delete(0, END)
                                                topattc.destroy()
                                                nom_fichier_pdf = os.path.join(
                                                    nom_dossier, str(D_22_09_10) + ".pdf")

                                                pdf.output(nom_fichier_pdf)

                                            topattc = Toplevel()
                                            topattc.title(D_22_09_10)
                                            topattc.geometry("500x220")
                                            topattc.resizable(width=0, height=0)
                                            icon = PhotoImage(file='logo-light.png')
                                            window.tk.call('wm', 'iconphoto', topattc._w, icon)
                                            l900 = Label(topattc, text=Type+" ("+Montant+" DH) de "+Nom,
                                                        font=('Times', 11, 'bold'))
                                            l900.place(x=50, y=25)
                                            def update_label():
                                                if var.get() == 1:
                                                    labelc.config(text="en tant que membre de l'association au titre de l'année")
                                                else:
                                                    labelc.config(text="en tant que membre de l'association au titre des années :")

                                            var = IntVar()
                                            var.set(1)  # Coche le premier checkbutton par défaut

                                            checkbuttonc1 = Checkbutton(topattc, text="Une année", variable=var, onvalue=1, offvalue=0, command=update_label)
                                            checkbuttonc1.place(x=90, y=60)

                                            checkbuttonc2 = Checkbutton(topattc, text="Plusieurs années", variable=var, onvalue=0, offvalue=1, command=update_label)
                                            checkbuttonc2.place(x=240, y=60)

                                            labelc = Label(topattc, text="en tant que membre de l'association au titre de l'année")
                                            labelc.place(x=50, y=90)

                                            e908 = Entry(topattc, width=25)
                                            e908.pack()
                                            e908.place(x=150, y=120)
                                            def hantaTchoufattc():
                                                e908.delete(0, END)
                                                topattc.destroy()
                                            submitbuttonattc = Button(
                                                topattc, text='Enregistrer', command=lambda: swbattc())
                                            submitbuttonattc.configure(
                                                font=('Times', 11, 'bold'), bg='green', fg='white')
                                            submitbuttonattc.place(x=250, y=170)
                                            cancelbuttonattc = Button(
                                                topattc, text='Annulé', command=lambda: hantaTchoufattc())
                                            cancelbuttonattc.configure(
                                                font=('Times', 11, 'bold'), bg='red', fg='white')
                                            cancelbuttonattc.place(x=150, y=170)
                                            topattc.protocol(
                                                "WM_DELETE_WINDOW", hantaTchoufattc)
                                            topattc.bind(
                                                "<Return>", lambda e: swbattc())
                                            topattc.bind(
                                                "<Escape>", lambda e: hantaTchoufattc())
                                            topattc.wait_window()
                                        laDate = x
                                        citrop = 1

                                        try:
                                            if Dattta[j][3] == ' ':
                                                Dattta[j][3] = ''
                                            if Dattta[j][2] == ' ':
                                                Dattta[j][2] = ''
                                            if Dattta[j][1] == ' ':
                                                Dattta[j][1] = ''
                                            if Dattta[j][0] == ' ':
                                                Dattta[j][0] = ''
                                            if Dattta[j][4] == ' ':
                                                Dattta[j][4] = ''
                                            if Dattta[j][5] == ' ':
                                                Dattta[j][5] = ''
                                            if Dattta[j][6] == ' ':
                                                Dattta[j][6] = ''
                                            if Dattta[j][7] == ' ':
                                                Dattta[j][7] = ''
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '01':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'1'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '02':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'2'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '03':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'3'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '04':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'4'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '05':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'5'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '06':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'6'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '07':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'7'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '08':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'8'
                                            if Dattta[j][7][8:len(Dattta[j][7])] == '09':
                                                Dattta[j][7] = Dattta[j][7][0:8]+'9'
                                        except:
                                            donothing = 0
                                        l9it = 1
                                j += 1
                    if citrop == 1:
                        messagebox.showinfo(
                            title='', message="Fichier(s) enregistré(s).", parent=window)
                    return
                rowscolor = Button(
                    frame4, text='', background='gray', command=changerowcolor)
                rowscolor.place(x=75, y=5)

                selall = Button(frame5, text='}', command=selectibabahom)
                selall.place(x=2, y=25)
                moveup = Button(frame5, text='⤊', command=up)
                moveup.place(x=2, y=130)
                movedown = Button(frame5, text='⤋', command=down)
                movedown.place(x=2, y=194)
                idinput = Button(frame5, text='•', command=idin)
                idinput.place(x=3, y=162)
                selall1 = Button(frame5, text='}', command=selectibabahom)
                selall1.place(x=2, y=296)

                télécharger = Button(
                    frame1, text='Télécharger xlsx', command=lambda: telechCSV(my_tree))
                télécharger.pack(pady=20)
                lo8 = Label(frame1, text='Ctrl-X', fg='white',
                            background='black', font=('Times', 10))
                lo8.pack()
                télécharger_dossier_complet = Button(
                    frame1, text='Dossier complet', command=lambda: telechDossCom(my_tree))
                télécharger_dossier_complet.pack(pady=40)

                lodo = Label(frame1, text='Ctrl-D', fg='white',
                             background='black', font=('Times', 10))
                lodo.pack()
                rechercher = Button(frame3, text='Rechercher',
                                    command=lambda: menuRecherch())
                delete = Button(frame3, text='Supprimer',
                                command=lambda: deleteData(my_tree))
                add = Button(frame3, text='Ajouter',
                             command=lambda: addData(my_tree))
                edite = Button(frame3, text='Modifier',
                               command=lambda: editData(my_tree))
                AttTypeD = Button(window, text='D',
                                  command=lambda: editTypeAttD(my_tree))
                AttTypeS = Button(window, text='S',background='yellow',
                                  command=lambda: editTypeAttS(my_tree))
                AttTypeP = Button(window, text='P',background='orange',
                                  command=lambda: editTypeAttP(my_tree))
                AttTypeC = Button(window, text='C',background='brown',
                                  command=lambda: editTypeAttC(my_tree))
                organise = Button(window, text='O',background='black',foreground='white',
                                  command=lambda: organiser(my_tree))
                def Mmee(tree) :
                    global ope
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][3].startswith('MLLE '):
                                                Dattta[j][3] = 'MME ' + Dattta[j][3][5:]
                                            elif Dattta[j][3].startswith('MR '):
                                                Dattta[j][3] = 'MME ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('DR '):
                                                Dattta[j][3] = 'MME ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('MME '):
                                                Dattta[j][3] = 'MME ' + Dattta[j][3][4:]
                                            elif Dattta[j][3].startswith('M '):
                                                Dattta[j][3] = 'MME ' + Dattta[j][3][2:]
                                            else :
                                                Dattta[j][3] = 'MME ' + Dattta[j][3]
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0


                def Mllee(tree) :
                    global ope
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][3].startswith('MLLE '):
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3][5:]
                                            elif Dattta[j][3].startswith('MR '):
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('DR '):
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('MME '):
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3][4:]
                                            elif Dattta[j][3].startswith('M '):
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3][2:]
                                            else :
                                                Dattta[j][3] = 'MLLE ' + Dattta[j][3]
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0

                def Mrr(tree) :
                    global ope
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][3].startswith('MLLE '):
                                                Dattta[j][3] = 'MR ' + Dattta[j][3][5:]
                                            elif Dattta[j][3].startswith('MR '):
                                                Dattta[j][3] = 'MR ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('DR '):
                                                Dattta[j][3] = 'MR ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('MME '):
                                                Dattta[j][3] = 'MR ' + Dattta[j][3][4:]
                                            elif Dattta[j][3].startswith('M '):
                                                Dattta[j][3] = 'MR ' + Dattta[j][3][2:]
                                            else :
                                                Dattta[j][3] = 'MR ' + Dattta[j][3]
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0

                def Drr(tree) :
                    global ope
                    global Dattta
                    global changerowcolo
                    global kn9lb3la
                    global reloulo
                    global bach
                    if ope == 0:
                        ope = 1
                        if len(tree.selection()) > 0:
                            lololo = tree.selection()
                            for i in lololo:
                                j = 0
                                l9it = 0
                                for y in Dattta:
                                    if l9it == 1:
                                        break
                                    if l9it == 0:
                                        if tree.selection()[tree.selection().index(i)].split()[0] == y[7] or y[7] == tree.selection()[tree.selection().index(i)].split()[len(tree.selection()[tree.selection().index(i)].split())-1]:
                                            if Dattta[j][3].startswith('MLLE '):
                                                Dattta[j][3] = 'DR ' + Dattta[j][3][5:]
                                            elif Dattta[j][3].startswith('MR '):
                                                Dattta[j][3] = 'DR ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('DR '):
                                                Dattta[j][3] = 'DR ' + Dattta[j][3][3:]
                                            elif Dattta[j][3].startswith('MME '):
                                                Dattta[j][3] = 'DR ' + Dattta[j][3][4:]
                                            elif Dattta[j][3].startswith('M '):
                                                Dattta[j][3] = 'DR ' + Dattta[j][3][2:]
                                            else :
                                                Dattta[j][3] = 'DR ' + Dattta[j][3]
                                            l9it = 1
                                    j += 1
                            stflhzak(tree)
                        else:
                            messagebox.showinfo(
                                title='!!', message='Merci de selectionner la (les) ligne(s) à modifier')
                        ope = 0
                    
                Mme = Button(window, text='Mme',
                                  command=lambda: Mmee(my_tree))
                Mlle = Button(window, text='Mlle',
                                  command=lambda: Mllee(my_tree))
                Mr = Button(window, text='Mr',
                                  command=lambda: Mrr(my_tree))
                Dr = Button(window, text='Dr',
                                  command=lambda: Drr(my_tree))
                Attestation = Button(
                    frame3, text='  Attestation(s)', command=lambda: att(my_tree))

                window.bind('<Control-k>', lambda *args: selectibabahom())
                window.bind('<Control-K>', lambda *args: selectibabahom())
                window.bind('<Control-D>', lambda e: telechDossCom(my_tree))
                window.bind('<Control-d>', lambda e: telechDossCom(my_tree))
                window.bind('<Control-O>', lambda e: organiser(my_tree))
                window.bind('<Control-o>', lambda e: organiser(my_tree))
                window.bind("<Control-s>", lambda e: deleteData(my_tree))
                window.bind("<Control-f>", lambda e: menuRecherch())
                window.bind("<Control-F>", lambda e: menuRecherch())
                window.bind("<Control-a>", lambda e: addData(my_tree))
                window.bind("<Control-P>", lambda e: showPDF())
                window.bind("<Control-p>", lambda e: showPDF())
                window.bind("<Control-Z>", lambda e: rj3Data(my_tree))
                window.bind("<Control-z>", lambda e: rj3Data(my_tree))
                window.bind("<Control-m>", lambda e: editData(my_tree))
                my_tree.bind("<Double-1>", lambda e: editData(my_tree))
                window.bind("<Control-S>", lambda e: deleteData(my_tree))
                window.bind("<Control-A>", lambda e: addData(my_tree))
                window.bind("<Control-M>", lambda e: editData(my_tree))
                window.bind("+", lambda e: up())
                window.bind("=", lambda e: down())
                window.bind("<Control-I>", lambda e: idin())
                window.bind("<Control-i>", lambda e: idin())
                window.bind("<Control-T>", lambda e: att(my_tree))
                window.bind("<Control-t>", lambda e: att(my_tree))
                window.bind("<Control-x>", lambda e: telechCSV(my_tree))
                window.bind("<Control-X>", lambda e: telechCSV(my_tree))
                rechercher.pack(pady=20, padx=20)
                lo1 = Label(frame3, text='Ctrl-F', fg='white',
                            background='black', font=('Times', 10))
                lo1.pack()
                delete.pack(pady=20, padx=20)
                lo2 = Label(frame3, text='Ctrl-S', fg='white',
                            background='black', font=('Times', 10))
                lo2.pack()
                add.pack(pady=20, padx=20)
                lo3 = Label(frame3, text='Ctrl-A', fg='white',
                            background='black', font=('Times', 10))
                lo3.pack()
                edite.pack(pady=20, padx=20)
                lo4 = Label(frame3, text='Ctrl-M', fg='white',
                            background='black', font=('Times', 10))
                lo4.pack()
                AttTypeD.place(x=1525, y=10)
                AttTypeS.place(x=1560, y=10)
                AttTypeP.place(x=1595, y=10)
                AttTypeC.place(x=1630, y=10)
                organise.place(x=1660, y = 50)
                Dr.place(x=545,y=10)
                Mme.place(x=600, y=10)
                Mlle.place(x=670, y=10)
                Mr.place(x=740, y=10)

                Attestation.pack(pady=20, padx=20)
                lo5 = Label(frame3, text='Ctrl-T', fg='white',
                            background='black', font=('Times', 10))
                lo5.pack()
                return
            return

        frame1.pack(fill='y', pady=10, padx=10, side='right')
        frame2.pack(fill='both', side='left', expand=True, padx=10, pady=10)
        init = Button(frame1, text="Relancer l'application", command=reset)
        lo12 = Label(frame1, text='Ctrl-R', fg='white',
                     background='black', font=('Times', 10))
        lo12.pack(side='bottom')
        init.pack(side='bottom', pady=20, padx=5)
        l = Label(frame2, text=os.path.basename(RBUmnia))
        l.pack(pady=5, padx=5)
        window.bind("<Return>", lambda e: traitement())
        window.bind("<Control-r>", lambda e: reset())
        window.bind("<Control-R>", lambda e: reset())

    b = Button(text="Créer attestation", command=lambda: secondaire())
    a = Button(text="Charger relevé bancaire PDF", command=lambda: principal())
    window.bind("<Control-T>", lambda e: secondaire())
    window.bind("<Control-t>", lambda e: secondaire())
    a.place(relx=0.5, rely=0.3, anchor=CENTER)
    lo14 = Label(text='Entrée', fg='white',
                 background='black', font=('Times', 10))
    lo14.place(relx=0.5, rely=0.4, anchor=CENTER)
    b.place(relx=0.5, rely=0.7, anchor=CENTER)
    lo13 = Label(text='Ctrl-T', fg='white',
                 background='black', font=('Times', 10))
    lo13.place(relx=0.5, rely=0.8, anchor=CENTER)
    window.title(" JADARA ")
    window.bind("<Return>", lambda e: principal())


window.mainloop()
